import json, requests, re, boto3, http.client
from .utils.email import enviar_email_empresa
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from .models import Cliente, Pedidos, Produtos, Itens_Pedido, Colaborador, Bairro, TipoPgto, Contas_Receber, Contas_Pagar, Fornecedor, \
    Empresas, Registra, Layout, Producao, Tipo_Pedido, Revenda, Produtos_Revenda, Temp_Pedidos, \
    Temp_Contas_Receber, Bancos, Rota, Reparos, atendimento, Cob_Unificada, Inadimplencia, RegistrosLog, Status_Colaborador, Motivos_Desligamento, TipoMedida
from .models import *
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required

from datetime import datetime, date, timedelta
from django.db.models import Q, F, Sum, OuterRef, Subquery, Max, Case, When, Value, IntegerField, Func, ExpressionWrapper, fields, Count
import locale

from babel.numbers import format_currency
from itertools import chain
from django.utils import timezone

import urllib.parse
from django.views.decorators.csrf import csrf_exempt

from django.db.models.functions import TruncMonth, ExtractMonth, ExtractYear
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from collections import defaultdict

#import pandas as pd
#import plotly.express as px


def novo_log (numero_pedido, colaborador, texto):
    novolog = RegistrosLog()
    #novolog.pedido = Pedidos.objects.filter(numero=numero_pedido).first()
    novolog.colaborador = colaborador
    novolog.operacao = texto
    novolog.data = datetime.today()
    novolog.save()

def obter_vendas_por_mes_ano(empresa_id):
    pedidos=Pedidos.objects.filter(empresa_pedido=empresa_id)
    vendas = (
       pedidos
       .annotate(mes=ExtractMonth('datapedido'), ano=ExtractYear('datapedido'))
       .values('mes', 'ano')
       .annotate(total_vendas=Sum('valorTotal'))
       .order_by('ano', 'mes')

    )
    return vendas


def vendas_por_modelo_mes_ano(empresa_id):
    # Filtrar os pedidos pela empresa
    pedidos = Pedidos.objects.filter(empresa_pedido=empresa_id)

    # Obter todos os modelos de produtos
    modelos = Produtos.objects.values_list('modelo', flat=True)


    # Obter os itens que contêm no nome parte do modelo
    itens_com_modelos = Itens_Pedido.objects.filter(pedido_itens__in=pedidos)

    # Criar um filtro para encontrar os itens cujos nomes contêm algum dos modelos
    itens_com_modelos = itens_com_modelos.filter(
        # Verifica se qualquer parte do nome do modelo está contida no nome do item
        *[Q(nome__icontains=modelo) for modelo in modelos]
    )
    print(itens_com_modelos)

    # Agrupar por modelo e ano, somando as quantidades
    itens_venda = (
        itens_com_modelos
        .annotate(ano=ExtractYear('pedido_itens__datapedido'))
        .values('nome', 'ano')
        .annotate(quantidade_total=Sum('quantidade'))
        .order_by('ano', 'nome')
    )


    return list(itens_venda)

def teste_email (request):

    empresa = 'temtapetes'
    assunto = 'EMAIL DE TESTE SSL'
    mensagem = 'EMAIL DE TESTE SSL'
    destinatarios =['gustavocterra@yahoo.com','gterra@conectivasolucoes.com.br', 'suporte@conectivasolucoes.com.br']

    enviar_email_empresa(empresa, assunto, mensagem, destinatarios)

    return HttpResponse('Teste Efetuado')



def envio_email_cobranca(obj_parcela):


    empresa_obj = obj_parcela.pedido_conta.empresa
    if 'tem' in empresa_obj.nome:
        empresa = 'temtapetes'
    else:
        empresa = 'guermatapetes'

    mensagem = f'Favor verificar sua fatura que venceu no ultimo dia {obj_parcela.data_vencimento} - {obj_parcela.caminho_boleto}'

    destinatarios = [obj_parcela.pedido_conta.comprador_email, obj_parcela.cliente.email]
    assunto = 'Fatura Pendente'
    enviar_email_empresa(empresa, assunto, mensagem, destinatarios)




# def indicadores(request):
#     ###GERANDO DADOS GRF #########
#     vendas_grf = obter_vendas_por_mes_ano(1)
#
#     df = pd.DataFrame(vendas_grf)
#     df['mes_nome'] = pd.to_datetime(df['mes'], format='%m').dt.strftime('%b')  # Nome do mês
#     fig = px.bar(
#         df,
#         x='mes_nome',
#         y='total_vendas',
#         color='ano',
#         barmode='group',  # Agrupa por ano
#         title="Vendas por Mês e Ano GRF",
#         labels={'mes_nome': 'Mês', 'total_vendas': 'Total de Vendas'}
#     )
#     graph_grf = fig.to_html(full_html=False)
#
#     ###GERANDO DADOS TEM #########
#     vendas_tem = obter_vendas_por_mes_ano(2)
#
#     df = pd.DataFrame(vendas_tem)
#     df['mes_nome'] = pd.to_datetime(df['mes'], format='%m').dt.strftime('%b')  # Nome do mês
#     figtem = px.bar(
#         df,
#         x='mes_nome',
#         y='total_vendas',
#         color='ano',
#         barmode='group',  # Agrupa por ano
#         title="Vendas por Mês e Ano TEM",
#         labels={'mes_nome': 'Mês', 'total_vendas': 'Total de Vendas'}
#     )
#     graph_tem = figtem.to_html(full_html=False)
#
#     ####Venda por Itens GRF##############
#     itens_venda = vendas_por_modelo_mes_ano(1)
#     dados_agrupados = {}
#     for item in itens_venda:
#         descricao = item['nome']
#         quantidade = item['quantidade_total']
#         if descricao in dados_agrupados:
#             dados_agrupados[descricao] += quantidade
#         else:
#             dados_agrupados[descricao] = quantidade
#
#     # Criar um DataFrame a partir dos dados agrupados
#     df = pd.DataFrame(itens_venda)  # itens_venda contém 'descricao', 'quantidade_total', 'ano'
#
#     # Agrupar por item e somar as quantidades por ano
#     df_agrupado = (
#         df.groupby(['nome', 'ano'])
#         .agg(quantidade_total=('quantidade_total', 'sum'))
#         .reset_index()
#     )
#
#     # Criar uma coluna de rótulo que une os anos e as quantidades
#     df_agrupado['info_hover'] = (
#             'Ano: ' + df_agrupado['ano'].astype(str) +
#             ' | Quantidade: ' + df_agrupado['quantidade_total'].astype(str)
#     )
#
#     # Somar a quantidade total de cada item para o gráfico
#     df_resumo = (
#         df_agrupado.groupby('nome')
#         .agg(quantidade_total=('quantidade_total', 'sum'))
#         .reset_index()
#     )
#
#     # Gerar o gráfico de pizza
#     fig_grf_itens = px.pie(
#         df_resumo,
#         names='nome',  # Nome do item
#         values='quantidade_total',  # Quantidade total por item
#         title='Distribuição de Vendas por Item e Ano',
#         hover_data={'nome': False, 'quantidade_total': False},
#         custom_data=['nome']
#     )
#
#     # Adicionar detalhes personalizados no hover
#     fig_grf_itens.update_traces(
#         textinfo='none',  # Remove o texto padrão das fatias
#         hovertemplate=(
#                 '<b>%{label}</b><br>' +
#                 'Quantidade Total: %{value}<br>' +
#                 '<extra></extra>'
#         )
#     )
#     graph_grf_itens = fig_grf_itens.to_html(full_html=False)
#
#
#
#     # Criando o gráfico
#
#
#     context = {
#         'graph_grf': graph_grf,
#         'graph_tem': graph_tem,
#         'graph_grf_itens': graph_grf_itens
#
#     }
#
#     return render (request, 'indicadores/index.html', context)
def contagem_numeros(mensagem):
    # Usando expressão regular para encontrar dígitos
    numeros = re.findall(r'\d', mensagem)
    return len(numeros)

def enviar_whatsapp(request):

    url = 'https://api.z-api.io/instances/3C42B3C8B120B003CC20FA99999AD217/token/40B45D278087BB716C9FC13F/send-text'
    payload = {
        "phone": "5521998280320",
        "message": "Welcome to *Z-API*"
    }


    print(payload)

    response = requests.post(url, json=payload)

    print(response.text)

def enviar_whatsapp_cliente(telefone, mensagem):

    url = 'https://api.z-api.io/instances/3C42B3C8B120B003CC20FA99999AD217/token/40B45D278087BB716C9FC13F/send-text'
    print(telefone, mensagem)
    data_atual = date.today()
    atendimento_atual = atendimento.objects.filter(telefone=telefone, data=data_atual).order_by('-id').first()

    if telefone == '5521975733934':

        if mensagem == '1':
            total_vendedor = 0
            data_atual = date.today()
            mes_atual = data_atual.month
            pedidos = Pedidos.objects.filter(datapedido__month=mes_atual, datapedido__year=data_atual.year)
            vendedores = Colaborador.objects.filter(Q(funcao__icontains='Vendedor')|Q(funcao='Gerente'))
            mensagem = ('Olá, Márcia. Segue o Relatorio dos Vendedores: \n'
                        )
            payload = {
                "phone": telefone,
                "message": mensagem

            }
            response = requests.post(url, json=payload)
            for vendedor in vendedores:
                pedidos = Pedidos.objects.filter(datapedido__month=mes_atual, datapedido__year=data_atual.year)
                pedidos_vendedor = pedidos.filter(Q(Vendedor=vendedor.nome)|Q(Vendedor2=vendedor.nome))
                for pedido in pedidos_vendedor:
                    if vendedor.nome in pedido.Vendedor:
                        total_vendedor += pedido.vendedor_comissaov
                    elif vendedor.nome in pedido.Vendedor2:
                        total_vendedor += pedido.vendedor2_comissaov
                mensagem = f"Vendedor: {vendedor.nome}, Total Comissao: {total_vendedor}"
                payload = {
                    "phone": telefone,
                    "message": mensagem

                }
                response = requests.post(url, json=payload)




            mensagem = ('Olá, Márcia. Segue o Relatorio dos Vendedores: \n'
                       )
            payload = {
                "phone": telefone,
                "message": mensagem

            }
            response = requests.post(url, json=payload)



        mensagem = ('Olá, Márcia. Selecione o Relatorio Desejado: \n'
                    '1: Resumo de Comissao Vendedores Mês Atual\n'
                    '2: Resumo de Comissao Vendedores Mês Anterior\n'
                    '3: Resumo Financeiro GRF\n'
                    '4: Resumo Financeiro TEM\n'
                    '5: Resumo Inadimplentes\n'

                    )
        payload = {
            "phone": telefone,
            "message": mensagem

        }
        response = requests.post(url, json=payload)



    if atendimento_atual is None:

        novo_atendimento = atendimento()
        novo_atendimento.telefone = telefone
        novo_atendimento.status='aberto'
        novo_atendimento.data = date.today()
        novo_atendimento.tempo = datetime.now()
        novo_atendimento.save()
        mensagem = ('Olá, me informe o seguinte: Digite:\n'
                    '1: Consultar Pedidos\n'
                    '2: Novo pedido\n'
                    '3: 2a via de boleto')
        payload = {
            "phone": telefone,
            "message": mensagem

        }
        response = requests.post(url, json=payload)
        print(response)


    mensagem_recebida = mensagem
    numero_de_digitos = contagem_numeros(mensagem_recebida)




    if mensagem_recebida == '1' and atendimento_atual.status == 'aberto':

        if atendimento_atual.cliente is None:
            mensagem = 'Digite seu documento. Somente os numeros'
            payload = {
                "phone": telefone,
                "message": mensagem
            }
            atendimento_atual.ultima_opcao = 1
            atendimento_atual.save()

            print(payload)

            response = requests.post(url, json=payload)

            print(response.text)
        else:
            cliente = atendimento_atual.cliente
            pedidos = Pedidos.objects.filter(cliente_pedido=cliente)
            if pedidos:
                payload = {
                    "phone": telefone,
                    "message": 'Aqui estão os pedidos já realizados e seus respectivos status'

                }
                response = requests.post(url, json=payload)
                print('mensagem enviada', response.text)
                for pedido in pedidos:
                    mensagem = f"Número do pedido: {pedido.numero}, Status do pedido: {pedido.status}, Valor Total: {pedido.valorTotal}"
                    payload = {
                        "phone": telefone,
                        "message": mensagem

                    }
                    response = requests.post(url, json=payload)
                    print('mensagem enviada', response.text)
            else:
                mensagem = "Não foi encontrado nenhum pedido para o documento fornecido"
                payload = {
                    "phone": telefone,
                    "message": mensagem

                }
                response = requests.post(url, json=payload)
                print('mensagem enviada', response.text)


    elif mensagem_recebida == '2':
        mensagem = 'Você será atendido por um dos nossos vendedores em instantes'
        payload = {
            "phone": telefone,
            "message": mensagem
        }

        print(payload)

        response = requests.post(url, json=payload)

        print(response.text)

    elif mensagem_recebida == '3' and atendimento_atual.status == 'aberto' :
        if atendimento_atual.cliente is None:
            atendimento_atual.ultima_opcao = 3
            atendimento_atual.save()
            mensagem = 'Digite seu documento. Somente os numeros'
            payload = {
                "phone": telefone,
                "message": mensagem
            }


            print(payload)

            response = requests.post(url, json=payload)

            print(response.text)
        else:
            cliente = atendimento_atual.cliente

            # Suponha que você tem um modelo chamado `Contas_Receber`
            contas = Contas_Receber.objects.filter(cliente=cliente, id_cobranca__isnull=False)
            contas = contas.filter((Q(status_cobranca='PENDING') | Q(status_cobranca='OVERDUE')))
            if contas:
                mensagem = 'Seus boletos são os seguintes:'
                payload = {
                    "phone": telefone,
                    "message": mensagem
                }

                print(payload)

                response = requests.post(url, json=payload)

                for conta in contas:

                    mensagem = f"Vencimento: {conta.data_vencimento}, Valor do Boleto: {conta.valor}, Numero Parcela / Total Parcela: {conta.numero_parcela}/{conta.total_parcelas}\n{conta.caminho_boleto}"

                    payload = {
                        "phone": telefone,
                        "message": mensagem
                    }

                    print(payload)

                    response = requests.post(url, json=payload)

                    print(response.text)
            else:
                mensagem = 'Nenhum boleto em aberto encontrado'
                payload = {
                    "phone": telefone,
                    "message": mensagem
                }

    elif atendimento_atual.ultima_opcao == 1 and numero_de_digitos > 10:
        documento = mensagem_recebida


        cliente = Cliente.objects.filter(documento=documento).first()

        if cliente:

            atendimento_atual.cliente = cliente
            atendimento_atual.save()

            pedidos = Pedidos.objects.filter(cliente_pedido=cliente)
            if pedidos:
                payload = {
                    "phone": telefone,
                    "message": 'Aqui estão os pedidos já realizados e seus respectivos status'

                }
                response = requests.post(url, json=payload)
                print('mensagem enviada', response.text)
                for pedido in pedidos:
                    mensagem = f"Número do pedido: {pedido.numero}, Status do pedido: {pedido.status}, Valor Total: {pedido.valorTotal}"
                    payload = {
                        "phone": telefone,
                        "message": mensagem

                    }
                    response = requests.post(url, json=payload)
                    print('mensagem enviada', response.text)
            else:
                mensagem = "Não foi encontrado nenhum pedido para o documento fornecido"
                payload = {
                    "phone": telefone,
                    "message": mensagem

                }
                response = requests.post(url, json=payload)
                print('mensagem enviada', response.text)
        else:
            mensagem = "Cliente não encontrado"
            payload = {
                "phone": telefone,
                "message": mensagem

            }
            response = requests.post(url, json=payload)
            print('mensagem enviada', response.text)

    elif atendimento_atual.ultima_opcao == 3 and numero_de_digitos > 10:
        documento = mensagem_recebida

        cliente = Cliente.objects.filter(documento=documento).first()

        if cliente:
            atendimento_atual.cliente = cliente
            atendimento_atual.save()
            contas = Contas_Receber.objects.filter(cliente=cliente, id_cobranca__isnull=False)
            contas = contas.filter((Q(status_cobranca='PENDING') | Q(status_cobranca='OVERDUE')))
            if contas:

                for conta in contas:
                    mensagem = conta.caminho_boleto
                    payload = {
                        "phone": telefone,
                        "message": mensagem
                    }

                    print(payload)

                    response = requests.post(url, json=payload)

                    print(response.text)
            else:
                mensagem = 'Nenhum boleto em aberto encontrado'
                payload = {
                    "phone": telefone,
                    "message": mensagem
                }



    else:
        mensagem = ('Olá, me informe o seguinte: Digite:\n'
                    '1: Consultar Pedidos\n'
                    '2: Novo pedido\n'
                    '3: 2a via de boleto')
        payload = {
            "phone": telefone,
            "message": mensagem

        }
        response = requests.post(url, json=payload)
        print('mensagem enviada', response.text)




def check_data(dt_abertura):

    if dt_abertura:
        if isinstance(dt_abertura, str):
            # Assuming dt_abertura is a string in the format "YYYY-MM-DD"
            try:
                # Convert the input string to a datetime.date object
                dt_abertura = datetime.strptime(dt_abertura, "%Y-%m-%d").date()
            except ValueError:
                # If dt_abertura is not in the correct format, return False
                return False

        if isinstance(dt_abertura, date):
            # Get the current date as a datetime.date object
            current_date = date.today()
            # Calculate the difference between the current date and dt_abertura
            time_difference = current_date - dt_abertura
            # Get the number of days in the time difference
            days_difference = time_difference.days

            # Check if the difference is at least 365 days (considering leap years)
            return days_difference >= 365

        else:
            return False
    else:
        return False

def formatar_documento(cpf_cnpj):
    # Remove caracteres não numéricos
    cpf_cnpj = re.sub('[^0-9]', '', cpf_cnpj)

    if len(cpf_cnpj) == 11:
        # Formatação para CPF: 999.999.999-99
        cpf_cnpj = f'{cpf_cnpj[:3]}.{cpf_cnpj[3:6]}.{cpf_cnpj[6:9]}-{cpf_cnpj[9:]}'
    elif len(cpf_cnpj) == 14:
        # Formatação para CNPJ: 99.999.999/9999-99
        cpf_cnpj = f'{cpf_cnpj[:2]}.{cpf_cnpj[2:5]}.{cpf_cnpj[5:8]}/{cpf_cnpj[8:12]}-{cpf_cnpj[12:]}'
    return cpf_cnpj

def consultar_cep(request):
    cep = request.GET.get('cep')
    url = f'https://viacep.com.br/ws/{cep}/json/'
    response = requests.get(url)

    if response.status_code == 200:
        data = response.json()
        return JsonResponse(data)
    else:
        return JsonResponse({'error': 'CEP not found'})


def consultar_cnpj(request):

    cnpj = request.GET.get('cnpj')  # Obtém o número do CNPJ do formulário

    # Realiza a consulta à API de CNPJ
    url = f'https://receitaws.com.br/v1/cnpj/{cnpj}'
    headers = {
        'Authorization': 'Bearer seu_token_de_autenticacao',
        'Content-Type': 'application/json'
    }
    response = requests.get(url, headers=headers)

    if response.status_code == 200:  # Verifica se a requisição foi bem-sucedida
        data = response.json()  # Obtém os dados da empresa da resposta da API

        return JsonResponse(data)
    else:
        # Trata o caso de falha na consulta (ex: CNPJ inválido)
        return JsonResponse({'error': 'CNPJ not found'})

def compara_pedidos (numero_pedido):
    pedidoN = Pedidos.objects.filter(numero=numero_pedido).first()
    data_atual = date.today()
    data_inicial = data_atual - timedelta(10)

    pedidos_lista = Pedidos.objects.filter(datapedido__range=[data_inicial, data_atual])
    for pedido in pedidos_lista:
        print (pedido)
        if pedido.numero!=pedidoN.numero and pedidoN.cliente_pedido == pedido.cliente_pedido and pedidoN.valorTotal == pedido.valorTotal:
            itens_n = Itens_Pedido.objects.filter(pedido=numero_pedido)
            itens_p = Itens_Pedido.objects.filter(pedido=pedido.numero)
            for item_n, item_p in zip(itens_n, itens_p):
                if (
                        item_n.nome == item_p.nome and
                        item_n.preco == item_p.preco and
                        item_n.comprimento == item_p.comprimento and
                        item_n.largura == item_p.largura
                ):
                    pedidoN.status = 'duplicado'
                    pedidoN.save()
                    print('pedido duplicado!')

        else:
         pass

@csrf_exempt
def flag_impressao(request):
    data = json.loads(request.body)
    pedido_numero = data.get('numero_pedido')
    print(pedido_numero)
    pedido = Pedidos.objects.filter (numero=pedido_numero).first()
    if pedido and pedido.flag_impressao == 1:
        response_data = 0
    elif pedido:
        pedido.flag_impressao = 1
        pedido.save()
        print('flag_pedido_impressao = 1', pedido_numero)
        response_data = 1
    else:
        response_data = {'error': 'Pedido não encontrado'}


    return JsonResponse(response_data, safe=False)



def ajustar_itens():
    itens = Itens_Pedido.objects.all()

    for item in itens:
        item.pedido_itens = Pedidos.objects.filter(numero=item.pedido).first()
        item.save()

def painel (request):



    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    if colaborador.empresa.flag_pagamento == 1:
        return redirect('/dologin/')
    elif not colaborador:
        return redirect('/dologin/')
    else:
        pass
    pedidos = Pedidos.objects.all()

    for pedido in pedidos:
        if pedido.vendedor_comissaov == 0 or pedido.vendedor_comissaov is None:
            pedido.vendedor2_comissaov =0.00
            pedido.vendedor_comissaov =0.00

            pedido.vendedor_comissaov = ((pedido.valorTotalParcial-pedido.descontototal)/100) * pedido.vendedor_comissao
            pedido.vendedor_comissaov = round(pedido.vendedor_comissaov,2)
            if pedido.Vendedor2:
                pedido.vendedor2_comissaov= ((pedido.valorTotalParcial-pedido.descontototal)/100) * pedido.vendedor2_comissao
                pedido.vendedor2_comissaov = round(pedido.vendedor2_comissaov,2)
            else:
                pass
            pedido.save()
        else:
            pass

    if 'Vendedor' in colaborador.funcao:

        return redirect('meus_pedidos')

    if 'Gerente' in colaborador.funcao:

        return redirect('dashboard')

    if 'Produção' in colaborador.funcao:

        return redirect('producao')

    if 'Logistica' in colaborador.funcao:

        return redirect ('entregas')

    if 'Expedicao' in colaborador.funcao:
        return redirect('expedicao')



    return render(request, 'painel.html')

@csrf_exempt
def dologin(request):
    data = {}
    if request.method == 'POST':
        username = request.POST.get('username', '')  # Obtém o valor da chave 'username' ou retorna uma string vazia
        password = request.POST.get('pass', '')
        user = authenticate(username=username, password=password)
        colaborador = Colaborador.objects.filter(usuario__username=username).first()
        print(colaborador)

        if user is not None and colaborador.empresa.flag_pagamento == 0:
            login(request, user)
            return redirect('/painel/')
        elif colaborador.empresa.flag_pagamento == 1:
            data['msg'] = ' Sistema Bloqueado! Favor Contactar o Suporte!\nEmail:suporte@conectivasolucoes.com.br'
            data['class'] = 'alert-danger'

            return render(request, 'login.html', data)

        else:
            data['msg'] = ' Usuario ou Senha Invalido(s)'
            data['class'] = 'alert-danger'

            return render(request, 'login.html', data)
    else:
        return render(request, 'login.html')
def logout_view(request):
    logout(request)
    return redirect('index')

def index(request):
    # Sua lógica de visualização aqui
    return redirect('dologin')

@login_required
def form_Colaborador(request):
    FUNCAO_CHOICES = (
        ('Gerente', 'Gerente'),
        ('Administrador', 'Administrador'),
        ('Vendedor Interno', 'Vendedor Interno'),
        ('Vendedor Externo', 'Vendedor Externo'),
        ('Producao', 'Producao'),
        ('Auxiliar', 'Auxiliar'),
        ('Logistica', 'Logística'),
        ('Financeiro', 'Financeiro'),
        ('Recepcionista', 'Recepcionista'),
    )
    empresas = Empresas.objects.all()

    context = {
        'funcoes': FUNCAO_CHOICES,
        'empresas': empresas

    }

    if request.method == 'POST':
        nome = request.POST['nome']
        data_nascimento = request.POST['data']
        print (data_nascimento)
        data_admissao = request.POST['data_admissao']
        print (data_admissao)
        email = request.POST['email']
        telefone = request.POST['telefone']
        telefone_comercial = request.POST['telefone2']
        cpf = request.POST['cpf']
        empresa = request.POST['empresa']
        funcao = request.POST['funcao']
        endereco = request.POST['endereco']
        empresa = Empresas.objects.filter(nome=empresa).first()
        status = Status_Colaborador.objects.filter(descricao='ATIVO').first()

        novo_colaborador = Colaborador(
            nome=nome, data_nascimento=data_nascimento, status=status, data_admissao=data_admissao, cpf=cpf, email=email, telefone=telefone,
            funcao=funcao, empresa=empresa, endereco=endereco, telefone_comercial=telefone_comercial
        )
        novo_colaborador.save()

        novo_registro = RegistrosLog()
        novo_registro.colaborador = Colaborador.objects.filter(nome=request.user.username).first()  # Relaciona o log com o colaborador
        novo_registro.data = timezone.now()  # Usa timezone.now() para obter a data e hora atuais
        novo_registro.operacao = f'Criado o Colaborador: {nome}'  # Formata a string corretamente
        novo_registro.usuario = request.user.username  # Assume que há um campo 'usuario' em RegistrosLog
        novo_registro.save()

        return redirect('/consulta_colaboradores/')


    return render(request, 'cad_Colaborador2.html', context)

def limpar_cadastro_clientes():


    pedidos = Pedidos.objects.filter(cliente_pedido__flag=0)
    if not pedidos:
        print('Nao foram encontrados pedidos')
    for pedido in pedidos:
        print (pedido.numero,pedido.cliente_pedido)
        pedido.cliente_pedido.total_pedidos +=1
        pedido.cliente_pedido.flag =1
        pedido.cliente_pedido.save()






@login_required
def consulta_colaboradores (request):
    cpf = request.GET.get('cpf_filter')
    nome = request.GET.get('nome_filter')
    status_id = request.GET.get('status_filter')
    status = Status_Colaborador.objects.all()
    if (cpf or nome or status_id):
        query = Q()
        if cpf:
            query &=Q(cpf=cpf)
        if nome:
            query &= Q(nome__icontains=nome)
        if status_id:
            status_colaborador = Status_Colaborador.objects.filter(id=status_id).first()
            query &= Q(status=status_colaborador)
        colaboradores = Colaborador.objects.filter(query)
        print (colaboradores)


        context = {
            'colaboradores': colaboradores,
            'status': status

        }

        return render(request, 'consulta_colaboradores.html', context)

    else:
        colaboradores = Colaborador.objects.filter(status__descricao='ATIVO').order_by('nome')
        status = Status_Colaborador.objects.all()
        context = {
            'colaboradores': colaboradores,
            'status': status

        }

        return render (request, 'consulta_colaboradores.html', context)

@login_required
def editar_colaborador (request, id):
    colaborador = Colaborador.objects.filter(id=id).first()
    FUNCAO_CHOICES = (
        ('Gerente', 'Gerente'),
        ('Administrador', 'Administrador'),
        ('Vendedor Interno', 'Vendedor Interno'),
        ('Vendedor Externo', 'Vendedor Externo'),
        ('Producao', 'Producao'),
        ('Auxiliar', 'Auxiliar'),
        ('Logistica', 'Logística'),
        ('Financeiro', 'Financeiro'),
        ('Recepcionista', 'Recepcionista'),
    )
    status = Status_Colaborador.objects.all()

    context = {

        'colaborador': colaborador,
        'funcoes': FUNCAO_CHOICES,
        'status': status,
        'empresas': Empresas.objects.all(),
        'motivos': Motivos_Desligamento.objects.all()
    }
    if request.method == 'POST':
        colaborador.nome = request.POST['nome']
        colaborador.data_nascimento = request.POST['data']
        colaborador.data_admissao = request.POST['data_admissao']
        colaborador.email = request.POST['email']
        colaborador.telefone = request.POST['telefone']
        colaborador.telefone_comercial = request.POST['telefone_comercial']
        colaborador.cpf = request.POST['cpf']

        colaborador.funcao = request.POST['funcao']
        status_id = request.POST['status']
        status = Status_Colaborador.objects.filter(id=status_id).first()
        colaborador.status = status
        colaborador.endereco = request.POST['endereco']


        if status.descricao != 'ATIVO':
            data_desligamento = request.POST['data_desligamento']
            motivo_desligamento_id = request.POST['motivo_desligamento']
            motivo = Motivos_Desligamento.objects.filter(id=motivo_desligamento_id).first()
            colaborador.data_desligamento=data_desligamento
            colaborador.motivo_desligamento = motivo


        colaborador.save()
        novo_registro = RegistrosLog()
        novo_registro.colaborador = Colaborador.objects.filter(nome=request.user.username).first()  # Relaciona o log com o colaborador
        novo_registro.data = timezone.now()  # Usa timezone.now() para obter a data e hora atuais
        novo_registro.operacao = f'Editado o Colaborador: {colaborador}'  # Formata a string corretamente

        novo_registro.save()
        return redirect('/consulta_colaboradores/')
    else:

        return render (request, 'colaborador_editar.html', context)



def form_Clientes(request):
    if request.method == 'POST':
        print('entrou')
        nome = request.POST['nome']
        email = request.POST['email']
        documento = request.POST['documento']
        documento = ''.join(filter(str.isdigit, documento))
        cliente = Cliente.objects.filter(documento=documento).first()
        if cliente:
            return HttpResponse('Cliente Já Cadastrado!')
        nome_contato = request.POST['contato-cnpj']
        inscricao_estadual = request.POST['inscricao_estadual']
        dt_abertura = request.POST['data_abertura']
        telefone1 = request.POST['telefone1']
        telefone1= remover_caracteres_especiais(telefone1)
        telefone2 = request.POST['telefone2']
        if telefone2:
            telefone2 = remover_caracteres_especiais(telefone2)
        cep = request.POST['cep']
        bairro = request.POST['bairro']
        cidade = request.POST['cidade']
        uf = request.POST['uf']
        endereco = request.POST['endereco']
        numero_endereco = request.POST['numero_endereco']
        complemento_endereco = request.POST['complemento_endereco']
        ponto_referencia = request.POST['ponto_referencia']
        modo_id =  request.POST['modo_indicacao']
        modo = Tipo_Indicacao.objects.filter(id=modo_id).first()

        pagamento_boleto = 'NAO'

        if len(documento) == 14:
            if check_data(dt_abertura):
                pagamento_boleto='SIM'
            else:
                pagamento_boleto = 'NAO'
        else:
            pass



        novo_cliente = Cliente(
            nome=nome, nome_contato=nome_contato, dt_abertura=dt_abertura if dt_abertura else None,
            inscricao_estadual=inscricao_estadual if inscricao_estadual else None, email=email, documento=documento, telefone1=telefone1, telefone2=telefone2,
            CEP=cep, estado=uf, cidade=cidade, bairro=bairro, endereco=endereco,
            numero_endereco=numero_endereco,ponto_referencia=ponto_referencia, complemento=complemento_endereco, pagamento_boleto=pagamento_boleto, tipo_entrada = modo, data_cadastro =datetime.today()
        )
        novo_cliente.save()
        novo_registro = RegistrosLog()
        novo_registro.colaborador = Colaborador.objects.filter(
            nome=request.user.username).first()  # Relaciona o log com o colaborador
        novo_registro.data = timezone.now()  # Usa timezone.now() para obter a data e hora atuais
        novo_registro.operacao = f'Cadastrado Cliente {nome},{documento}'  # Formata a string corretamente
        novo_registro.usuario = request.user.username  # Assume que há um campo 'usuario' em RegistrosLog
        novo_registro.save()

        return redirect('/painel/')

    # Verificar se o cliente foi encontrado
    cliente_encontrado = False  # Defina essa variável com base na lógica de busca do cliente

    if not cliente_encontrado:

        modos = Tipo_Indicacao.objects.all()

        # Renderizar o formulário de cadastro diretamente
        context = {
            'bairros': Bairro.objects.all(),
            'modos': modos
        }
        return render(request, 'CadCliente2.html', context)

    # Renderizar o formulário normalmente se o cliente for encontrado


    context = {
        'bairros': Bairro.objects.all()

    }
    return render(request, 'CadCliente.html', context)

def editar_clientes(request,documento):
    cliente = Cliente.objects.filter(documento=documento).first()

    if request.method == 'POST':
        cliente.nome = request.POST['nome']
        cliente.email = request.POST['email']
        cliente.nome_contato = request.POST['contato-cnpj']
        cliente.inscricao_estadual = request.POST['inscricao_estadual']
        cliente.dt_abertura = request.POST['data_abertura']
        cliente.telefone1 = request.POST['telefone1']
        cliente.telefone2 = request.POST['telefone2']
        cliente.CEP = request.POST['cep']
        cliente.bairro = request.POST['bairro']
        cliente.cidade = request.POST['cidade']
        cliente.estado = request.POST['uf']
        cliente.endereco = request.POST['endereco']
        cliente.numero_endereco = request.POST['numero_endereco']
        cliente.complemento = request.POST['complemento_endereco']
        cliente.ponto_referencia = request.POST['ponto_referencia']

        cliente.save()
        return redirect('/consulta_cliente/')

    context = {
        'bairros': Bairro.objects.all(),
        'cliente': cliente
    }
    return render(request, 'editar_clientes.html', context)

def form_fornecedor(request):
    if request.method == 'POST':
        nome = request.POST['nome']
        email = request.POST['email']
        documento = request.POST['documento']
        documento = ''.join(filter(str.isdigit, documento))
        telefone1 = request.POST['telefone1']
        telefone1 = ''.join(filter(str.isdigit, telefone1))
        cep = request.POST['cep']
        cep = ''.join(filter(str.isdigit, cep))
        bairro=request.POST['bairro']
        cidade = request.POST['cidade']
        uf = request.POST['uf']
        endereco = request.POST['endereco']
        numero_endereco = request.POST['numero_endereco']
        complemento = request.POST['complemento_endereco']


        novo_fornecedor = Fornecedor(
            nome=nome, email=email, documento=documento, telefone=telefone1,
            cep=cep, uf=uf, cidade=cidade, bairro=bairro, logradouro=endereco, numero=numero_endereco, complemento=complemento
        )
        novo_fornecedor.save()

    return render (request,'cadastro_fornecedor.html')

@login_required
def pedido_create(request):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    if colaborador.funcao == 'Gerente' or colaborador.funcao == 'Financeiro' or colaborador.funcao == 'Vendedor Interno' or colaborador.funcao == 'Vendedor Externo':
        return redirect('/pedido_gerente/')


    vendedores = Colaborador.objects.filter(Q(funcao='Vendedor Interno') | Q(funcao='Vendedor Externo'))
    clientes = Cliente.objects.all()
    produtos = Produtos.objects.all()
    bairros = Bairro.objects.all()
    tipos_pagamento_avista = TipoPgto.objects.filter(aceita_parcela='NÃO')
    tipos_pagamento_parcelado = TipoPgto.objects.filter(aceita_parcela='SIM')
    fornecedores = Fornecedor.objects.filter(tipo_fornecedor='1')

    cep_entrega = ''
    uf_entrega = ''
    cidade_entrega = ''
    bairro_entrega = ''
    endereco_entrega = ''
    numero_end_entrega = '0'
    complemento = ''
    total_avista ='0'
    total_parcelado = '0'
    desconto ='0'
    context = {
        'vendedor': colaborador.nome,
        'vendedores': vendedores,
        'clientes': clientes,
        'produtos': produtos,
        'bairros': bairros,
        'tipos_pagamento_avista': tipos_pagamento_avista,
        'tipos_pagamento_parcelado' : tipos_pagamento_parcelado,
        'fornecedores': fornecedores,
        'colaborador': colaborador,
    }

    if request.method == 'POST':
        numero_pedido = request.POST.get('numero_pedido')
        datapedido = request.POST.get('data-pedido')
        prazo_entrega = request.POST.get('data_entrega')
        vendedor = request.POST.get('vendedor1')
        vendedor_comissao = request.POST.get('comissao1')
        vendedor_comissao = vendedor_comissao.replace('%', '') if vendedor_comissao else '0'

        vendedor2 = request.POST.get('vendedor2')
        vendedor2_comissao = request.POST.get('comissao2')
        vendedor2_comissao = vendedor2_comissao.replace('%', '') if vendedor2_comissao else '0'
        documento_cliente = request.POST.get('documento_cliente')
        documento = ''.join(filter(str.isdigit, documento_cliente))
        endereco_opcao = request.POST.get('endereco-opcao')
        frete_valor = request.POST.get('frete')
        frete_valor = round(float(frete_valor), 2) if frete_valor else 0
        frete_fornecedor = request.POST.get('fretador')
        colaborador = Colaborador.objects.get(nome=vendedor)

        empresa = colaborador.empresa



        cliente_pedido = Cliente.objects.get(documento=documento)

        if cliente_pedido:
            cliente_id = cliente_pedido.id
        else:
            cliente_id = None  # Trate o caso em que o cliente não é encontrado

        if endereco_opcao == 'cadastro':
            if cliente_pedido:
                cep_entrega = cliente_pedido.CEP
                uf_entrega = cliente_pedido.estado
                cidade_entrega = cliente_pedido.cidade
                bairro_entrega = cliente_pedido.bairro
                endereco_entrega = cliente_pedido.endereco
                numero_end_entrega = cliente_pedido.numero_endereco
                complemento = cliente_pedido.complemento
        elif endereco_opcao == 'opcional':
            cep_entrega = request.POST.get('cep-entrega-opcional')
            endereco_entrega = request.POST.get('endereco-entrega-opcional')
            numero_end_entrega = request.POST.get('numero-endereco-opcional')
            complemento = request.POST.get('complemento-entrega-opcional')

            bairro_nome, bairro_cidade_uf = request.POST.get('bairro-entrega-opcional').split(', ')
            bairro = Bairro.objects.get(nome=bairro_nome, cidade=bairro_cidade_uf.split('-')[0], uf=bairro_cidade_uf.split('-')[1])
            uf_entrega = bairro.uf
            cidade_entrega = bairro.cidade
            bairro_entrega = bairro.nome

        total_pedido_parcial = Itens_Pedido.objects.filter(pedido=numero_pedido).aggregate(total=Sum('total_item'))['total']
        desconto = request.POST.get('desconto')
        desconto = float(desconto) if desconto else 0
        total_pedido =round(float(total_pedido_parcial - desconto), 2)


        pedido_obj = Pedidos(
            numero=numero_pedido,
            datapedido=datapedido,
            valorTotal=total_pedido,
            valorTotalParcial=total_pedido_parcial,
            descontototal=desconto,
            cliente_pedido=cliente_pedido,
            prazo_entrega=prazo_entrega,
            Vendedor=vendedor,
            Vendedor2=vendedor2,
            vendedor_comissao=vendedor_comissao,
            vendedor2_comissao=vendedor2_comissao,
            cep_entrega=cep_entrega,
            uf_entrega=uf_entrega,
            cidade_entrega=cidade_entrega,
            bairro_entrega=bairro_entrega,
            endereco_entrega=endereco_entrega,
            numero_end_entrega=numero_end_entrega,
            complemento=complemento,
            empresa_pedido=empresa

        )
        pedido_obj.save()


        total_avista = request.POST.get('total_avista')
        total_avista = round(float(total_avista), 2) if total_avista else 0
        descricao_avista = request.POST.get('forma_pgto_avista', '')
        pagamento_avista = TipoPgto.objects.get(descricao=descricao_avista) if descricao_avista else None

        total_parcelado = request.POST.get('total_parcelado')
        total_parcelado = round(float(total_parcelado), 2) if total_parcelado else 0

        descricao_parcelado = request.POST.get('forma_pgto_parcelado', '')
        pagamento_parcelado = TipoPgto.objects.get(descricao=descricao_parcelado) if descricao_parcelado else None


        if total_avista > 0:
            if pagamento_avista:
                contas_avista_existentes = Contas_Receber.objects.filter(
                    pedido=numero_pedido,
                    tipo_pgto=pagamento_avista.descricao,
                    numero_parcela=1,
                    total_parcelas=1
                )
                if contas_avista_existentes.exists():
                    # As contas a receber já existem, faça algo caso necessário
                    pass
                else:
                    # Calcular a data de vencimento para pagamento à vista
                    if pagamento_avista:
                        prazo_dias = int(pagamento_avista.prazo_dias)
                    else:
                        prazo_dias = 0

                    data_vencimento_avista = datetime.now().date() + timedelta(days=prazo_dias)

                    # Gerar conta a receber para pagamento à vista
                    contas_receber_avista = Contas_Receber(
                        pedido=numero_pedido,
                        descricao=cliente_id,
                        data_vencimento=data_vencimento_avista,
                        data_pagamento=None,
                        tipo_pgto=pagamento_avista.descricao,
                        numero_parcela=1,
                        total_parcelas=1,
                        valor=total_avista
                    )
                    contas_receber_avista.save()

        if total_parcelado > 0:
            contas_parcelado_existentes = Contas_Receber.objects.filter(
                pedido=numero_pedido,
                tipo_pgto=pagamento_parcelado.descricao,
                total_parcelas=pagamento_parcelado.numero_parcelas,
            )
            if contas_parcelado_existentes.exists():
                # As contas a receber já existem, faça algo caso necessário
                pass
            else:
                numero_parcelas = pagamento_parcelado.numero_parcelas
                prazo_dias = int(pagamento_parcelado.prazo_dias)

                # Gerar contas a receber para pagamento parcelado
                for parcela in range(numero_parcelas):
                    data_vencimento = datetime.now().date() + timedelta(days=prazo_dias + parcela * 30)

                    contas_receber_parcelado = Contas_Receber(
                        pedido=numero_pedido,
                        descricao=cliente_id,
                        data_vencimento=data_vencimento,
                        data_pagamento=None,
                        tipo_pgto=pagamento_parcelado.descricao,
                        numero_parcela=parcela + 1,
                        total_parcelas=numero_parcelas,
                        valor=round((total_parcelado / numero_parcelas), 2)
                    )
                    contas_receber_parcelado.save()

        if frete_valor > 0:
            frete_fornecedor = Fornecedor.objects.filter(nome=frete_fornecedor).first()
            conta_obj = Contas_Pagar(
                pedido=numero_pedido,
                descricao="FRETE REF PEDIDO",
                fornecedor=frete_fornecedor,
                valor=frete_valor,

            )
            conta_obj.save()
            atualiza_pedido = Pedidos.objects.get(numero=numero_pedido)
            atualiza_pedido.frete_valor = frete_valor
            atualiza_pedido.frete_fornecedor=frete_fornecedor
            atualiza_pedido.save()

            return redirect('/painel/')

        else:
            return redirect('/painel/')


    return render(request, 'novopedido.html', context)

def get_clientes(request):

    tipo_pedido = request.GET.get('tipo_pedido')
    documento_cliente_1 = request.GET.get('documento')
    numero_pedido = request.GET.get('numero_pedido')
    if documento_cliente_1:
        documento = ''.join(filter(str.isdigit, documento_cliente_1))
        cliente = Cliente.objects.filter(documento=documento).first()
    else:
        documento_cliente_2 = request.GET.get('documento2')
        documento = ''.join(filter(str.isdigit, documento_cliente_2))
        cliente = Cliente.objects.filter(documento=documento).first()

    if cliente:

        cliente_tipo = cliente.tipo_cliente.descricao.strip()
        tipo_pedido_tipo = tipo_pedido.strip()

        if cliente_tipo == tipo_pedido_tipo:
            documento = formatar_documento(cliente.documento)


            data = { 'cliente':{

                'nome': cliente.nome,
                'email': cliente.email,
                'documento': documento,
                'telefone1': cliente.telefone1,
                'telefone2': cliente.telefone2,
                'CEP': cliente.CEP,
                'estado': cliente.estado,
                'cidade': cliente.cidade,
                'bairro': cliente.bairro,
                'endereco': cliente.endereco,
                'numero_endereco': cliente.numero_endereco,
                'complemento': cliente.complemento,
                'ponto_referencia': cliente.ponto_referencia,
                'dt_abertura':cliente.dt_abertura,
                'contato_cnpj':cliente.nome_contato,
                'tipo_cliente': cliente.tipo_cliente.descricao,
                'score_cliente': cliente.score,
                'boleto': cliente.pagamento_boleto
            }
            }
            pedido = Temp_Pedidos.objects.filter(pedido=numero_pedido).first()
            if pedido:
                pedido.cliente=cliente
                pedido.tipo_pedido = Tipo_Pedido.objects.filter(descricao=tipo_pedido).first()
                pedido.save()


        else: data = {'message':'CLIENTE INCOMPATIVEL COM TIPO DO PEDIDO','cliente': None}
    else:
        data = {'cliente': None}



    return JsonResponse(data)

def form_Clientes2(request):
    if request.method == 'POST':
        nome = request.POST['nome']
        email = request.POST['email']
        documento = request.POST['documento']
        documento = ''.join(filter(str.isdigit, documento))
        telefone1 = request.POST['telefone1']
        telefone1 = ''.join(filter(str.isdigit, telefone1))
        telefone2 = request.POST['telefone2']
        telefone2 = ''.join(filter(str.isdigit, telefone2))
        cep = request.POST['cep']
        cep = ''.join(filter(str.isdigit, cep))
        bairro=request.POST['bairro']
        cidade = request.POST['cidade']
        uf = request.POST['uf']
        endereco = request.POST['endereco']
        numero_endereco = request.POST['numero_endereco']
        complemento = request.POST['complemento_endereco']


        novo_cliente = Cliente(
            nome=nome, email=email, documento=documento, telefone1=telefone1, telefone2=telefone2,
            CEP=cep, estado=uf, cidade=cidade, bairro=bairro, endereco=endereco, numero_endereco=numero_endereco, complemento=complemento
        )
        novo_cliente.save()
        return redirect('/painel/')
    context = {
        'bairros': Bairro.objects.all()
    }

    return render(request, 'CadCliente.html', context)

def consulta_clientes(request):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    if colaborador.empresa.flag_pagamento == 1:
        return redirect('/dologin/')
    elif not colaborador:
        return redirect('/dologin/')
    else:
        pass
    clientes_boleto=Cliente.objects.filter(pagamento_boleto='NAO')
    pedidos = Pedidos.objects.all()

    for cliente in clientes_boleto:
        # Filtrar os pedidos do cliente atual
        pedidos_cliente = pedidos.filter(cliente_pedido=cliente)
        if cliente.pagamento_boleto == 'NAO':

            if len(cliente.documento) == 14:
                if check_data(cliente.dt_abertura):
                    cliente.pagamento_boleto='SIM'
                else:
                    cliente.pagamento_boleto = 'NAO'
            else:
                cliente.pagamento_boleto = 'NAO'
        else:
            pass

        # Atualizar o total de pedidos do cliente
        total_pedidos = pedidos_cliente.count()
        cliente.total_pedidos = total_pedidos
        cliente.save()

        # Obter o último pedido do cliente
        ultimo_pedido = pedidos_cliente.aggregate(ultimo_pedido=Max('numero'))['ultimo_pedido']
        cliente.ultimo_pedido = ultimo_pedido
        cliente.save()


    lista_clientes = Cliente.objects.all()
    for cliente in lista_clientes:
        if cliente.total_pedidos is None:
            print (cliente, cliente.total_pedidos)
            cliente.total_pedidos = 0
            cliente.save()
        else:
            pass


    clientes = lista_clientes.order_by('-id')[:10]


    context ={
        'colaborador':colaborador,
        'clientes':clientes

    }
    if request.method == 'GET':
        nome_cliente = request.GET.get('nome_cliente_filter')
        documento_cliente = request.GET.get('documento_cliente_filter')
        documento_cliente = ''.join(filter(str.isdigit, documento_cliente)) if documento_cliente else None
        email_cliente = request.GET.get('email_cliente_filter')
        if nome_cliente or documento_cliente or email_cliente:
            query = Q()

            if nome_cliente:
                query &= Q(nome__icontains=nome_cliente)
            if documento_cliente:
                query &= Q(documento__icontains=documento_cliente)
            if email_cliente:
                query &= Q(email__icontains=email_cliente)

            clientes = Cliente.objects.filter(query)
        else:
            clientes = Cliente.objects.all().order_by('-id')[:300]

        context = {

            'colaborador': colaborador,
            'clientes': clientes

        }
        return render (request,'consulta_clientes.html',context)
    else:
        return render(request, 'consulta_clientes.html', context)

def cadastra_revenda (request):
    if request.method == 'POST':
        data = json.loads(request.body)  # Obtém o valor do parâmetro 'data' da URL
        documento = data.get('documento')
        nivel = data.get('nivel')
        cliente = Cliente.objects.get(documento=documento)
        tipo_cliente = Tipo_Pedido.objects.get(descricao='REVENDA')
        cliente.tipo_cliente = tipo_cliente
        cliente.save()
        revenda = Revenda (


            cliente=cliente,
            nivel = nivel

        )
        revenda.save()
        return JsonResponse({'message': 'Dados recebidos com sucesso'})

    else:
        return JsonResponse({'error': 'Erro no cadastro revenda!'}, status=405)
@csrf_exempt
def cliente_faturamento (request):
    if request.method == 'POST':
        data = json.loads(request.body)  # Obtém o valor do parâmetro 'data' da URL
        documento = data.get('documento')
        cliente = Cliente.objects.get(documento=documento)
        if cliente.pagamento_boleto == 'SIM':
            cliente.pagamento_boleto = 'NAO'
            cliente.save()
            return JsonResponse({'message': 'Cliente Desabilitado Para Faturamento!'})
        elif cliente.pagamento_boleto == 'NAO':
            cliente.pagamento_boleto = 'SIM'
            cliente.save()
            return JsonResponse({'message': 'Cliente Habilitado Para Faturamento!'})
        else:
            return JsonResponse({'error': 'Cliente Não Encontrado!'}, status=405)
    else:
        return JsonResponse({'error': 'Erro na alteração de Status de Faturamento!'}, status=405)

@csrf_exempt
def get_pedidos(request):
    data_atual = date.today()
    empresa_pedido = request.POST.get('empresa_pedido')
    print(empresa_pedido)

    vendedor = request.POST.get ('vendedor')
    print (vendedor)
    colaborador = Colaborador.objects.get(nome=vendedor)
    if colaborador.funcao == 'Gerente' or colaborador.funcao == 'Financeiro':
        empresa = Empresas.objects.get(nome=empresa_pedido)

    else:
        empresa = Empresas.objects.get(nome=colaborador.empresa.nome)

    empresa_identificador = empresa.id
    # Obtém o número do último pedido cadastrado no dia atual
    ultimo_pedido = empresa.ultimo_pedido_ref
    if ultimo_pedido:
        ultimo_numero_pedido = int(str(ultimo_pedido)[-5:]) + 1
    else:
        ultimo_numero_pedido = 1  # Caso não haja pedidos anteriores

    id_pedido = ultimo_numero_pedido

    while Pedidos.objects.filter(numero__endswith=str(id_pedido)[-5:],empresa_pedido=empresa).exists():
        id_pedido += 1

    # Combine a data atual com o ID do pedido para formar o número do pedido
    numero_pedido = f'{data_atual.strftime("%d%m%y")}{empresa_identificador:02}{colaborador.id:02}{id_pedido:05}'
    numero_pedido = int(numero_pedido)

    #Cria uma Instancia Temporaria do Pedido mas se ja existir ele apenas atualiza ela
    pedido = Temp_Pedidos.objects.filter(pedido=numero_pedido).first()
    if pedido:
        pedido.empresa=empresa
        pedido.save()
    else:
        pedido = Temp_Pedidos(

            pedido=numero_pedido,
            empresa=empresa

        )
        pedido.save()
    empresa.ultimo_pedido_ref = numero_pedido
    empresa.save()

    # Retorna o número do pedido como uma resposta JSON
    return JsonResponse({'numero_pedido': numero_pedido})


@csrf_exempt
def item_save(request):
    if request.method == 'POST':

        items = json.loads(request.POST.get('items'))
        tipo_pedido = request.POST.get('tipo_pedido')

        pedido = request.POST.get('pedido')
        print(items)
        Itens_Pedido.objects.filter(pedido=pedido).delete()
        ultimo_comprimento =0.00
        ultimo_largura = 0.00
        ultimo_quantidade = 0

        try:
            for item_data in items:

                item = Itens_Pedido()
                item.pedido = pedido
                item.observacao_item = item_data['observacao']
                print(item.observacao_item)
                item.quantidade = int(item_data['quantidade'])
                item.quantidade=int(item.quantidade)
                item.nome = item_data['produto']
                item.tipo = item_data['tipo_produto']
                item.comprimento =item_data['comprimento']
                item.comprimento= float(item.comprimento)
                item.comprimento = round(item.comprimento,3)
                if tipo_pedido == 'PADRAO':
                    produto = Produtos.objects.get(descricao=item.nome)
                    item.produto_id = produto.id
                else:
                    produto = Produtos_Revenda.objects.get(descricao=item.nome)
                    item.produto_revenda_id = produto.id

                if item.tipo == 'PADRAO' or item.tipo == 'ADICIONAL':
                    ultimo_comprimento = item.comprimento
                    ultimo_largura = item.largura


                    if (produto.tipo_medida == 'linear'):
                        item.largura = 1
                    else:
                        item.largura = item_data['largura']
                        item.largura = float(item.largura)
                        item.largura = round(item.largura,3)

                    ultimo_quantidade = item.quantidade
                    ultimo_largura = item.largura
                    item.medida_final = item.largura * item.comprimento
                    item.medida_final = round((item.medida_final),3)
                    print(item.nome, item.medida_final)
                    if (item_data['desconto']):
                        item.desconto = float(item_data['desconto'])
                    else: item.desconto = 0.00
                    item.preco = item_data['preco']
                    item.preco = round(float(item.preco),2)

                    item.total_item = (item.quantidade*item.preco)-item.desconto
                    item.total_item = round((item.total_item), 2)

                    item.valor_base = (item.preco) / item.medida_final
                    item.valor_base = round((item.valor_base), 2)

                    if tipo_pedido == 'PADRAO':
                        if item.tipo == 'BRINDE':
                            item.largura = "{:.3f}".format(item.largura)
                            item.comprimento = "{:.3f}".format(item.comprimento)

                            item.save()
                        else:
                            print(item.comprimento, item.largura, item.nome)
                            if item.valor_base < produto.preco_base_comissao1:
                                return JsonResponse({'error': 'Erro ao processar os itens', 'erro_valor': 'ERRO VALOR' })
                            item.largura = "{:.3f}".format(item.largura)
                            item.comprimento = "{:.3f}".format(item.comprimento)

                            item.save()

                    elif tipo_pedido == 'REVENDA':
                        documento = request.POST.get('documento')
                        documento_cliente = ''.join(filter(str.isdigit, documento))
                        cliente = Cliente.objects.get(documento=documento_cliente)
                        revenda = Revenda.objects.get(cliente=cliente)

                        if item.tipo == 'BRINDE':
                            item.largura = "{:.3f}".format(item.largura)
                            item.comprimento = "{:.3f}".format(item.comprimento)
                            item.save()
                        else:

                            if revenda.nivel == '1':
                                if item.valor_base < produto.preco_base_1:
                                    return JsonResponse({'error': 'Erro ao processar os itens', 'erro_valor': 'ERRO VALOR'})

                            elif revenda.nivel == '2':
                                if item.valor_base < produto.preco_base_2:
                                    return JsonResponse({'error': 'Erro ao processar os itens', 'erro_valor': 'ERRO VALOR'})

                            elif revenda.nivel == '3':
                                if item.valor_base < produto.preco_base_3:
                                    return JsonResponse({'error': 'Erro ao processar os itens', 'erro_valor': 'ERRO VALOR'})

                            elif revenda.nivel == '4':
                                if item.valor_base < produto.preco_base_4:
                                    return JsonResponse({'error': 'Erro ao processar os itens', 'erro_valor': 'ERRO VALOR'})

                            else:
                                if item.valor_base < produto.preco_base_3:
                                    return JsonResponse({'error': 'Erro ao processar os itens', 'erro_valor': 'ERRO VALOR'})

                            item.largura = "{:.3f}".format(item.largura)
                            item.comprimento = "{:.3f}".format(item.comprimento)

                            item.save()

                if item.tipo == 'EXCEPCIONAL':

                    if ultimo_comprimento > 0:
                        if 'PERFIL' in item.nome:
                            if '4 LADOS' in item.nome:
                                item.comprimento = (ultimo_comprimento*ultimo_largura)*ultimo_quantidade*2

                            elif '3 LADOS' in item.nome:
                                if '1C' in item.nome:
                                    item.comprimento = (ultimo_comprimento + (2*ultimo_largura))*ultimo_quantidade
                                else:
                                    item.comprimento = ((2*ultimo_comprimento)+ultimo_largura)*ultimo_quantidade
                            elif '2 LADOS' in item.nome:
                                if 'CL' in item.nome:
                                    item.comprimento = (ultimo_comprimento+ultimo_largura)*ultimo_quantidade
                                elif item.nome == 'PERFIL 2 LADOS C':
                                    item.comprimento = (ultimo_comprimento*2)*ultimo_quantidade
                                else:
                                    item.comprimento = (ultimo_largura*2)*ultimo_quantidade
                            else:
                                if 'C' in item.nome:
                                    item.comprimento = ultimo_comprimento*ultimo_quantidade
                                else:
                                    item.comprimento = ultimo_largura*ultimo_quantidade
                            item.largura = 1
                            item.medida_final = item.comprimento
                            item.preco = item_data['preco']
                            item.preco = round(float(item.preco), 2)
                            item.total_item = (item.quantidade * item.preco) - item.desconto
                            item.total_item = round((item.total_item), 2)
                            item.valor_base = (item.preco - item.desconto) / item.medida_final
                            item.valor_base = round((item.valor_base), 2)
                            if tipo_pedido == 'PADRAO':
                                if item.tipo == 'BRINDE':
                                    item.largura = "{:.3f}".format(item.largura)
                                    item.comprimento = "{:.3f}".format(item.comprimento)
                                    item.save()
                                else:
                                    if item.valor_base < produto.preco_base_comissao1:
                                        return JsonResponse(
                                            {'error': 'Erro ao processar os itens', 'erro_valor': 'ERRO VALOR'})
                                    item.largura = "{:.3f}".format(item.largura)
                                    item.comprimento = "{:.3f}".format(item.comprimento)


                                    item.save()

                            elif tipo_pedido == 'REVENDA':
                                documento = request.POST.get('documento')
                                documento_cliente = ''.join(filter(str.isdigit, documento))
                                cliente = Cliente.objects.get(documento=documento_cliente)
                                revenda = Revenda.objects.get(cliente=cliente)

                                if item.tipo == 'BRINDE':
                                    item.largura = "{:.3f}".format(item.largura)
                                    item.comprimento = "{:.3f}".format(item.comprimento)
                                    item.save()
                                else:

                                    if revenda.nivel == '1':
                                        if item.valor_base < produto.preco_base_1:
                                            return JsonResponse(
                                                {'error': 'Erro ao processar os itens', 'erro_valor': 'ERRO VALOR'})

                                    elif revenda.nivel == '2':
                                        if item.valor_base < produto.preco_base_2:
                                            return JsonResponse(
                                                {'error': 'Erro ao processar os itens', 'erro_valor': 'ERRO VALOR'})
                                    elif revenda.nivel == '3':
                                        if item.valor_base < produto.preco_base_3:
                                            return JsonResponse(
                                                {'error': 'Erro ao processar os itens', 'erro_valor': 'ERRO VALOR'})
                                    elif revenda.nivel == '4':
                                        if item.valor_base < produto.preco_base_4:
                                            return JsonResponse(
                                                {'error': 'Erro ao processar os itens', 'erro_valor': 'ERRO VALOR'})
                                    else:
                                        if item.valor_base < produto.preco_base_3:
                                            return JsonResponse(
                                                {'error': 'Erro ao processar os itens', 'erro_valor': 'ERRO VALOR'})

                                    item.largura = "{:.3f}".format(item.largura)
                                    item.comprimento = "{:.3f}".format(item.comprimento)



                                    item.save()
                            ultimo_comprimento = 0
                            ultimo_largura = 0
                            ultimo_quantidade = 0


            return JsonResponse({'message': 'Dados recebidos com sucesso'})
        except Exception as e:
            return JsonResponse({'error': 'Erro ao processar os itens', 'message': str(e)}, status=400)
    else:
        return JsonResponse({'error': 'Método não permitido'}, status=405)

@csrf_exempt
def saveitenspedido(request):


    if request.method == 'POST':

        total1 = request.POST.get('total1')


        total_pagamentos = 0
        if total1:
            total1 = float(total1)
            print(total1)
            total_pagamentos += total1

        else:
            total1 = 0
        totalc1 = float (request.POST.get('totalc1'))
        print(totalc1)
        if totalc1 >0:
            totalc1 = float(totalc1)
            total_pagamentos += totalc1

        else:
            totalc1 =0
        print(total_pagamentos)
        total2 = request.POST.get('total2')
        if total2:
            total2 = float(total2)
            total_pagamentos += total2
        else:
            total2 = 0
        totalc2 = request.POST.get('totalc2')
        if totalc2:
            totalc2 = float(totalc2)
            total_pagamentos += totalc2
        else:
            totalc2 = 0
        valorTotal = request.POST.get('valorTotal')
        if valorTotal:
            valorTotal = float(valorTotal)
        else:
            valorTotal = 0


        print (total_pagamentos, valorTotal)

        if total_pagamentos != valorTotal:
            return JsonResponse({'error': 'Os valores do Total do pedido e Total de Pagamento não conferem!'})

        else:
            print('entrou')
            items = json.loads(request.POST.get('items'))

            tipo_pedido = request.POST.get('tipopedido')
            print(tipo_pedido)
            frete = request.POST.get('frete')
            frete_embutido = frete
            quantidade_itens = 0

            quantidade_itens = len(items)
            print('total de itens:', quantidade_itens)
            frete_item = 0.00
            if frete:
                frete = float(frete)
                frete_item = frete / quantidade_itens
                frete_item = round (frete_item,2)

            print ('frete_item:', frete_item)
            pedido = request.POST.get('pedido')

            lista_itens = Itens_Pedido.objects.filter(pedido=pedido).first()
            if lista_itens:
                return JsonResponse({'error': 'Já existe um pedido com o mesmo número'})




            try:
                for item_data in items:


                    item = Itens_Pedido()
                    item.pedido = pedido
                    item.observacao_item = item_data['observacao']

                    item.quantidade = (item_data['quantidade'])
                    item.quantidade=int(item.quantidade)

                    item.nome = item_data['produto']
                    print('item_nome:',item.nome)
                    item.tipo = item_data['tipo_produto']

                    item.comprimento =item_data['comprimento']
                    item.comprimento= float(item.comprimento)
                    item.comprimento = round(item.comprimento,3)
                    comprimento_dig = item.comprimento
                    largura_dig = item_data['largura']
                    largura_dig = round(largura_dig,3)
                    adicional = 0.00
                    adicional = item_data['adicional']

                    if tipo_pedido == 'PADRAO':
                        print ('PEDIDO PADRAO')
                        produto = Produtos.objects.filter(descricao=item.nome).first()
                        if produto is None:
                            return JsonResponse({'error': f'Produto Não Encontrado: {item.nome}'})
                        item.produto_id = produto.id
                    else:
                        produto = Produtos_Revenda.objects.filter(descricao=item.nome).first()
                        if produto is None:
                            return JsonResponse({'error': f'Produto Não Encontrado: {item.nome}'})
                        item.produto_revenda_id = produto.id


                    if 'MPC' in produto.descricao:
                        item.comprimento = produto.tamanho_padrao
                        item.largura = item_data['largura']
                        item.largura = float(item.largura)
                        item.largura = round(item.largura, 3)
                    elif 'MPL' in produto.descricao:
                        item.largura = produto.tamanho_padrao
                        item.comprimento = item_data['comprimento']
                        item.comprimento = float(item.comprimento)
                        item.comprimento = round(item.comprimento, 3)
                    else:
                        item.comprimento = item_data['comprimento']
                        item.largura = item_data['largura']

                    if (produto.tipo_medida == 'linear'):
                        item.largura = 1
                        item.medida_final = item.largura * item.comprimento
                    elif (produto.tipo_medida == 'm2') :

                        item.medida_final = item.largura * item.comprimento
                    elif (produto.tipo_medida == 'pc'):
                        item.medida_final = item.quantidade

                    item.medida_final = round((item.medida_final),3)
                    print(item.nome, item.medida_final)

                    item.preco = item_data['preco']
                    item.preco = round(float(item.preco),2)
                    print(item.preco)

                    item.valor_base = (item.preco + adicional) / item.medida_final
                    item.valor_base = round((item.valor_base), 2)

                    item.preco = (frete_item/item.quantidade) + adicional + item.preco
                    item.preco =  round(item.preco,2)
                    item.total_item = (item.quantidade * item.preco)
                    item.total_item = round((item.total_item), 2)
                    print(item.comprimento, item.largura, item.nome)
                    item.largura = "{:.3f}".format(largura_dig)
                    item.comprimento = "{:.3f}".format(comprimento_dig)
                    item.adicional = adicional
                    if frete_embutido:
                        item.frete_embutido = frete_embutido
                    item.save()


                    quantidade_itens = quantidade_itens - 1

                    if frete > 0:
                        frete = frete - frete_item
                        if quantidade_itens > 0:
                            frete_item = frete / quantidade_itens
                            frete_item = round(frete_item, 2)


                return JsonResponse({'message': 'Dados recebidos com sucesso'})
            except Exception as e:
                return JsonResponse({'error': 'Erro ao processar os itens', 'message': str(e)}, status=400)

@csrf_exempt
def editaitenspedido(request):


    if request.method == 'POST':
        print('entrou')
        items = json.loads(request.POST.get('items'))

        tipo_pedido = request.POST.get('tipopedido')
        print(tipo_pedido)
        frete = request.POST.get('frete')
        quantidade_itens = 0
        quantidade_itens = len(items)
        print('total de itens:', quantidade_itens)
        frete_item = 0.00
        if frete:
            frete = float(frete)
            frete_item = frete / quantidade_itens
            frete_item = round (frete_item,2)

        print ('frete_item:', frete_item)
        npedido = request.POST.get('pedido')
        pedido = Pedidos.objects.filter(numero=npedido).first()

        Itens_Pedido.objects.filter(pedido=npedido).delete()
        try:
            for item_data in items:


                item = Itens_Pedido()
                item.pedido = npedido
                item.observacao_item = item_data['observacao']

                item.quantidade = (item_data['quantidade'])
                item.quantidade=int(item.quantidade)

                item.nome = item_data['produto']
                print('item_nome:',item.nome)
                item.tipo = item_data['tipo_produto']

                item.comprimento =item_data['comprimento']
                item.comprimento= float(item.comprimento)
                item.comprimento = round(item.comprimento,3)
                adicional = 0.00
                adicional = item_data['adicional']

                if pedido.tipo_pedido.descricao == 'PADRAO':
                    print ('PEDIDO PADRAO')
                    produto = Produtos.objects.filter(descricao=item.nome).first()
                    if produto is None:
                        return JsonResponse({'error': f'Produto Não Encontrado: {item.nome}'})
                    item.produto_id = produto.id
                else:
                    produto = Produtos_Revenda.objects.filter(descricao=item.nome).first()
                    if produto is None:
                        return JsonResponse({'error': f'Produto Não Encontrado: {item.nome}'})
                    item.produto_revenda_id = produto.id


                if 'MPC' in produto.descricao:
                    item.comprimento = produto.tamanho_padrao
                    item.largura = item_data['largura']
                    item.largura = float(item.largura)
                    item.largura = round(item.largura, 3)
                elif 'MPL' in produto.descricao:
                    item.largura = produto.tamanho_padrao
                    item.comprimento = item_data['comprimento']
                    item.comprimento = float(item.comprimento)
                    item.comprimento = round(item.comprimento, 3)
                else:
                    item.comprimento = item_data['comprimento']
                    item.largura = item_data['largura']

                if (produto.tipo_medida == 'linear'):
                    item.largura = 1
                    item.medida_final = item.largura * item.comprimento
                elif (produto.tipo_medida == 'm2') :

                    item.medida_final = item.largura * item.comprimento
                elif (produto.tipo_medida == 'pc'):
                    item.medida_final = item.quantidade

                item.medida_final = round((item.medida_final),3)
                print(item.nome, item.medida_final)

                item.preco = item_data['preco']
                item.preco = round(float(item.preco),2)
                print(item.preco)

                item.valor_base = (item.preco + adicional) / item.medida_final
                item.valor_base = round((item.valor_base), 2)

                item.preco = (frete_item/item.quantidade) + adicional + item.preco
                item.preco = round(item.preco,2)
                item.total_item = (item.quantidade * item.preco)
                item.total_item = round((item.total_item), 2)
                print(item.comprimento, item.largura, item.nome)
                item.largura = "{:.3f}".format(item.largura)
                item.comprimento = "{:.3f}".format(item.comprimento)

                item.save()


                quantidade_itens = quantidade_itens - 1

                if frete > 0:
                    frete = frete - frete_item
                    if quantidade_itens > 0:
                        frete_item = frete / quantidade_itens
                        frete_item = round(frete_item, 2)


            return JsonResponse({'message': 'Dados recebidos com sucesso'})
        except Exception as e:
            return JsonResponse({'error': 'Erro ao processar os itens', 'message': str(e)}, status=400)


def lista_pedidos (request):

    campos = Pedidos._meta.get_fields()
    campos_selecionados = request.POST.getlist('campos_selecionados')
    subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]
    lista_pedidos = Pedidos.objects.annotate(nome_cliente=Subquery(subquery))



    campos_para_exibir = []
    for campo in campos_selecionados:
        if campo.name in campos_selecionados:
            campos_para_exibir.append(campo)

    context = {
        'campos' : campos,
        'pedidos': Pedidos.objects.all(),
        'itens_pedidos': Itens_Pedido.objects.all(),
        'clientes': Cliente.objects.all(),
        'lista_pedidos' : lista_pedidos,
    }

    return render (request, 'list_pedidos.html', context)

def sincroniza_notas(request):
    data_atual = date.today()
    mes_atual = data_atual.month
    pedidos_mes_atual = Pedidos.objects.filter(datapedido__month=mes_atual, datapedido__year=data_atual.year)
    pedidos_nota = pedidos_mes_atual
    tpedidos_nota = pedidos_mes_atual.count()
    print (pedidos_nota)
    print (tpedidos_nota)
    novo_registro = RegistrosLog()
    novo_registro.colaborador = Colaborador.objects.filter(
        nome=request.user.username).first()  # Relaciona o log com o colaborador
    novo_registro.data = timezone.now()  # Usa timezone.now() para obter a data e hora atuais
    novo_registro.operacao = 'sincroniza notas executado'  # Formata a string corretamente

    novo_registro.save()

    for pedido in pedidos_nota:
        if pedido.numero_nfe is None and pedido.flag_nfe == 1:
            if (pedido.empresa_pedido.nome == 'TEM INDUSTRIA E COMERCIO DE TAPETES LTDA'):
                token = "OK0V8A4AKMOi5ywOFOUljMR8Ew4S3Imf"
            else:
                token = "Dzrt7d64Eq30parVmw1xT8swfZ2uMejS"
            url_consulta = "https://api.focusnfe.com.br/v2/nfe/"
            ref = str(pedido.numero) + str(pedido.operacao_nota)
            completa = 'completa=1'

            print(url_consulta + ref)
            response = requests.get(url_consulta + ref, params=completa, auth=(token, ""))
            print(response)
            if response.status_code == 200:
                json_response = response.json()
                print(json_response)
                status_nfe = json_response['status']
                if status_nfe == 'erro_autorizacao':
                    pass
                else:
                    xml = json_response['caminho_xml_nota_fiscal']
                    nfe = json_response['caminho_danfe'] if json_response['caminho_danfe'] else None
                    nfe_numero = json_response['numero'] if json_response['numero'] else None
                    url_focus = 'https://api.focusnfe.com.br'
                    data_emissao = json_response['requisicao_nota_fiscal']['data_emissao']
                    data_emissao = datetime.fromisoformat(data_emissao)
                    pedido.dataEmissao_nfe = data_emissao.date()
                    pedido.caminho_xml = url_focus + xml
                    pedido.caminho_nfe = url_focus + nfe
                    pedido.numero_nfe = nfe_numero
                    pedido.valor_nota = pedido.valorTotal
                    if pedido.flag_nfe_antecipada > 0:
                        pedido.status = 'Producao'
                    else:
                        pass

                    pedido.save()
            else:
                pass


    return HttpResponse('Notas Sincronizadas com sucesso!')

def sincroniza_nota_unica(request, numero_pedido):


    print(numero_pedido)
    pedido = Pedidos.objects.filter(numero=numero_pedido).first()
    print (pedido)

    if (pedido.empresa_pedido.nome == 'TEM INDUSTRIA E COMERCIO DE TAPETES LTDA'):
        token = "OK0V8A4AKMOi5ywOFOUljMR8Ew4S3Imf"
    else:
        token = "Dzrt7d64Eq30parVmw1xT8swfZ2uMejS"
    url_consulta = "https://api.focusnfe.com.br/v2/nfe/"
    ref = str(pedido.numero) + str(pedido.operacao_nota)
    completa = 'completa=1'

    print(url_consulta + ref)
    response = requests.get(url_consulta + ref, params=completa, auth=(token, ""))
    print(response)
    if response.status_code == 200:
        json_response = response.json()
        print(json_response)
        status_nfe = json_response['status']
        if status_nfe == 'erro_autorizacao':
            return HttpResponse('Erro Autorização')
        else:
            xml = json_response['caminho_xml_nota_fiscal']
            nfe = json_response['caminho_danfe'] if json_response['caminho_danfe'] else None
            nfe_numero = json_response['numero'] if json_response['numero'] else None
            url_focus = 'https://api.focusnfe.com.br'
            data_emissao = json_response['requisicao_nota_fiscal']['data_emissao']
            data_emissao = datetime.fromisoformat(data_emissao)
            pedido.dataEmissao_nfe = data_emissao.date()
            pedido.caminho_xml = url_focus + xml
            pedido.caminho_nfe = url_focus + nfe
            pedido.numero_nfe = nfe_numero
            pedido.valor_nota = pedido.valorTotal
            pedido.referencia_nfe = ref
            if pedido.flag_nfe_antecipada > 0:
                pedido.status = 'Producao'
            else:
                pass

            pedido.save()
    else:
        return HttpResponse('Nota não encontrada!')

    return HttpResponse('Nota Sincronizadas com sucesso!')

def remover_caracteres_especiais(telefone):
    # Usando regex para substituir caracteres especiais por vazio
    telefone_sem_especiais = re.sub(r'[^0-9]', '', telefone)
    return telefone_sem_especiais

@login_required
def meuspedidos (request):

    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)

    if colaborador.empresa.flag_pagamento == 1:
        return redirect ('/dologin/')
    elif not colaborador:
        return redirect('/dologin/')
    else:
        pass

    total_pedidos_vendedor = 0
    total_comissao = 0
    total_venda = 0
    print(colaborador.funcao)


    campos = Pedidos._meta.get_fields()
    campos_selecionados = request.POST.getlist('campos_selecionados')
    subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]
    bairros = Bairro.objects.all()
    vendedores = Colaborador.objects.filter(Q(funcao='Vendedor Interno') | Q(funcao='Vendedor Externo'))
    empresas = Empresas.objects.all()
    fretes = Fornecedor.objects.filter(tipo_fornecedor=1)

    status_choices = Pedidos.STATUS_CHOICES
    bancos = Bancos.objects.all()
    resultados = None
    total_mes_anterior = 0
    comissao_mes_anterior = 0


    campos_para_exibir = []
    for campo in campos_selecionados:
        if campo.name in campos_selecionados:
            campos_para_exibir.append(campo)

    if request.method == 'GET':
        total_vendido_grf = 0
        total_vendido_tem = 0
        resultados_tem =None
        resultados_grf = None
        numero_pedido = request.GET.get('numero_pedido_filter')
        nome_cliente = request.GET.get('nome_cliente_filter')
        documento_cliente = request.GET.get('documento_cliente_filter')
        documento_cliente = ''.join(filter(str.isdigit, documento_cliente)) if documento_cliente else None
        status_pedido = request.GET.get('status_pedido_filter')
        vendedor = request.GET.get('vendedor_pedido_filter')
        localidade_entrega = request.GET.get('localidade_pedido_filter')
        datapedido_inicial = request.GET.get('filter_data_pedido_inicial')
        datapedido_final = request.GET.get('filter_data_pedido_final')
        dataprevista_inicial = request.GET.get('filter_data_pedido_entrega_inicial')
        dataprevista_final = request.GET.get('filter_data_pedido_entrega_final')
        empresa_pedido = request.GET.get ('filter_empresa')
        opcao_bonus = request.GET.get('bonificado')
        opcao_nfe = request.GET.get('nfe')
        opcao_reparo = request.GET.get('reparo')
        valor_pedido = request.GET.get ('valorpedido')
        comprador = request.GET.get ('nome_comprador_filter')




        pedidos_busca = Pedidos.objects.none()


        if numero_pedido or valor_pedido or nome_cliente or documento_cliente or status_pedido or vendedor or localidade_entrega\
                or (datapedido_inicial and datapedido_final)\
                or (dataprevista_inicial and dataprevista_final) or empresa_pedido or opcao_bonus or opcao_nfe or opcao_reparo or comprador:
            query = Q()

            if numero_pedido:
                query &= Q(numero__icontains=numero_pedido)

            if documento_cliente:
                cliente_pedido = Cliente.objects.filter(documento=documento_cliente).first()

                if cliente_pedido:
                    query &= Q(cliente_pedido=cliente_pedido)
                else:
                    return HttpResponse (f'Nenhum cliente encontrado com esse documento: {documento_cliente}')

            if nome_cliente:
                cliente_pedido = Cliente.objects.filter(nome__icontains=nome_cliente).first()
                if cliente_pedido:
                    query &= Q(cliente_pedido__nome__icontains=nome_cliente)
                else:
                    cliente_pedido = None

            if comprador:
                    query &= Q(comprador_nome__icontains=comprador)


            if status_pedido:
                query &= Q(status__icontains=status_pedido)

            if valor_pedido:
                valor_pedido = re.sub(r'^R\$|\.', '', valor_pedido).replace(',', '.')
                query &= Q(valorTotal=valor_pedido)

            if vendedor:
                query &=  (Q(Vendedor__icontains=vendedor) | Q(Vendedor2__icontains=vendedor))

            if localidade_entrega:
                bairro_nome, bairro_cidade_uf = localidade_entrega.split(', ')
                bairro = Bairro.objects.get(nome=bairro_nome, cidade=bairro_cidade_uf.split('-')[0],
                                            uf=bairro_cidade_uf.split('-')[1])
                query &= Q(cidade_entrega__icontains=bairro.cidade, uf_entrega__icontains=bairro.uf,
                           bairro_entrega__icontains=bairro.nome)

            if datapedido_inicial and datapedido_final:
                query &= Q(datapedido__range=[datapedido_inicial, datapedido_final])

            if dataprevista_inicial and dataprevista_final:
                query &= Q(prazo_entrega__range=[dataprevista_inicial, dataprevista_final])

            if empresa_pedido:
                empresa_pedido = Empresas.objects.get(nome=empresa_pedido)
                query &= Q(empresa_pedido=empresa_pedido)

            if opcao_bonus:
                query &= Q(valor_repasse__gt=0)

            if opcao_nfe:
                query &= Q(numero_nfe__isnull=False)

            if opcao_reparo:
                ids_reparos = Reparos.objects.all().values_list('pedido_id', flat=True)

                # Filtrando os pedidos que têm IDs presentes na lista de reparos
                query &= Q(id__in=ids_reparos)

            if 'Vendedor' in colaborador.funcao:
                pedidos_vendedor = Pedidos.objects.filter(Vendedor=colaborador.nome)

                pedidos_busca = pedidos_vendedor.filter(query).order_by('status')

            elif 'Expedicao' in colaborador.funcao:
                lista_pedidos_producao = Pedidos.objects.filter(
                    Q(status='Producao') | Q(status='Expedicao') | Q(status='Liberado Para Entrega') |
                    Q(status='Liberado Para Entrega') | Q(status='Em Rota de Entrega') | Q(status='Entregue'))
                pedidos_busca = lista_pedidos_producao.filter(query)

            else:
                total_vendido_grf = 0
                total_vendido_tem = 0
                pedidos_grf = Pedidos.objects.filter(empresa_pedido=1)
                pedidos_grf_mes = pedidos_grf.annotate(month=TruncMonth('datapedido')).values('month').annotate(
                    total_venda=Sum('valorTotal')).order_by('month')
                resultado_grf = {}
                for item in pedidos_grf_mes:
                    mes = item['month'].strftime('%Y-%m')
                    total = item['total_venda']
                    resultado_grf[mes] = total
                resultados_grf = pedidos_grf.values('datapedido__year', 'datapedido__month') \
                    .annotate(total_venda=Sum('valorTotal')) \
                    .order_by('datapedido__year', 'datapedido__month')



                for pedido in pedidos_grf:
                    total_vendido_grf += pedido.valorTotal

                pedidos_tem = Pedidos.objects.filter(empresa_pedido=2)
                pedidos_tem_mes = pedidos_tem.annotate(month=TruncMonth('datapedido')).values('month').annotate(
                    total_venda=Sum('valorTotal')).order_by('month')
                resultado_tem = {}
                for item in pedidos_tem_mes:
                    mes = item['month'].strftime('%Y-%m')
                    total = item['total_venda']
                    resultado_tem[mes] = total
                resultados_tem = pedidos_tem.values('datapedido__year', 'datapedido__month') \
                    .annotate(total_venda=Sum('valorTotal')) \
                    .order_by('datapedido__year', 'datapedido__month')



                for pedido in pedidos_tem:
                    total_vendido_tem += pedido.valorTotal

                subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]
                pedidos_busca = Pedidos.objects.filter(query).annotate(nome_cliente=Subquery(subquery))

            context ={
                'pedidos': pedidos_busca,
                'campos': campos_para_exibir,
                'itens_pedidos': Itens_Pedido.objects.all(),
                'clientes': Cliente.objects.all(),
                'bairros': bairros,
                'vendedores': vendedores,
                'status_choices': status_choices,
                'empresas': empresas,
                'colaborador':colaborador,
                'fretes': fretes,
                'total_pedidos_vendedor': total_pedidos_vendedor,
                'total_comissao': total_comissao,
                'resultados': resultados,
                'total_venda': total_venda,
                'resultados_grf': resultados_grf if resultados_grf else None,
                'total_vendido_grf': total_vendido_grf if total_vendido_grf else None,
                'resultados_tem': resultados_tem if resultados_tem else None,
                'total_vendido_tem': total_vendido_tem if total_vendido_tem else None,
                'total_mes_anterior': total_mes_anterior if total_mes_anterior else None
            }
            return render(request, 'meuspedidos.html', context)
        else:
            total_pedidos_vendedor=0
            total_comissao=0
            total_venda=0
            resultados_grf = None
            resultados_tem = None
            total_vendido_grf = 0
            total_vendido_tem = 0
            total_mes = 0
            data_atual = date.today()
            mes_atual = data_atual.month
            ano_atual = data_atual.year
            cpedidos_mes = 0
            comissao_parcial = 0

            if 'Vendedor Interno' in colaborador.funcao:
                pedidos_vendedor = Pedidos.objects.filter(Vendedor=colaborador.nome)
                lista_pedidos = pedidos_vendedor.order_by('-datapedido')
                total_pedidos_vendedor = lista_pedidos.order_by('-datapedido').count()
                comissoes_por_mes = pedidos_vendedor.annotate(month=TruncMonth('datapedido')).values('month').annotate(
                    total_comissao=Sum('vendedor_comissaov')).order_by('month')
                pedidos_mes_vendedor = pedidos_vendedor.filter(datapedido__month=mes_atual, datapedido__year=ano_atual)
                comissao_parcial = 0
                total_mes_anterior = 0
                comissao_mes_anterior =0

                if mes_atual == 1:
                    ano = ano_atual - 1
                    mes_anterior = 12
                else:
                    mes_anterior = mes_atual-1
                    ano = ano_atual
                pedidos_mes_anterior = pedidos_vendedor.filter(datapedido__month=mes_anterior, datapedido__year=ano)

                for pedido in pedidos_mes_anterior:
                    total_mes_anterior += pedido.valorTotal
                    comissao_mes_anterior += pedido.vendedor_comissaov

                for pedido in pedidos_mes_vendedor:
                    total_mes += pedido.valorTotal
                    comissao_parcial += pedido.vendedor_comissaov

                cpedidos_mes = pedidos_mes_vendedor.count()



                resultado = {}
                for item in comissoes_por_mes:

                    mes = item['month'].strftime('%Y-%m')
                    total = item['total_comissao']
                    resultado[mes] = total
                for lista_pedido in lista_pedidos:
                    total_comissao += lista_pedido.vendedor_comissaov
                    total_venda += lista_pedido.valorTotal

                resultados = pedidos_vendedor.values('datapedido__year', 'datapedido__month') \
                    .annotate(total_comissao=Sum('vendedor_comissaov')) \
                    .order_by('datapedido__year', 'datapedido__month')


            elif 'Vendedor Externo' in colaborador.funcao:
                pedidos_vendedor = Pedidos.objects.filter(Vendedor2=colaborador.nome)
                lista_pedidos = pedidos_vendedor.order_by('-datapedido')
                total_pedidos_vendedor = lista_pedidos.order_by('-datapedido').count()
                comissoes_por_mes = pedidos_vendedor.annotate(month=TruncMonth('datapedido')).values('month').annotate(
                    total_comissao=Sum('vendedor2_comissaov')).order_by('month')
                comissao_parcial =0
                pedidos_mes_vendedor = pedidos_vendedor.filter(datapedido__month=mes_atual, datapedido__year=ano_atual)

                total_mes_anterior = 0
                comissao_mes_anterior = 0

                pedidos_mes_anterior = pedidos_vendedor.filter(datapedido__month=mes_atual - 1,
                                                               datapedido__year=ano_atual)
                for pedido in pedidos_mes_anterior:
                    total_mes_anterior += pedido.valorTotal
                    comissao_mes_anterior += pedido.vendedor2_comissaov

                for pedido in pedidos_mes_vendedor:
                    total_mes += pedido.valorTotal
                    comissao_parcial += pedido.vendedor2_comissaov

                cpedidos_mes = pedidos_mes_vendedor.count()

                resultado = {}




                for item in comissoes_por_mes:
                    mes = item['month'].strftime('%Y-%m')
                    total = item['total_comissao']
                    resultado[mes] = total
                for lista_pedido in lista_pedidos:
                    total_comissao += lista_pedido.vendedor_comissaov
                    total_venda += lista_pedido.valorTotal

                resultados = pedidos_vendedor.values('datapedido__year', 'datapedido__month') \
                    .annotate(total_comissao=Sum('vendedor2_comissaov')) \
                    .order_by('datapedido__year', 'datapedido__month')


            elif 'Gerente' in colaborador.funcao:
                total_vendido_grf = 0
                total_vendido_tem = 0
                pedidos_grf = Pedidos.objects.filter(empresa_pedido=1, datapedido__year=ano_atual)
                pedidos_grf_mes = pedidos_grf.annotate(month=TruncMonth('datapedido')).values('month').annotate(
                    total_venda=Sum('valorTotal')).order_by('month')
                resultado_grf = {}
                for item in pedidos_grf_mes:
                    mes = item['month'].strftime('%Y-%m')
                    total = item['total_venda']
                    resultado_grf[mes] = total
                resultados_grf = pedidos_grf.values('datapedido__year', 'datapedido__month') \
                    .annotate(total_venda=Sum('valorTotal')) \
                    .order_by('datapedido__year', 'datapedido__month')



                for pedido in pedidos_grf:
                    total_vendido_grf += pedido.valorTotal

                pedidos_tem = Pedidos.objects.filter(empresa_pedido=2, datapedido__year=ano_atual )
                pedidos_tem_mes = pedidos_tem.annotate(month=TruncMonth('datapedido')).values('month').annotate(
                    total_venda=Sum('valorTotal')).order_by('month')
                resultado_tem = {}
                for item in pedidos_tem_mes:
                    mes = item['month'].strftime('%Y-%m')
                    total = item['total_venda']
                    resultado_tem[mes] = total
                resultados_tem = pedidos_tem.values('datapedido__year', 'datapedido__month') \
                    .annotate(total_venda=Sum('valorTotal')) \
                    .order_by('datapedido__year', 'datapedido__month')



                for pedido in pedidos_tem:
                    total_vendido_tem += pedido.valorTotal

                subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]



                lista_pedidos = Pedidos.objects.annotate(nome_cliente=Subquery(subquery)).order_by('-datapedido',
                                                                                                   '-numero')
                order_dict = {
                    'duplicado': 1,
                    'Nfe Antecipada': 3,
                    'Confirma os Valores': 4,
                    'Pendente Financeiro': 5,
                    'Conferir Item': 2,
                    'Pendente Logística': 6,
                    'Producao': 7,
                    'Expedicao': 8,
                    'Liberado para Entrega': 9,
                    'Liberado Coleta': 10,
                    'Liberado Retirada Cliente': 11,
                    'Em Rota de Entrega': 12,

                }

                # Crie um campo "order" baseado no status usando o Case
                lista_pedidos = lista_pedidos.annotate(
                    order=Case(
                        *[When(status=status, then=Value(order)) for status, order in order_dict.items()],
                        default=Value(8),  # Define um valor padrão maior para os outros status
                        output_field=IntegerField(),
                    )
                ).order_by('flag_impressao', 'order', 'status')




            elif 'Expedicao' in colaborador.funcao:
                lista_pedidos_producao = Pedidos.objects.filter(
                    Q(status='Producao') | Q(status='Expedicao') | Q(status='Liberado Para Entrega') |
                    Q(status='Liberado Para Entrega') | Q(status='Em Rota de Entrega') | Q(status='Entregue'))
                lista_pedidos = lista_pedidos_producao.order_by('status')


            else:
                subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]
                lista_pedidos = Pedidos.objects.annotate(nome_cliente=Subquery(subquery)).order_by('-datapedido', '-numero')
                order_dict = {
                    'Nfe Antecipada':1,
                    'Conferir Item': 2,
                    'Pendente Financeiro': 3,
                    'Pendente Logística': 4,
                    'Producao': 5,
                    'Expedicao':6,
                    'Liberado para Entrega':7,
                    'Em Rota de Entrega': 8

                }

                # Crie um campo "order" baseado no status usando o Case
                lista_pedidos = lista_pedidos.annotate(
                    order=Case(
                        *[When(status=status, then=Value(order)) for status, order in order_dict.items()],
                        default=Value(8),  # Define um valor padrão maior para os outros status
                        output_field=IntegerField(),
                    )
                ).order_by('order', 'status')
                lista_pedidos = lista_pedidos.filter(datapedido__month=mes_atual, datapedido__year=ano_atual)

            context = {

                'pedidos':lista_pedidos.filter(Q(datapedido__month=mes_atual,
                                                 datapedido__year=ano_atual)
                                               |Q(status = 'Nfe Antecipada')
                                               |Q(status = 'Confirma os Valores')
                                               |Q(status = 'Pendente Financeiro')
                                               |Q(status = 'Pendente Logistica')
                                               |Q(status = 'Conferir Item')) if colaborador.funcao=='Gerente' else lista_pedidos,
                'itens_pedidos': Itens_Pedido.objects.all(),
                'campos': campos_para_exibir,
                'clientes': Cliente.objects.all(),
                'bairros': bairros,
                'vendedores': vendedores,
                'status_choices': status_choices,
                'empresas': empresas,
                'colaborador': colaborador,
                'fretes': fretes,
                'bancos':bancos,
                'total_pedidos_vendedor': total_pedidos_vendedor,
                'total_comissao': total_comissao,
                'resultados': resultados,
                'total_venda': total_venda,
                'resultados_grf': resultados_grf if resultados_grf else None,
                'total_vendido_grf': total_vendido_grf if total_vendido_grf else None,
                'resultados_tem': resultados_tem if resultados_tem else None,
                'total_vendido_tem': total_vendido_tem if total_vendido_tem else None,
                'total_mes': total_mes,
                'cpedidos_mes': cpedidos_mes,
                'comissao_parcial': comissao_parcial,
                'total_mes_anterior': total_mes_anterior if total_mes_anterior else None,
                'comissao_mes_anterior': comissao_mes_anterior if comissao_mes_anterior else None

            }

            return render(request, 'meuspedidos.html', context)

def meuspedidos_new (request):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    total_pedidos_vendedor = 0
    total_comissao = 0
    total_venda = 0
    print(colaborador.funcao)


    campos = Pedidos._meta.get_fields()
    campos_selecionados = request.POST.getlist('campos_selecionados')
    subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]
    bairros = Bairro.objects.all()
    vendedores = Colaborador.objects.filter(Q(funcao='Vendedor Interno') | Q(funcao='Vendedor Externo'))
    empresas = Empresas.objects.all()
    fretes = Fornecedor.objects.filter(tipo_fornecedor=1)

    status_choices = Pedidos.STATUS_CHOICES
    bancos = Bancos.objects.all()
    resultados = None
    total_mes_anterior = 0
    comissao_mes_anterior = 0


    campos_para_exibir = []
    for campo in campos_selecionados:
        if campo.name in campos_selecionados:
            campos_para_exibir.append(campo)

    if request.method == 'GET':
        total_vendido_grf = 0
        total_vendido_tem = 0
        resultados_tem =None
        resultados_grf = None
        numero_pedido = request.GET.get('numero_pedido_filter')
        nome_cliente = request.GET.get('nome_cliente_filter')
        documento_cliente = request.GET.get('documento_cliente_filter')
        documento_cliente = ''.join(filter(str.isdigit, documento_cliente)) if documento_cliente else None
        status_pedido = request.GET.get('status_pedido_filter')
        vendedor = request.GET.get('vendedor_pedido_filter')
        localidade_entrega = request.GET.get('localidade_pedido_filter')
        datapedido_inicial = request.GET.get('filter_data_pedido_inicial')
        datapedido_final = request.GET.get('filter_data_pedido_final')
        dataprevista_inicial = request.GET.get('filter_data_pedido_entrega_inicial')
        dataprevista_final = request.GET.get('filter_data_pedido_entrega_final')
        empresa_pedido = request.GET.get ('filter_empresa')
        opcao_bonus = request.GET.get('bonificado')
        opcao_nfe = request.GET.get('nfe')
        opcao_reparo = request.GET.get('reparo')
        valor_pedido = request.GET.get ('valorpedido')
        comprador = request.GET.get ('nome_comprador_filter')




        pedidos_busca = Pedidos.objects.none()


        if numero_pedido or valor_pedido or nome_cliente or documento_cliente or status_pedido or vendedor or localidade_entrega\
                or (datapedido_inicial and datapedido_final)\
                or (dataprevista_inicial and dataprevista_final) or empresa_pedido or opcao_bonus or opcao_nfe or opcao_reparo or comprador:
            query = Q()

            if numero_pedido:
                query &= Q(numero__icontains=numero_pedido)

            if documento_cliente:
                cliente_pedido = Cliente.objects.filter(documento=documento_cliente).first()

                if cliente_pedido:
                    query &= Q(cliente_pedido=cliente_pedido)
                else:
                    return HttpResponse (f'Nenhum cliente encontrado com esse documento: {documento_cliente}')

            if nome_cliente:
                cliente_pedido = Cliente.objects.filter(nome__icontains=nome_cliente).first()
                if cliente_pedido:
                    query &= Q(cliente_pedido__nome__icontains=nome_cliente)
                else:
                    cliente_pedido = None

            if comprador:
                    query &= Q(comprador_nome__icontains=comprador)


            if status_pedido:
                query &= Q(status__icontains=status_pedido)

            if valor_pedido:
                valor_pedido = re.sub(r'^R\$|\.', '', valor_pedido).replace(',', '.')
                query &= Q(valorTotal=valor_pedido)

            if vendedor:
                query &=  (Q(Vendedor__icontains=vendedor) | Q(Vendedor2__icontains=vendedor))

            if localidade_entrega:
                bairro_nome, bairro_cidade_uf = localidade_entrega.split(', ')
                bairro = Bairro.objects.get(nome=bairro_nome, cidade=bairro_cidade_uf.split('-')[0],
                                            uf=bairro_cidade_uf.split('-')[1])
                query &= Q(cidade_entrega__icontains=bairro.cidade, uf_entrega__icontains=bairro.uf,
                           bairro_entrega__icontains=bairro.nome)

            if datapedido_inicial and datapedido_final:
                query &= Q(datapedido__range=[datapedido_inicial, datapedido_final])

            if dataprevista_inicial and dataprevista_final:
                query &= Q(prazo_entrega__range=[dataprevista_inicial, dataprevista_final])

            if empresa_pedido:
                empresa_pedido = Empresas.objects.get(nome=empresa_pedido)
                query &= Q(empresa_pedido=empresa_pedido)

            if opcao_bonus:
                query &= Q(valor_repasse__gt=0)

            if opcao_nfe:
                query &= Q(numero_nfe__isnull=False)

            if opcao_reparo:
                ids_reparos = Reparos.objects.all().values_list('pedido_id', flat=True)

                # Filtrando os pedidos que têm IDs presentes na lista de reparos
                query &= Q(id__in=ids_reparos)

            if 'Vendedor' in colaborador.funcao:
                pedidos_vendedor = Pedidos.objects.filter(Vendedor=colaborador.nome)

                pedidos_busca = pedidos_vendedor.filter(query).order_by('status')

            elif 'Expedicao' in colaborador.funcao:
                lista_pedidos_producao = Pedidos.objects.filter(
                    Q(status='Producao') | Q(status='Expedicao') | Q(status='Liberado Para Entrega') |
                    Q(status='Liberado Para Entrega') | Q(status='Em Rota de Entrega') | Q(status='Entregue'))
                pedidos_busca = lista_pedidos_producao.filter(query)

            else:
                total_vendido_grf = 0
                total_vendido_tem = 0
                pedidos_grf = Pedidos.objects.filter(empresa_pedido=1)
                pedidos_grf_mes = pedidos_grf.annotate(month=TruncMonth('datapedido')).values('month').annotate(
                    total_venda=Sum('valorTotal')).order_by('month')
                resultado_grf = {}
                for item in pedidos_grf_mes:
                    mes = item['month'].strftime('%Y-%m')
                    total = item['total_venda']
                    resultado_grf[mes] = total
                resultados_grf = pedidos_grf.values('datapedido__year', 'datapedido__month') \
                    .annotate(total_venda=Sum('valorTotal')) \
                    .order_by('datapedido__year', 'datapedido__month')



                for pedido in pedidos_grf:
                    total_vendido_grf += pedido.valorTotal

                pedidos_tem = Pedidos.objects.filter(empresa_pedido=2)
                pedidos_tem_mes = pedidos_tem.annotate(month=TruncMonth('datapedido')).values('month').annotate(
                    total_venda=Sum('valorTotal')).order_by('month')
                resultado_tem = {}
                for item in pedidos_tem_mes:
                    mes = item['month'].strftime('%Y-%m')
                    total = item['total_venda']
                    resultado_tem[mes] = total
                resultados_tem = pedidos_tem.values('datapedido__year', 'datapedido__month') \
                    .annotate(total_venda=Sum('valorTotal')) \
                    .order_by('datapedido__year', 'datapedido__month')



                for pedido in pedidos_tem:
                    total_vendido_tem += pedido.valorTotal

                subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]
                pedidos_busca = Pedidos.objects.filter(query).annotate(nome_cliente=Subquery(subquery))

            context ={
                'pedidos': pedidos_busca,
                'campos': campos_para_exibir,
                'itens_pedidos': Itens_Pedido.objects.all(),
                'clientes': Cliente.objects.all(),
                'bairros': bairros,
                'vendedores': vendedores,
                'status_choices': status_choices,
                'empresas': empresas,
                'colaborador':colaborador,
                'fretes': fretes,
                'total_pedidos_vendedor': total_pedidos_vendedor,
                'total_comissao': total_comissao,
                'resultados': resultados,
                'total_venda': total_venda,
                'resultados_grf': resultados_grf if resultados_grf else None,
                'total_vendido_grf': total_vendido_grf if total_vendido_grf else None,
                'resultados_tem': resultados_tem if resultados_tem else None,
                'total_vendido_tem': total_vendido_tem if total_vendido_tem else None,
                'total_mes_anterior': total_mes_anterior if total_mes_anterior else None
            }
            return render(request, 'meuspedidos.html', context)
        else:
            total_pedidos_vendedor=0
            total_comissao=0
            total_venda=0
            resultados_grf = None
            resultados_tem = None
            total_vendido_grf = 0
            total_vendido_tem = 0
            total_mes = 0
            data_atual = date.today()
            mes_atual = data_atual.month
            ano_atual = data_atual.year
            cpedidos_mes = 0
            comissao_parcial = 0

            if 'Vendedor Interno' in colaborador.funcao:
                pedidos_vendedor = Pedidos.objects.filter(Vendedor=colaborador.nome)
                lista_pedidos = pedidos_vendedor.order_by('-datapedido')
                total_pedidos_vendedor = lista_pedidos.order_by('-datapedido').count()
                comissoes_por_mes = pedidos_vendedor.annotate(month=TruncMonth('datapedido')).values('month').annotate(
                    total_comissao=Sum('vendedor_comissaov')).order_by('month')
                pedidos_mes_vendedor = pedidos_vendedor.filter(datapedido__month=mes_atual, datapedido__year=ano_atual)
                comissao_parcial = 0
                total_mes_anterior = 0
                comissao_mes_anterior =0

                pedidos_mes_anterior = pedidos_vendedor.filter(datapedido__month=mes_atual-1, datapedido__year=ano_atual)
                for pedido in pedidos_mes_anterior:
                    total_mes_anterior += pedido.valorTotal
                    comissao_mes_anterior += pedido.vendedor_comissaov

                for pedido in pedidos_mes_vendedor:
                    total_mes += pedido.valorTotal
                    comissao_parcial += pedido.vendedor_comissaov

                cpedidos_mes = pedidos_mes_vendedor.count()



                resultado = {}
                for item in comissoes_por_mes:

                    mes = item['month'].strftime('%Y-%m')
                    total = item['total_comissao']
                    resultado[mes] = total
                for lista_pedido in lista_pedidos:
                    total_comissao += lista_pedido.vendedor_comissaov
                    total_venda += lista_pedido.valorTotal

                resultados = pedidos_vendedor.values('datapedido__year', 'datapedido__month') \
                    .annotate(total_comissao=Sum('vendedor_comissaov')) \
                    .order_by('datapedido__year', 'datapedido__month')


            elif 'Vendedor Externo' in colaborador.funcao:
                pedidos_vendedor = Pedidos.objects.filter(Vendedor2=colaborador.nome)
                lista_pedidos = pedidos_vendedor.order_by('-datapedido')
                total_pedidos_vendedor = lista_pedidos.order_by('-datapedido').count()
                comissoes_por_mes = pedidos_vendedor.annotate(month=TruncMonth('datapedido')).values('month').annotate(
                    total_comissao=Sum('vendedor2_comissaov')).order_by('month')
                comissao_parcial =0
                pedidos_mes_vendedor = pedidos_vendedor.filter(datapedido__month=mes_atual, datapedido__year=ano_atual)

                total_mes_anterior = 0
                comissao_mes_anterior = 0

                pedidos_mes_anterior = pedidos_vendedor.filter(datapedido__month=mes_atual - 1,
                                                               datapedido__year=ano_atual)
                for pedido in pedidos_mes_anterior:
                    total_mes_anterior += pedido.valorTotal
                    comissao_mes_anterior += pedido.vendedor2_comissaov

                for pedido in pedidos_mes_vendedor:
                    total_mes += pedido.valorTotal
                    comissao_parcial += pedido.vendedor2_comissaov

                cpedidos_mes = pedidos_mes_vendedor.count()

                resultado = {}




                for item in comissoes_por_mes:
                    mes = item['month'].strftime('%Y-%m')
                    total = item['total_comissao']
                    resultado[mes] = total
                for lista_pedido in lista_pedidos:
                    total_comissao += lista_pedido.vendedor_comissaov
                    total_venda += lista_pedido.valorTotal

                resultados = pedidos_vendedor.values('datapedido__year', 'datapedido__month') \
                    .annotate(total_comissao=Sum('vendedor2_comissaov')) \
                    .order_by('datapedido__year', 'datapedido__month')


            elif 'Gerente' in colaborador.funcao:
                total_vendido_grf = 0
                total_vendido_tem = 0
                pedidos_grf = Pedidos.objects.filter(empresa_pedido=1, datapedido__year=ano_atual)
                pedidos_grf_mes = pedidos_grf.annotate(month=TruncMonth('datapedido')).values('month').annotate(
                    total_venda=Sum('valorTotal')).order_by('month')
                resultado_grf = {}
                for item in pedidos_grf_mes:
                    mes = item['month'].strftime('%Y-%m')
                    total = item['total_venda']
                    resultado_grf[mes] = total
                resultados_grf = pedidos_grf.values('datapedido__year', 'datapedido__month') \
                    .annotate(total_venda=Sum('valorTotal')) \
                    .order_by('datapedido__year', 'datapedido__month')



                for pedido in pedidos_grf:
                    total_vendido_grf += pedido.valorTotal

                pedidos_tem = Pedidos.objects.filter(empresa_pedido=2, datapedido__year=ano_atual )
                pedidos_tem_mes = pedidos_tem.annotate(month=TruncMonth('datapedido')).values('month').annotate(
                    total_venda=Sum('valorTotal')).order_by('month')
                resultado_tem = {}
                for item in pedidos_tem_mes:
                    mes = item['month'].strftime('%Y-%m')
                    total = item['total_venda']
                    resultado_tem[mes] = total
                resultados_tem = pedidos_tem.values('datapedido__year', 'datapedido__month') \
                    .annotate(total_venda=Sum('valorTotal')) \
                    .order_by('datapedido__year', 'datapedido__month')



                for pedido in pedidos_tem:
                    total_vendido_tem += pedido.valorTotal

                subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]



                lista_pedidos = Pedidos.objects.annotate(nome_cliente=Subquery(subquery)).order_by('-datapedido',
                                                                                                   '-numero')
                order_dict = {
                    'duplicado': 1,
                    'Nfe Antecipada': 2,
                    'Confirma os Valores': 3,
                    'Pendente Financeiro': 4,
                    'Pendente Logística': 5,
                    'Producao': 6,
                    'Expedicao': 7,
                    'Liberado para Entrega': 8,
                    'Liberado Coleta': 8,
                    'Liberado Retirada Cliente': 8,
                    'Em Rota de Entrega': 11,

                }

                # Crie um campo "order" baseado no status usando o Case
                lista_pedidos = lista_pedidos.annotate(
                    order=Case(
                        *[When(status=status, then=Value(order)) for status, order in order_dict.items()],
                        default=Value(8),  # Define um valor padrão maior para os outros status
                        output_field=IntegerField(),
                    )
                ).order_by('order', 'status')




            elif 'Expedicao' in colaborador.funcao:
                lista_pedidos_producao = Pedidos.objects.filter(
                    Q(status='Producao') | Q(status='Expedicao') | Q(status='Liberado Para Entrega') |
                    Q(status='Liberado Para Entrega') | Q(status='Em Rota de Entrega') | Q(status='Entregue'))
                lista_pedidos = lista_pedidos_producao.order_by('status')


            else:
                subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]
                lista_pedidos = Pedidos.objects.annotate(nome_cliente=Subquery(subquery)).order_by('-datapedido', '-numero')
                order_dict = {
                    'Nfe Antecipada':1,
                    'Pendente Financeiro': 2,
                    'Pendente Logística': 3,
                    'Producao': 4,
                    'Expedicao':5,
                    'Liberado para Entrega':6,
                    'Em Rota de Entrega': 7,

                }

                # Crie um campo "order" baseado no status usando o Case
                lista_pedidos = lista_pedidos.annotate(
                    order=Case(
                        *[When(status=status, then=Value(order)) for status, order in order_dict.items()],
                        default=Value(8),  # Define um valor padrão maior para os outros status
                        output_field=IntegerField(),
                    )
                ).order_by('order', 'status')
                lista_pedidos = lista_pedidos.filter(datapedido__month=mes_atual, datapedido__year=ano_atual)

            context = {

                'pedidos':lista_pedidos.filter(Q(datapedido__month=mes_atual, datapedido__year=ano_atual )|Q(status = 'Nfe Antecipada')|Q(status = 'Confirma os Valores')|Q(status = 'Pendente Financeiro')|Q(status = 'Pendente Logistica')) if colaborador.funcao=='Gerente' else lista_pedidos,
                'itens_pedidos': Itens_Pedido.objects.all(),
                'campos': campos_para_exibir,
                'clientes': Cliente.objects.all(),
                'bairros': bairros,
                'vendedores': vendedores,
                'status_choices': status_choices,
                'empresas': empresas,
                'colaborador': colaborador,
                'fretes': fretes,
                'bancos':bancos,
                'total_pedidos_vendedor': total_pedidos_vendedor,
                'total_comissao': total_comissao,
                'resultados': resultados,
                'total_venda': total_venda,
                'resultados_grf': resultados_grf if resultados_grf else None,
                'total_vendido_grf': total_vendido_grf if total_vendido_grf else None,
                'resultados_tem': resultados_tem if resultados_tem else None,
                'total_vendido_tem': total_vendido_tem if total_vendido_tem else None,
                'total_mes': total_mes,
                'cpedidos_mes': cpedidos_mes,
                'comissao_parcial': comissao_parcial,
                'total_mes_anterior': total_mes_anterior if total_mes_anterior else None,
                'comissao_mes_anterior': comissao_mes_anterior if comissao_mes_anterior else None

            }

            return render(request, 'pedidos/painel_pedidos.html', context)
@login_required
def consultar_pedidos(request, numero_pedido):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)

    if 'Produção' in colaborador.funcao:
        return redirect('/painel/')

    pedido = Pedidos.objects.get(numero=numero_pedido)
    cliente_id = pedido.cliente_pedido.id
    cliente = Cliente.objects.get(id=cliente_id)
    itens = Itens_Pedido.objects.filter(pedido=numero_pedido)
    resumo_pagamento = Contas_Receber.objects.filter(pedido=numero_pedido)

    # Definir o contexto com os dados do pedido
    context = {
        'pedido': pedido,
        'cliente': cliente,
        'itens': itens,
        'parcelas': resumo_pagamento,
        'colaborador': colaborador
    }

    return render(request, 'consulta_pedido.html', context)


def editar_pedidos(request, numero_pedido):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    if 'Vendedor' or 'Gerente' or 'Financeiro' in colaborador.funcao:
        pedido = Pedidos.objects.get(numero=numero_pedido)
        pagamento_entrega_atual = pedido.pagamento_entrega
        cliente_id = pedido.cliente_pedido.id
        Contas = Contas_Receber.objects.filter(pedido=numero_pedido)
        itens = Itens_Pedido.objects.filter(pedido=numero_pedido)
        resumo_pagamento = Contas_Receber.objects.filter(pedido=numero_pedido)
        status_choices = Pedidos.STATUS_CHOICES
        localidades = Bairro.objects.all()
        vendedores = Colaborador.objects.filter(Q(funcao='Vendedor Interno') | Q(funcao='Vendedor Externo'))
        produtos = Produtos.objects.all()
        produtos_revenda = Produtos_Revenda.objects.all()
        fornecedores = Fornecedor.objects.filter(tipo_fornecedor='1')
        bonificados = Fornecedor.objects.filter(tipo_fornecedor='8')
        tipos_pagamento = TipoPgto.objects.all()
        lista_produtos = []

        for produto in produtos:
            lista_produtos.append(produto)
        for produto in produtos_revenda:
            lista_produtos.append(produto)

        momentos = [
            'Antecipado',
            'Entrega',
            'Faturado'
        ]

        if request.method == 'POST':



            opcao_endereco = request.POST.get('endereco-opcao')
            if opcao_endereco != 'manter_endereco':
                pedido.cep_entrega = request.POST.get('cep-entrega')
                pedido.endereco_entrega = request.POST.get('endereco-entrega')
                pedido.numero_end_entrega = request.POST.get('numero-endereco-entrega')
                pedido.complemento = request.POST.get('complemento_endereco')
                pedido.uf_entrega = request.POST.get('uf-entrega')
                pedido.cidade_entrega = request.POST.get('cidade-entrega')
                pedido.bairro_entrega = request.POST.get('bairro-entrega')
                pedido.ponto_referencia = request.POST.get('ponto_referencia_entrega')

            pedido.tipo_entrega = request.POST.get('opcao_frete')
            print (pedido.tipo_entrega)

            flag_pedido = request.POST.get('flag_pedido')
            pedido.Vendedor = request.POST.get('vendedor1')
            pedido.Vendedor2 = request.POST.get('vendedor2')

            pedido.valorTotalParcial = 0.0
            frete_valor = request.POST.get('frete')
            frete_valor = re.sub(r'^R\$|\.', '', frete_valor).replace(',', '.')
            pedido.frete_valor = round(float(frete_valor), 2) if frete_valor else 0
            pedido.prazo_entrega = request.POST.get('data_entrega')

            itens = Itens_Pedido.objects.filter(pedido=numero_pedido)
            for item in itens:
                pedido.valorTotalParcial += float (item.total_item)

            tipo_frete = request.POST.get('opcao_frete')
            nfe_opcao = request.POST.get('pedido-nfe-opcao')
            pedido.retirada_flag = 0
            if nfe_opcao == 'sim':
                pedido.nfe_flag = 1
                pedido.status = 'Nfe Antecipada'
            else:
                pedido.nfe_flag = 0

            if tipo_frete == 'Transportadora':
                pedido.frete_flag = 1
            else:
                pedido.frete_flag = 0
            if tipo_frete == 'retirada':
                pedido.retirada_flag = 1
            else:
                pass

            pedido.informacoes_adicionais =  informacoes_complementares = request.POST.get('informacoes_complementares')
            repasse = request.POST.get('acerto')
            repasse = re.sub(r'^R\$|\.', '', repasse).replace(',', '.') if repasse else None
            repasse = float(repasse)

            bonificado=''
            if pedido.valor_repasse != repasse and repasse > 0 :
                pedido.valor_repasse = float(repasse)
                conta_pagar = Contas_Pagar.objects.filter(pedido=pedido.numero,descricao='BONIFICACAO').first()
                if conta_pagar :
                    conta_pagar.valor = float(repasse)
                    bonificado = request.POST.get('bonificado')
                    conta_pagar.fornecedor = Fornecedor.objects.filter(nome=bonificado).first()
                    pedido.bonificado = conta_pagar.fornecedor
                    conta_pagar.status = 'A Vencer'
                    conta_pagar.save()

                else:
                    conta_pagar = Contas_Pagar()
                    conta_pagar.valor = float(repasse)
                    bonificado = request.POST.get('bonificado')
                    conta_pagar.fornecedor = Fornecedor.objects.filter(nome=bonificado).first()
                    pedido.bonificado = conta_pagar.fornecedor
                    conta_pagar.status = 'A Vencer'
                    prazo_entrega_datetime = datetime.strptime(pedido.prazo_entrega, '%Y-%m-%d')
                    data_vencimento = prazo_entrega_datetime + timedelta(days=30)
                    conta_pagar.data_vencimento = data_vencimento
                    conta_pagar.numero_parcela = 1
                    conta_pagar.total_parcelas = 1
                    conta_pagar.descricao = "BONIFICACAO"
                    conta_pagar.save()

            if repasse > 0:
                pedido.valorTotal = repasse + pedido.valorTotalParcial
            else:
                pedido.valorTotal = pedido.valorTotalParcial

            pedido.Vendedor = request.POST.get('vendedor1')
            pedido.vendedor_comissao = request.POST.get('comissao1')
            vendedor2 = request.POST.get('vendedor2')

            if vendedor2:
                pedido.Vendedor2 = vendedor2
                if pedido.vendedor_comissao == '1':
                    pedido.vendedor2_comissao = 1
                elif pedido.vendedor_comissao == '3':
                    pedido.vendedor2_comissao = 3
                elif pedido.vendedor_comissao == '5':
                    pedido.vendedor2_comissao = 7
                elif pedido.vendedor_comissao == '6':
                    pedido.vendedor2_comissao = 10
                elif pedido.vendedor_comissao == '7':
                    pedido.vendedor2_comissao = 15

            pedido.observacao = request.POST.get('observacao')
            pedido.save()

            redirect_url = f'/geradocpedido/{numero_pedido}/'

            return redirect(redirect_url)

        cliente=Cliente.objects.get(id=cliente_id)
        context = {
            'pedido': pedido,
            'cliente': cliente,
            'itens': itens,
            'parcelas': resumo_pagamento,
            'status': status_choices,
            'localidades': localidades,
            'vendedores': vendedores,
            'produtos': lista_produtos,
            'fornecedores': fornecedores,
            'tipos_pagamento': tipos_pagamento,
            'momentos':momentos,
            'colaborador':colaborador,
            'bonificados': bonificados
        }

        return render(request, 'edita_pedido.html', context)
    else:
        return redirect ('/painel/')

def editar_pedidos_b(request, numero_pedido):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    if 'Vendedor' or 'Gerente' or 'Financeiro' in colaborador.funcao:
        pedido = Pedidos.objects.get(numero=numero_pedido)
        pagamento_entrega_atual = pedido.pagamento_entrega
        cliente_id = pedido.cliente_pedido.id
        Contas = Contas_Receber.objects.filter(pedido=numero_pedido)
        itens = Itens_Pedido.objects.filter(pedido=numero_pedido)
        resumo_pagamento = Contas_Receber.objects.filter(pedido=numero_pedido)
        status_choices = Pedidos.STATUS_CHOICES
        localidades = Bairro.objects.all()
        vendedores = Colaborador.objects.filter(Q(funcao='Vendedor Interno') | Q(funcao='Vendedor Externo'))
        produtos = Produtos.objects.all()
        produtos_revenda = Produtos_Revenda.objects.all()
        fornecedores = Fornecedor.objects.filter(tipo_fornecedor='1')
        bonificados = Fornecedor.objects.filter(tipo_fornecedor='8')
        tipos_pagamento = TipoPgto.objects.all()
        lista_produtos = []

        for produto in produtos:
            lista_produtos.append(produto)
        for produto in produtos_revenda:
            lista_produtos.append(produto)

        momentos = [
            'Antecipado',
            'Entrega',
            'Faturado'
        ]

        if request.method == 'POST':



            opcao_endereco = request.POST.get('endereco-opcao')

            pedido.cep_entrega = request.POST.get('cep-entrega')
            pedido.endereco_entrega = request.POST.get('logradouro-entrega')
            pedido.numero_end_entrega = request.POST.get('numero-endereco-entrega')
            pedido.complemento = request.POST.get('complemento_endereco')
            pedido.uf_entrega = request.POST.get('uf-entrega')
            pedido.cidade_entrega = request.POST.get('cidade-entrega')
            pedido.bairro_entrega = request.POST.get('bairro-entrega')
            pedido.ponto_referencia_entrega = request.POST.get('ponto_referencia_entrega')


            flag_pedido = request.POST.get('flag_pedido')
            pedido.Vendedor = request.POST.get('vendedor1')
            pedido.Vendedor2 = request.POST.get('vendedor2')

            pedido.valorTotalParcial = 0.0
            frete_valor = request.POST.get('frete')
            frete_valor = re.sub(r'^R\$|\.', '', frete_valor).replace(',', '.')
            pedido.frete_valor = round(float(frete_valor), 2) if frete_valor else 0
            pedido.prazo_entrega = request.POST.get('data_entrega')

            itens = Itens_Pedido.objects.filter(pedido=numero_pedido)
            for item in itens:
                pedido.valorTotalParcial += float (item.total_item)

            tipo_frete = request.POST.get('opcao_frete')
            nfe_opcao = request.POST.get('pedido-nfe-opcao')
            pedido.retirada_flag = 0
            if nfe_opcao == 'sim':
                pedido.nfe_flag = 1
            else:
                pedido.nfe_flag = 0

            if tipo_frete == 'Transportadora':
                pedido.frete_flag = 1
            else:
                pedido.frete_flag = 0
            if tipo_frete == 'retirada':
                pedido.retirada_flag = 1
            else:
                pass

            pedido.tipo_entrega=tipo_frete
            pedido.informacoes_adicionais =  informacoes_complementares = request.POST.get('informacoes_complementares')
            repasse = request.POST.get('acerto')
            repasse = re.sub(r'^R\$|\.', '', repasse).replace(',', '.') if repasse else None
            repasse = float(repasse)

            bonificado=''
            if pedido.valor_repasse != repasse and repasse > 0 :
                pedido.valor_repasse = float(repasse)
                conta_pagar = Contas_Pagar.objects.filter(pedido=pedido.numero,descricao='BONIFICACAO').first()
                if conta_pagar :
                    conta_pagar.valor = float(repasse)
                    bonificado = request.POST.get('bonificado')
                    conta_pagar.fornecedor = Fornecedor.objects.filter(nome=bonificado).first()
                    pedido.bonificado = conta_pagar.fornecedor
                    conta_pagar.status = 'A Vencer'
                    conta_pagar.save()

                else:
                    conta_pagar = Contas_Pagar()
                    conta_pagar.valor = float(repasse)
                    bonificado = request.POST.get('bonificado')
                    conta_pagar.fornecedor = Fornecedor.objects.filter(nome=bonificado).first()
                    pedido.bonificado = conta_pagar.fornecedor
                    conta_pagar.status = 'A Vencer'
                    prazo_entrega_datetime = datetime.strptime(pedido.prazo_entrega, '%Y-%m-%d')
                    data_vencimento = prazo_entrega_datetime + timedelta(days=30)
                    conta_pagar.data_vencimento = data_vencimento
                    conta_pagar.numero_parcela = 1
                    conta_pagar.total_parcelas = 1
                    conta_pagar.descricao = "BONIFICACAO"
                    conta_pagar.save()



            pedido.Vendedor = request.POST.get('vendedor1')
            pedido.vendedor_comissao = request.POST.get('comissao1')
            vendedor2 = request.POST.get('vendedor2')

            if vendedor2:
                pedido.Vendedor2 = vendedor2
                if pedido.vendedor_comissao == '1':
                    pedido.vendedor2_comissao = 1
                elif pedido.vendedor_comissao == '3':
                    pedido.vendedor2_comissao = 3
                elif pedido.vendedor_comissao == '5':
                    pedido.vendedor2_comissao = 7
                elif pedido.vendedor_comissao == '6':
                    pedido.vendedor2_comissao = 10
                elif pedido.vendedor_comissao == '7':
                    pedido.vendedor2_comissao = 15

            if repasse > 0:
                pedido.valorTotal = pedido.valorTotalParcial
                pedido.vendedor_comissaov = float(pedido.valorTotal - repasse) * float(pedido.vendedor_comissao) / 100
                if vendedor2:
                    pedido.vendedor2_comissaov = float(pedido.valorTotal - repasse) * float(
                        pedido.vendedor2_comissao) / 100
            else:
                pedido.valorTotal = pedido.valorTotalParcial
                pedido.vendedor_comissaov = float(pedido.valorTotal) * float(pedido.vendedor_comissao) / 100
                if vendedor2:
                    pedido.vendedor2_comissaov = float(pedido.valorTotal) * float(
                        pedido.vendedor2_comissao) / 100


            pedido.observacao = request.POST.get('observacao')
            pedido.save()

            redirect_url = f'/geradocpedido/{numero_pedido}/'

            return redirect(redirect_url)

        cliente=Cliente.objects.get(id=cliente_id)
        context = {
            'pedido': pedido,
            'cliente': cliente,
            'itens': itens,
            'parcelas': resumo_pagamento,
            'status': status_choices,
            'localidades': localidades,
            'vendedores': vendedores,
            'produtos': lista_produtos,
            'fornecedores': fornecedores,
            'tipos_pagamento': tipos_pagamento,
            'momentos':momentos,
            'colaborador':colaborador,
            'bonificados': bonificados
        }

        return render(request, 'edita_pedido_basico.html', context)
    else:
        return redirect ('/painel/')

@login_required
def delete_pedido(request):
    if request.method == 'POST':
        pedido_numero = request.POST.get('pedido_numero')
        pedido = Pedidos.objects.filter(numero=pedido_numero).first()
        itens = Itens_Pedido.objects.filter(pedido=pedido_numero)
        conta = Contas_Receber.objects.filter(pedido=pedido_numero)
        layout = Layout.objects.filter (numero_pedido=pedido).first()
        colaborador = Colaborador.objects.filter(usuario__username=request.user.username).first()


        if conta:
            Contas_Receber.objects.filter(pedido=pedido_numero).delete()
            print ('#### Conta a Receber deletado: ####', conta)
        else:
            print ('Conta nao encontrada')

        if itens:
            Itens_Pedido.objects.filter(pedido=pedido_numero).delete()
            print('Itens Pedido Deletado ', itens)
        else:
            print ('Itens Nao Encontrados')

        if layout:
            layout.delete()
            print('Layout Deletado')
        else:
            pass

        if pedido:

           novo_log(pedido.numero, colaborador, f'Exclusao de Pedido {pedido.numero}')
           pedido.delete()
           print('Pedido Deletado ', pedido)
           return JsonResponse({'mensagem': 'Pedido deletado com sucesso!'})
        else:
            print ('Pedido Não Encontrado')
            return JsonResponse({'mensagem': 'Falha ao deletar o pedido. Contactar o SUPORTE'})






    # Se a requisição não for do tipo POST, você pode retornar um erro
    return JsonResponse({'mensagem': 'Método inválido.'}, status=400)

@login_required
def delete_cliente(request):
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente_id')

        try:

            Cliente.objects.filter(id=cliente_id).delete()


            return JsonResponse({'mensagem': 'Cliente Excluido com sucesso!'})
        except:
            return JsonResponse({'mensagem': 'Falha ao deletar o pedido.'})

    # Se a requisição não for do tipo POST, você pode retornar um erro
    return JsonResponse({'mensagem': 'Método inválido.'}, status=400)

def confirma_pedido(request):
    if request.method == 'POST':
        pedido_numero = request.POST.get('pedido_numero')
        pedido = Pedidos.objects.filter(numero=pedido_numero).first()
        data_atual = date.today()
        prazo_entrega = pedido.prazo_entrega
        diferenca_dias = (prazo_entrega - data_atual).days
        if pedido.status == 'Confirma os Valores':
            try:


               if pedido.flag_nfe_antecipada == 1:
                   pedido.status = 'Nfe Antecipada'
                   pedido.save()
                   return JsonResponse({'mensagem': f'Pedido Atualizado com sucesso! Novo Status do Pedido é: {pedido.status}'})

               elif pedido.pagamento_antecipado > 0:
                   pedido.status = 'Pendente Financeiro'
                   pedido.save()
                   return JsonResponse(
                       {'mensagem': f'Pedido Atualizado com sucesso! Novo Status do Pedido é: {pedido.status}'})

               elif diferenca_dias < 15:
                   pedido.status  = 'Pendente Logistica'
                   pedido.save()
                   return JsonResponse(
                       {'mensagem': f'Pedido Atualizado com sucesso! Novo Status do Pedido é: {pedido.status}'})

               else:
                   pedido.status = 'Producao'
                   pedido.save()
                   return JsonResponse({'mensagem': f'Pedido Atualizado com sucesso! Novo Status do Pedido é: {pedido.status}'})
            except:
                return JsonResponse({'mensagem': 'Falha ao atualizar o pedido.'})
        else:
            return JsonResponse({'mensagem': f'Erro! O status do Pedido é: {pedido.status}'})


    # Se a requisição não for do tipo POST, você pode retornar um erro
    return JsonResponse({'mensagem': 'Método inválido.'}, status=400)

def gerar_recibo(request, numero_pedido):
    pedido = Pedidos.objects.filter(numero=numero_pedido).first()
    if pedido.status == 'Entregue':
        registro_entrega = Registra.objects.filter(numero_registro=numero_pedido).first()
    else:
        registro_entrega = None
    itens =  Itens_Pedido.objects.filter(pedido=numero_pedido)
    data_atual = date.today()
    data = data_atual.isoformat()
    hora_atual = datetime.now()

    # Passo 2: Subtrair 3 horas da hora atual
    nova_hora = hora_atual - timedelta(hours=3)
    hora_atual = nova_hora.strftime('%H:%M:%S')

    context = {
            'pedido': pedido,
            'itens': itens,
            'data':data,
            'hora_atual':hora_atual,
            'registro_entrega':registro_entrega
        }
    novo_registro = RegistrosLog()
    novo_registro.colaborador = Colaborador.objects.filter(
        nome=request.user.username).first()  # Relaciona o log com o colaborador
    novo_registro.data = timezone.now()  # Usa timezone.now() para obter a data e hora atuais
    novo_registro.operacao = f'Gerado Recibo: {pedido.numero}'  # Formata a string corretamente

    novo_registro.save()

    return render(request, 'modelo_recibo.html', context)
def gerar_docpedido(request, numero_pedido):
    pedido = Pedidos.objects.get(numero=numero_pedido)
    itens =  Itens_Pedido.objects.filter(pedido=numero_pedido)
    data_atual = date.today()
    contas = Contas_Receber.objects.filter(pedido=numero_pedido)
    data = data_atual.isoformat()
    hora_atual = datetime.now()

    # Passo 2: Subtrair 3 horas da hora atual
    nova_hora = hora_atual - timedelta(hours=3)
    hora_atual = nova_hora.strftime('%H:%M:%S')

    context = {
            'pedido': pedido,
            'itens': itens,
            'data':data,
            'hora_atual':hora_atual,
            'contas': contas
        }
    return render(request, 'recibo_tem.html', context)

def gerar_guia_reparo (request, numero_pedido):


    pedido = Pedidos.objects.get(numero=numero_pedido)
    reparo = Reparos.objects.filter(pedido=pedido).first()
    itens =  Itens_Pedido.objects.filter(pedido=numero_pedido)
    data_atual = date.today()
    contas = Contas_Receber.objects.filter(pedido=numero_pedido)
    data = data_atual.isoformat()
    hora_atual = datetime.now()

    # Passo 2: Subtrair 3 horas da hora atual
    nova_hora = hora_atual - timedelta(hours=3)
    hora_atual = nova_hora.strftime('%H:%M:%S')

    context = {
            'pedido': pedido,
            'itens': itens,
            'data':data,
            'hora_atual':hora_atual,
            'contas': contas,
            'reparo_observacao': reparo.observacao
        }
    return render(request, 'guia_reparo.html', context)

def documento_guia_reparo (request, numero_pedido):


    pedido = Pedidos.objects.filter(numero=numero_pedido).first()
    guia = Guia_Reparo.objects.filter(pedido=pedido).first()
    itens =  Reparos.objects.filter(guia=guia)
    itens_pedido = Itens_Pedido.objects.filter(pedido=pedido.numero)
    data_atual = date.today()

    data = data_atual.isoformat()
    hora_atual = datetime.now()

    # Passo 2: Subtrair 3 horas da hora atual
    nova_hora = hora_atual - timedelta(hours=3)
    hora_atual = nova_hora.strftime('%H:%M:%S')

    context = {
            'pedido': pedido,
            'itens_reparo': itens,
            'itens_pedido': itens_pedido,
            'guia': guia,
            'data':data,
            'hora_atual':hora_atual

        }
    return render(request, 'guiareparo/guia_reparo_documento.html', context)
@login_required
def painel_vendas(request):
    campos = Pedidos._meta.get_fields()
    campos_selecionados = request.POST.getlist('campos_selecionados')
    subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]
    bairros = Bairro.objects.all()
    vendedores = Colaborador.objects.filter(Q(funcao='Vendedor Interno') | Q(funcao='Vendedor Externo'))
    status_choices = Pedidos.STATUS_CHOICES

    campos_para_exibir = []
    for campo in campos_selecionados:
        if campo.name in campos_selecionados:
            campos_para_exibir.append(campo)

    if request.method == 'GET':
        numero_pedido = request.GET.get('numero_pedido_filter')
        documento_cliente = request.GET.get('documento_cliente_filter')
        documento_cliente = ''.join(filter(str.isdigit, documento_cliente)) if documento_cliente else None
        status_pedido = request.GET.get('status_pedido_filter')
        vendedor = request.GET.get('vendedor_pedido_filter')
        localidade_entrega = request.GET.get('localidade_pedido_filter')
        datapedido_inicial = request.GET.get('filter_data_pedido_inicial')
        datapedido_final = request.GET.get('filter_data_pedido_final')
        dataprevista_inicial = request.GET.get('filter_data_pedido_entrega_inicial')
        dataprevista_final = request.GET.get('filter_data_pedido_entrega_final')

        pedidos_busca = Pedidos.objects.none()

        if numero_pedido or documento_cliente or status_pedido or vendedor or localidade_entrega or (
                datapedido_inicial and datapedido_final) or (dataprevista_inicial and dataprevista_final):
            query = Q()

            if numero_pedido:
                query &= Q(numero__icontains=numero_pedido)

            if documento_cliente:
                cliente_pedido = Cliente.objects.filter(documento=documento_cliente)
                if cliente_pedido.exists():
                    cliente_id = cliente_pedido[0].id
                else:
                    cliente_id = None
                query &= Q(cliente_pedido__icontains=cliente_id)

            if status_pedido:
                query &= Q(status__icontains=status_pedido)

            if vendedor:
                query &= Q(Vendedor__icontains=vendedor)

            if localidade_entrega:
                bairro_nome, bairro_cidade_uf = localidade_entrega.split(', ')
                bairro = Bairro.objects.get(nome=bairro_nome, cidade=bairro_cidade_uf.split('-')[0],
                                            uf=bairro_cidade_uf.split('-')[1])
                query &= Q(cidade_entrega__icontains=bairro.cidade, uf_entrega__icontains=bairro.uf,
                           bairro_entrega__icontains=bairro.nome)

            if datapedido_inicial and datapedido_final:
                query &= Q(datapedido__range=[datapedido_inicial, datapedido_final])

            if dataprevista_inicial and dataprevista_final:
                query &= Q(prazo_entrega__range=[dataprevista_inicial, dataprevista_final])

            subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]
            pedidos_busca = Pedidos.objects.filter(query).annotate(nome_cliente=Subquery(subquery))




            context = {
                'pedidos': pedidos_busca,
                'itens_pedidos': Itens_Pedido.objects.all(),
                'clientes': Cliente.objects.all(),
                'bairros': bairros,
                'vendedores': vendedores,
                'status_choices': status_choices
            }
            return render(request, 'meuspedidos.html', context)
        else:
            nome_usuario = request.user.username
            vendedor = Colaborador.objects.get(usuario__username=nome_usuario)
            lista_pedidos_vendedor = Pedidos.objects.filter(Vendedor=vendedor.nome)


            subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]
            lista_pedidos = lista_pedidos_vendedor.annotate(nome_cliente=Subquery(subquery))
            lista_pedidos = lista_pedidos.order_by('status', 'datapedido')


            context = {

                'pedidos': lista_pedidos,
                'itens_pedidos': Itens_Pedido.objects.all(),
                'clientes': Cliente.objects.all(),
                'bairros': bairros,
                'vendedores': vendedores,
                'status_choices': status_choices
            }

            return render(request, 'meuspedidos.html', context)

def painel_producao(request):
    campos = Pedidos._meta.get_fields()
    campos_selecionados = request.POST.getlist('campos_selecionados')
    subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]
    bairros = Bairro.objects.all()
    colaboradores_ativos =  Colaborador.objects.filter (status__descricao='ATIVO')

    vendedores = colaboradores_ativos.filter(Q(funcao='Vendedor Interno') | Q(funcao='Vendedor Externo'))
    status_choices = Pedidos.STATUS_CHOICES

    campos_para_exibir = []
    for campo in campos_selecionados:
        if campo.name in campos_selecionados:
            campos_para_exibir.append(campo)

    if request.method == 'GET':
        numero_pedido = request.GET.get('numero_pedido_filter')
        documento_cliente = request.GET.get('documento_cliente_filter')
        documento_cliente = ''.join(filter(str.isdigit, documento_cliente)) if documento_cliente else None
        status_pedido = request.GET.get('status_pedido_filter')
        vendedor = request.GET.get('vendedor_pedido_filter')
        localidade_entrega = request.GET.get('localidade_pedido_filter')
        datapedido_inicial = request.GET.get('filter_data_pedido_inicial')
        datapedido_final = request.GET.get('filter_data_pedido_final')
        dataprevista_inicial = request.GET.get('filter_data_pedido_entrega_inicial')
        dataprevista_final = request.GET.get('filter_data_pedido_entrega_final')

        pedidos_busca = Pedidos.objects.filter()

        if numero_pedido or documento_cliente or status_pedido or vendedor or localidade_entrega or (
                datapedido_inicial and datapedido_final) or (dataprevista_inicial and dataprevista_final):
            query = Q()

            if numero_pedido:
                query &= Q(numero__icontains=numero_pedido)

            if documento_cliente:
                cliente_pedido = Cliente.objects.filter(documento=documento_cliente)
                if cliente_pedido.exists():
                    cliente_id = cliente_pedido[0].id
                else:
                    cliente_id = None
                query &= Q(cliente_pedido__icontains=cliente_id)

            if status_pedido:
                query &= Q(status__icontains=status_pedido)

            if vendedor:
                query &= Q(Vendedor__icontains=vendedor)

            if localidade_entrega:
                bairro_nome, bairro_cidade_uf = localidade_entrega.split(', ')
                bairro = Bairro.objects.get(nome=bairro_nome, cidade=bairro_cidade_uf.split('-')[0],
                                            uf=bairro_cidade_uf.split('-')[1])
                query &= Q(cidade_entrega__icontains=bairro.cidade, uf_entrega__icontains=bairro.uf,
                           bairro_entrega__icontains=bairro.nome)

            if datapedido_inicial and datapedido_final:
                query &= Q(datapedido__range=[datapedido_inicial, datapedido_final])

            if dataprevista_inicial and dataprevista_final:
                query &= Q(prazo_entrega__range=[dataprevista_inicial, dataprevista_final])

            subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]
            pedidos_busca = Pedidos.objects.filter(query).annotate(nome_cliente=Subquery(subquery))

            zonas = set()
            bairros = set()
            cidades = set()

            for p in pedidos_busca:
                zona = p.zona_entrega
                cidade = p.cidade_entrega
                bairro = p.bairro_entrega

                if zona not in zonas:
                    if zona:
                        zonas.add(zona)
                if cidade not in cidades:
                    cidades.add(cidade)
                if bairro not in bairros:
                    bairros.add(bairro)

            pedidos_busca = pedidos_busca.filter(Q(status='Producao')|Q(status='Expedicao')|Q(status='Liberado Para Entrega')|
                                                            Q(status='Liberado Para Entrega')|Q(status='Em Rota de Entrega')|Q(status='Entregue'))

            context = {
                'pedidos': pedidos_busca,
                'itens_pedidos': Itens_Pedido.objects.all(),
                'clientes': Cliente.objects.all(),

                'vendedores': vendedores,
                'status_choices': status_choices,
                'zonas': zonas,
                'cidades': cidades,
                'bairros': bairros
            }
            return render(request, 'meuspedidos.html', context)
        else:
            nome_usuario = request.user.username
            usuario = Colaborador.objects.get(usuario__username=nome_usuario)
            lista_pedidos_producao = Pedidos.objects.filter(Q(status='Producao')|Q(status='Expedicao')|Q(status='Liberado Para Entrega')|
                                                            Q(status='Liberado Para Entrega')|Q(status='Em Rota de Entrega')|Q(status='Entregue'))


            subquery = Cliente.objects.filter(id=OuterRef('cliente_pedido')).values('nome')[:1]

            lista_pedidos_producao = lista_pedidos_producao.order_by('status')
            lista_produzindo = Pedidos.objects.filter(status='Produzindo')
            total_produzindo = lista_produzindo.len() if lista_produzindo else 0



            context = {

                'pedidos': lista_pedidos_producao,
                'itens_pedidos': Itens_Pedido.objects.all(),
                'clientes': Cliente.objects.all(),
                'bairros': bairros,
                'vendedores': vendedores,
                'status_choices': status_choices
            }

            return render(request, 'meuspedidos.html', context)



# Define a função de chave de ordenação fora da função de consulta
def atualiza_inadimplencia():
    today = date.today()
    data_check = today - timedelta(days=1)
    mes_atual = today.month
    ano_atual = today.year
    # VERIFICA AS CONTAS DO MES COM OS PEDIDOS
    contas_mes = Contas_Receber.objects.filter(data_vencimento__month=mes_atual, data_vencimento__year=ano_atual)
    for conta in contas_mes:

        numero_pedido = conta.pedido
        pedido = Pedidos.objects.filter(numero=numero_pedido)
        if pedido:
            pass
        else:
            conta.delete()

    #VERIFICAS CONTAS VENCIDAS

    contas_vencidas = Contas_Receber.objects.filter(data_vencimento__lt=today)
    for conta in contas_vencidas:
        if conta.data_vencimento:

            if not conta.data_pagamento:

                if conta.tipo_pgto == 'ACERTO':

                    conta.status_conta = 'A Vencer'
                    conta.save()
                elif conta.tipo_pgto == 'BOLETO' and (
                        conta.status_cobranca == 'OVERDUE' or conta.status_cobranca is None):
                    conta.status_conta = 'Vencida'

                    conta.save()
                elif conta.tipo_pgto == 'BOLETO' and conta.status_cobranca == 'PENDING':
                    conta.status_conta = 'Vencida a Validar'

                    conta.save()

                elif conta.tipo_pgto != 'BOLETO':
                    conta.status_conta = 'Vencida'
                    conta.save()
                else:
                    conta.status_conta = 'Vencida a Validar'

                conta.save()
            else:
                pass

        else:
            pass

    novo_check = Inadimplencia ()
    novo_check.data = date.today()
    novo_check.flag = 1
    novo_check.save()

@login_required
def condicoes_pagamento (request):

    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()
    if colaborador.funcao == 'Gerente':

        if request.method == 'POST':
            nova_condicao = Condicoes_Pagamentos()
            nova_condicao.descricao = request.POST.get('descricao')
            nova_condicao.save()


        context = {
            'condicoes_pagamento': Condicoes_Pagamentos.objects.all()
        }

        return render (request, 'condicoes_pagamento.html', context)

    else:

        return HttpResponse ('Usuário Sem Permissão')

@login_required
def contas_receber (request):

    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    if colaborador.empresa.flag_pagamento == 1:
        return redirect('/dologin/')
    elif not colaborador:
        return redirect('/dologin/')
    else:
        pass


    if 'Vendedor' in colaborador.funcao:
        return redirect('/painel/')


    if 'Financeiro' or 'Gerente' in colaborador.funcao:


        today = date.today()
        #data_check = today - timedelta(days=1)
        mes_atual = today.month
        ano_atual = today.year
        contas = Contas_Receber.objects.filter(data_vencimento__month = mes_atual, data_vencimento__year = ano_atual)

        check = Inadimplencia.objects.filter(data=today).first()

        if check is None:
            print('contas checkando')
            atualiza_inadimplencia()
        else:
            print ('contas checkadas já')
            pass

        pedidos = Pedidos.objects.all()
        status_conta = Contas_Receber.STATUS_CONTA
        formas_pagamento = TipoPgto.objects.all()
        empresas = Empresas.objects.all()
        mes_atual = date.today()
        mes_atual = mes_atual.month
        ano_atual = today.year

        consultas = Contas_Receber.objects.filter(data_vencimento__month=mes_atual, data_vencimento__year = ano_atual ).order_by('data_vencimento')

        order_dict = {
            'Vencida': 1,
            'Vencida a Validar': 2,
            'A Vencer': 3,
            'Pago': 4,
        }
        consultas = consultas.annotate(
            order=Case(
                *[When(status_conta=status_conta, then=Value(order)) for status_conta, order in order_dict.items()],
                default=Value(4),  # Define um valor padrão maior para os outros status
                output_field=IntegerField(),
            )
        ).order_by('data_vencimento', 'order', 'status_conta' )



        tipos_pagamento = TipoPgto.objects.all()

        context = {
            'contas': contas,
            'pedidos': pedidos,
            'consultas': consultas,
            'status_conta': status_conta,
            'empresas': empresas,
            'tipos_pagamento': tipos_pagamento

        }
        if request.method == 'GET':
            numero_pedido = request.GET.get('numero_pedido_filter')
            documento_cliente = request.GET.get('documento_cliente_filter')
            documento_cliente = ''.join(filter(str.isdigit, documento_cliente)) if documento_cliente else None
            datavencimento_inicial = request.GET.get('filter_vencimento_inicial')
            datavencimento_final = request.GET.get('filter_vencimento_final')
            datapagamento_inicial = request.GET.get('filter_data_pagamento_inicial')
            datapagamento_final = request.GET.get('filter_data_pagamento_final')
            datapedido_inicial = request.GET.get('filter_data_pedido_inicial')
            datapedido_final = request.GET.get('filter_data_pedido_final')
            empresa_pedido = request.GET.get('filter_empresa')
            status_conta = request.GET.get('status_pagamento_filter')
            forma_pagamento = request.GET.get('forma_pagamento_filter')
            nome_cliente = request.GET.get('nome_cliente_filter')
            valor_parcela = request.GET.get('valor_parcela_filter')

            pedidos_busca = Pedidos.objects.none()
            contas_busca = consultas

            if numero_pedido or valor_parcela or nome_cliente or documento_cliente or status_conta or forma_pagamento or empresa_pedido or (
                    datavencimento_inicial and datavencimento_final) or (
                    datapagamento_inicial and datapagamento_final) or (
                    datapedido_inicial and datapedido_final):
                pedidos_busca = Pedidos.objects.all()

                query = Q()

                if empresa_pedido:
                    empresa = Empresas.objects.filter(nome=empresa_pedido).first()
                    query &=Q(pedido_conta__empresa_pedido=empresa)



                if nome_cliente:
                    cliente = Cliente.objects.filter(nome__icontains=nome_cliente).first()
                    if cliente:
                        query &= Q(cliente__nome__icontains=nome_cliente)
                    else:
                        pass

                if valor_parcela:
                    valor_parcela = re.sub(r'^R\$|\.', '', valor_parcela).replace(',', '.')
                    query &= Q(valor=valor_parcela)

                if numero_pedido:
                    #pedidos_busca= pedidos_busca.filter(numero__icontains=numero_pedido)

                    #numeros_pedidos_filtrados = pedidos_busca.values_list('numero', flat=True)

                    query &= Q(Q(pedido__icontains=numero_pedido)|Q(descricao__icontains=numero_pedido))

                if datapedido_inicial and datapedido_final:
                    query &= Q(
                        pedido_conta__datapedido__range=[datapedido_inicial, datapedido_final])

                if documento_cliente:
                    cliente_pedido = Cliente.objects.filter(documento=documento_cliente).first()
                    query &= Q(cliente=cliente_pedido)


                if datavencimento_inicial and datavencimento_final:
                    query &= Q(
                        data_vencimento__range=[datavencimento_inicial, datavencimento_final])

                if datapagamento_inicial and datapagamento_final:
                    query &= Q(
                        data_pagamento__range=[datapagamento_inicial, datapagamento_final])

                if forma_pagamento:
                    query &= Q(tipo_pgto__icontains=forma_pagamento)

                if status_conta:
                    query &= Q(status_conta__icontains=status_conta)

                contas_busca = Contas_Receber.objects.filter(query).order_by('data_vencimento')
                print (contas_busca)




            else:
                status_conta = Contas_Receber.STATUS_CONTA
        total_vencido = 0
        total_receber = 0
        total_pago = 0
        for conta in contas_busca:

            if conta.status_conta == 'Vencida':
                total_vencido += conta.valor
            elif conta.status_conta == 'A Vencer':
                total_receber += conta.valor
            elif conta.status_conta == 'Pago':
                total_pago += conta.valor


        context = {
                    #'pedidos': pedidos_busca,
                    'consultas': contas_busca,
                    #'itens_pedidos': Itens_Pedido.objects.all(),

                    'forma_pagamento': forma_pagamento,
                    'status_conta': status_conta,
                    'formas_pagamento': formas_pagamento,
                    'empresas': Empresas.objects.all(),
                    'bancos': Bancos.objects.all(),
                    'data_atual': date.today(),
                    'colaborador': colaborador,
                    'total_vencido': format_currency(float(total_vencido),'BRL', locale='pt_BR'),
                    'total_receber': format_currency(float(total_receber),'BRL', locale='pt_BR'),
                    'total_pago': format_currency(float(total_pago),'BRL', locale='pt_BR')


                }

        return render(request, 'contas_receber.html', context)
    else:
        return redirect('/painel/')

@login_required
def editar_conta(request, numero_pedido):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    if 'Financeiro' or 'Gerente' in colaborador.funcao:
        pedido = Pedidos.objects.filter(numero=numero_pedido).first()

        consultas = Contas_Receber.objects.filter(pedido_conta=pedido)
        parcelas = consultas.filter(pedido=numero_pedido)
        pedido = Pedidos.objects.filter(numero=numero_pedido).first()
        tipos_pagamento = TipoPgto.objects.all()
        bancos = Bancos.objects.filter(empresa=pedido.empresa_pedido)
        status_parcela = Contas_Receber.STATUS_CONTA

        contaref = Contas_Receber.objects.filter(pedido=numero_pedido).first()
        flag_uni = contaref.flag_uni
        cob_unificado = Cob_Unificada.objects.filter(conta=contaref)
        data_atual = datetime.now()
        flag = 1
        for consulta in consultas:
            if consulta.status_conta == 'Pendente Faturamento' and consulta.momento_pagamento == 'FATURADO':
                flag = 0
            else:
                flag = 1

        if pedido.condicao_pagamento and flag == 0:
            condicao_pagamento = pedido.condicao_pagamento
            dias_vencimento = list(map(int, condicao_pagamento.descricao.split(", ")))
            contador = 0
            for dia in dias_vencimento:
                contador +=1
            if contador == parcelas.count():
                vencimentos = [data_atual + timedelta(days=dias) for dias in dias_vencimento]
                for i, (parcela, vencimento) in enumerate(zip(parcelas, vencimentos), start=1):
                    parcela.data_vencimento = vencimento
                    parcela.save()
                context = {

                    'parcelas': parcelas,
                    'pedido': pedido if pedido else None,
                    'tipos_pagamento': tipos_pagamento,
                    'status_conta': status_parcela,
                    'bancos': bancos,
                    'unificados': cob_unificado if cob_unificado else None,
                    'conta': contaref
                }
                return render(request, 'editar_contas_receber.html', context)
            else:
                return HttpResponse('Pedido com Condição de Pagamento Diferente do Total de Parcelas do Conta a Receber')
        else:
            context = {

                'parcelas': consultas,
                'pedido': pedido if pedido else None,
                'tipos_pagamento': tipos_pagamento,
                'status_conta': status_parcela,
                'bancos': bancos,
                'unificados': cob_unificado if cob_unificado else None,
                'conta': contaref
            }
            url_atual = request.get_full_path()
        if request.method == 'POST':
            # Atualizar a lista de pagamentos do Contas_Receber
            pedido = Pedidos.objects.filter(numero=numero_pedido).first()

            conta_atual = list(Contas_Receber.objects.filter(pedido=numero_pedido).values())

            # Criar instâncias de Temp_Contas_Receber para armazenar temporariamente os dados
            for item in conta_atual:
                temp_conta = Temp_Contas_Receber(
                    pedido=item['pedido'],
                    cliente_id=item['cliente_id'],
                    descricao=item['descricao'],
                    data_vencimento=item['data_vencimento'] if item['data_vencimento'] else None,
                    data_pagamento=item['data_pagamento'] if item['data_pagamento'] else None,
                    tipo_pgto=item['tipo_pgto'],
                    numero_parcela=item['numero_parcela'],
                    total_parcelas=item['total_parcelas'],
                    valor=item['valor'],
                    status_conta=item['status_conta'],
                    momento_pagamento=item['momento_pagamento'],
                    caminho_boleto=item['caminho_boleto'],
                    id_cobranca=item['id_cobranca'],
                    status_cobranca=item['status_cobranca'],
                    flag_uni = item['flag_uni'],


                )
                temp_conta.save()
            caminho_boleto =  None
            id_cobranca =  None
            status_cobranca = None
            Contas_Receber.objects.filter(pedido=numero_pedido).delete()

            cliente = Cliente.objects.filter(documento=pedido.cliente_pedido.documento).first()

            # Obter os dados dos campos do formulário
            status_contas = request.POST.getlist('status_conta')
            data_vencimentos = request.POST.getlist('data_vencimento')
            data_pagamentos = request.POST.getlist('data_pagamento')
            tipos_pagamento = request.POST.getlist('tipo_pagamento')
            bancos = request.POST.getlist('banco')
            numeros_parcela = request.POST.getlist('numero_parcela')
            valores_parcela = request.POST.getlist('valor_parcela')
            momentos_pagamento = request.POST.getlist ('momento_pagamento')
            valor_Total =0
            total_parcelas = len(data_vencimentos)

            # Iterar sobre os dados e criar os registros de Contas_Receber
            for i in range(len(data_vencimentos)):
                data_vencimento = data_vencimentos[i]
                if data_pagamentos :
                    data_pagamento = data_pagamentos[i]
                else: data_pagamento = None
                tipo_pagamento = tipos_pagamento[i]
                numero_parcela = numeros_parcela[i]
                valor_parcela = float(valores_parcela[i])
                if tipo_pagamento == 'BOLETO':
                    conta_antiga = Temp_Contas_Receber.objects.filter(pedido=numero_pedido,
                                                                      data_vencimento=data_vencimento,
                                                                      valor=valor_parcela,
                                                                      tipo_pgto=tipo_pagamento
                                                                      ).first()
                    if conta_antiga:

                        caminho_boleto = conta_antiga.caminho_boleto
                        id_cobranca = conta_antiga.id_cobranca
                        status_cobranca = conta_antiga.status_cobranca
                        flag_uni = conta_antiga.flag_uni if conta_antiga.flag_uni else 0
                if contaref.flag_uni != 1:
                    descricao = f"{cliente.nome} {numero_parcela} / {total_parcelas}"
                else:
                    descricao = contaref.descricao

                momento_pagamento = momentos_pagamento[i]
                valor_Total += valor_parcela
                status = status_contas[i]
                banco_nome = bancos[i]
                banco = Bancos.objects.filter(nome=banco_nome).first()

                if data_pagamento:

                    conta_receber = Contas_Receber(
                        pedido=numero_pedido,
                        cliente = cliente,
                        descricao= descricao,
                        data_vencimento=data_vencimento,
                        data_pagamento=data_pagamento,
                        tipo_pgto=tipo_pagamento,
                        numero_parcela=numero_parcela,
                        total_parcelas=total_parcelas,
                        valor=valor_parcela,
                        status_conta=status,
                        banco = banco,
                        momento_pagamento=momento_pagamento,
                        caminho_boleto = caminho_boleto if caminho_boleto else None,
                        id_cobranca = id_cobranca if id_cobranca else None,
                        status_cobranca = status_cobranca if status_cobranca else None,
                        pedido_conta = pedido,
                        flag_uni = flag_uni


                    )
                else:
                    conta_receber = Contas_Receber(
                        pedido=numero_pedido,
                        descricao=descricao,
                        cliente=cliente,
                        data_vencimento=data_vencimento,
                        tipo_pgto=tipo_pagamento,
                        numero_parcela=numero_parcela,
                        total_parcelas=total_parcelas,
                        valor=valor_parcela,
                        status_conta=status,
                        banco=banco,
                        momento_pagamento=momento_pagamento,
                        caminho_boleto=caminho_boleto if caminho_boleto else None,
                        id_cobranca=id_cobranca if id_cobranca else None,
                        status_cobranca=status_cobranca if status_cobranca else None,
                        pedido_conta=pedido,
                        flag_uni=flag_uni
                    )

                conta_receber.save()


            if contaref.flag_uni != 1:
                pedido.valorTotal = valor_Total
                pedido.save()

            else:
                unificados = cob_unificado
                for unificado in unificados:
                    unificado.conta = Contas_Receber.objects.filter(pedido=numero_pedido).first()
                    unificado.save()



                url_anterior = request.META.get('HTTP_REFERER')

            return redirect(url_atual)



        return render(request, 'editar_contas_receber.html', context)


    return redirect('/contas_receber/')
def atualiza_itens():
    produtos = Produtos.objects.all()
    produtos_revenda = Produtos_Revenda.objects.all()
    lista_produtos = ['CLEANKAP', 'LIFTKAP', 'LIFTKAP', 'CARPETE', 'CONECTADO' ]
    for p in produtos:
        for l in lista_produtos:
            if l in p.descricao:
                p.flag_conferir=1
                p.save()
    for p in produtos_revenda:
        for l in lista_produtos:
            if l in p.descricao:
                p.flag_conferir = 1
                p.save()


def busca_item(numero_pedido):

    pedido = Pedidos.objects.filter(numero=numero_pedido).first()
    itens = Itens_Pedido.objects.filter(pedido=numero_pedido)
    if itens:
        for item in itens:
            if item.produto:
                if item.produto.flag_conferir == 1:
                    pedido.status = 'Conferir Item'
                    pedido.save()
                else:
                    pass
            elif item.produto_revenda:
                if item.produto_revenda.flag_conferir == 1:
                    pedido.status = 'Conferir Item'
                    pedido.save()
                else:
                    pass
            else:
                print('NAO EXISTE PRODUTO')



@csrf_exempt
@login_required
def pedido_create_super(request):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    if colaborador.empresa.flag_pagamento == 1:
        return redirect ('/dologin/')
    elif not colaborador:
        return redirect('/dologin/')
    else:
        pass
    colaboradores_ativos =  Colaborador.objects.filter (status__descricao='ATIVO')
    vendedores = colaboradores_ativos.filter(Q(funcao='Vendedor Interno') | Q(funcao='Vendedor Externo')|Q(funcao='Gerente'))
    clientes = Cliente.objects.all()
    produtos = Produtos.objects.all()
    produtos_adicionais = Produtos.objects.filter(produto_adicional = 'SIM')
    empresas = Empresas.objects.all()
    bairros = Bairro.objects.all()
    tipos_pagamentos = TipoPgto.objects.all()
    fornecedores = Fornecedor.objects.filter(tipo_fornecedor='1')
    bonificados = Fornecedor.objects.filter(tipo_fornecedor='8')
    tipos_pedidos = Tipo_Pedido.objects.all()
    produtos_revenda = Produtos_Revenda.objects.all()
    lista_produtos = []


    for produto in produtos:
        lista_produtos.append(produto)
    for produto in produtos_revenda:
        lista_produtos.append(produto)

    revendas = Revenda.objects.all()
    zona_entrega = 'Sem Zona'
    momentos = [
        'Antecipado',
        'Entrega',
        'Faturado'
    ]
    cep_entrega = ''
    uf_entrega = ''
    cidade_entrega = ''
    bairro_entrega = ''
    endereco_entrega = ''
    numero_end_entrega = '0'
    complemento = ''
    total_avista ='0'
    total_parcelado = '0'
    desconto ='0'
    condicoes_pagamentos = Condicoes_Pagamentos.objects.all
    context = {
        'vendedor': colaborador.nome,
        'vendedores': vendedores,
        'clientes': clientes,
        'produtos': lista_produtos,
        'produtos_adicionais': produtos_adicionais,
        'bairros': bairros,
        'fornecedores': fornecedores,
        'empresas': empresas,
        'tipos_pagamentos':tipos_pagamentos,
        'momentos': momentos,
        'colaborador':colaborador,
        'revendas':revendas,
        'produtos_revenda':produtos_revenda,
        'tipos_pedidos':tipos_pedidos,
        'lista_produtos': lista_produtos,
        'bonificados': bonificados,
        'condicoes_pagamentos': condicoes_pagamentos

    }

    if request.method == 'POST':

        pagamento_antecipado = 0.00
        ponto_referencia_entrega =" "
        numero_pedido = request.POST['numero_pedido']
        print(numero_pedido)
        datapg_antecipado = None
        flag_pedido = request.POST.get('flag_pedido')
        print(flag_pedido)
        if flag_pedido:
            flag_pedido = 1
        else:
            flag_pedido = 0
        pedido_temp = Temp_Pedidos.objects.filter(pedido=numero_pedido).first()
        datapedido = request.POST.get('data-pedido')
        prazo_entrega = request.POST.get('data_entrega')

        if 'Vendedor or Gerente' in colaborador.funcao:
            vendedor = colaborador.nome
        else:

             vendedor = request.POST.get('vendedor1')

        vendedor2 = request.POST.get('vendedor2')

        total_pedido_parcial = 0.00


        documento = request.POST.get('documento')
        documento_cliente = ''.join(filter(str.isdigit, documento))
        cliente_pedido = Cliente.objects.filter(documento=documento_cliente).first()


        empresa = request.POST.get('empresa_pedido')
        empresa = Empresas.objects.get(nome=empresa) if empresa else colaborador.empresa


        tipo_pedido = request.POST.get('tipo_pedido')
        tipo_do_pedido = Tipo_Pedido.objects.filter(descricao=tipo_pedido).first()

        itens = Itens_Pedido.objects.filter(pedido=numero_pedido)
        for item in itens:
            total_pedido_parcial += item.total_item


        tipo_frete = request.POST.get('opcao_frete')
        nfe_opcao = request.POST.get('pedido-nfe-opcao')
        retirada_flag = 0
        if nfe_opcao == 'sim':
            nfe_flag = 1
        else:
            nfe_flag = 0


        if tipo_frete == 'Transportadora':
            frete_flag = 1
        else:
            frete_flag = 0
            if tipo_frete == 'retirada':
                retirada_flag = 1
            else:
                retirada_flag = 0



        endereco_opcao = request.POST.get('endereco-opcao')
        frete_valor = request.POST.get('frete')
        frete_valor = re.sub(r'^R\$|\.', '', frete_valor).replace(',', '.')
        frete_valor = round(float(frete_valor), 2) if frete_valor else 0
        status_pedido = 'Aberto'

        numero_pedido_str = str(numero_pedido)
        last_5_digits = numero_pedido_str[-5:]


        informacoes_complementares = request.POST.get('informacoes_complementares')

        if cliente_pedido:
            cliente_id = cliente_pedido
        else:
            cliente_id = None  # Trate o caso em que o cliente não é encontrado

        if endereco_opcao == 'cadastro':
            if cliente_pedido:
                cep_entrega = cliente_pedido.CEP
                uf_entrega = cliente_pedido.estado
                cidade_entrega = cliente_pedido.cidade
                bairro_entrega = cliente_pedido.bairro
                endereco_entrega = cliente_pedido.endereco
                numero_end_entrega = cliente_pedido.numero_endereco
                complemento = cliente_pedido.complemento
                ponto_referencia_entrega = cliente_pedido.ponto_referencia
                if (cidade_entrega == 'Rio de Janeiro' or cidade_entrega == 'RIO DE JANEIRO'):
                    localidade = Bairro.objects.filter(nome=bairro_entrega, cidade=cidade_entrega).first()
                    if localidade:
                        zona_entrega = localidade.zona if localidade.zona else None


        elif endereco_opcao == 'opcional':
            cep_entrega = request.POST.get('cep-entrega-opcional')
            bairro_entrega = request.POST.get('bairro-entrega-opcional')
            cidade_entrega = request.POST.get('cidade-entrega-opcional')
            uf_entrega = request.POST.get('uf-entrega-opcional')
            endereco_entrega = request.POST.get('endereco-entrega-opcional')
            numero_end_entrega = request.POST.get('numero-endereco-opcional')
            complemento = request.POST.get('complemento-entrega-opcional')
            ponto_referencia_entrega = request.POST.get('pontoreferencia-opcional')
            if (cidade_entrega == 'Rio de Janeiro'):
                localidade = Bairro.objects.get(nome=bairro_entrega, cidade=cidade_entrega)
                zona_entrega = localidade.zona if localidade.zona else None

        print (total_pedido_parcial)
        desconto = 0.00
        repasse = request.POST.get('acerto')
        repasse=re.sub(r'^R\$|\.', '', repasse).replace(',', '.') if repasse else None
        repasse = float(repasse)
        desconto = float(desconto) if desconto else 0
        if desconto > 0:
            total_pedido =round(float(total_pedido_parcial - desconto), 2)
        else:
            total_pedido = float(total_pedido_parcial)
            total_pedido = round(total_pedido,2)
        print(total_pedido)
        bonificado_pedido =""


        if repasse > 0:
            repasse = float(repasse)
            conta_pagar = Contas_Pagar()
            total_pagar = total_pedido - total_pedido_parcial
            conta_pagar.status = 'A Vencer'
            conta_pagar.pedido = numero_pedido
            bonificado = request.POST.get('bonificado')
            conta_pagar.fornecedor = Fornecedor.objects.filter(nome=bonificado).first()
            bonificado_pedido =  conta_pagar.fornecedor
            prazo_entrega_datetime = datetime.strptime(prazo_entrega, '%Y-%m-%d')
            data_vencimento = prazo_entrega_datetime + timedelta(days=30)
            conta_pagar.data_vencimento = data_vencimento
            conta_pagar.numero_parcela=1
            conta_pagar.total_parcelas =1
            conta_pagar.valor = repasse
            conta_pagar.descricao = "BONIFICACAO"
            conta_pagar.save()

        itens = Itens_Pedido.objects.filter(pedido=numero_pedido)
        menor_comissao = 100
        vendedor_comissao = 0
        vendedor = Colaborador.objects.get(nome=vendedor)


        vendedor_comissao = request.POST.get('comissao1')
        vendedor2_comissao = 0
        if vendedor2:
            vendedor2 = Colaborador.objects.get(nome=vendedor2)
            if vendedor_comissao =='1':
                vendedor2_comissao = 1
            elif vendedor_comissao =='3':
                vendedor2_comissao = 3
            elif vendedor_comissao == '5':
                vendedor2_comissao = 7
            elif vendedor_comissao == '6':
                vendedor2_comissao = 10
            elif vendedor_comissao == '7':
                vendedor2_comissao = 15



        bancopg_antecipado = ""
        informacoes_parcelas = ""
        momento_pagamento_1 = request.POST.get('momento-pagamento1')
        momento_pagamento_2 = request.POST.get('momento-pagamento2') if request.POST.get('momento-pagamento2') else None
        total_pagamento1 = request.POST.get('total_1')
        total_pagamento1 = re.sub(r'^R\$|\.', '', total_pagamento1).replace(',', '.') if total_pagamento1 else 0.00
        forma_pagamento1 = request.POST.get('forma_pagamento1')
        parcelas_cartao_1 = request.POST.get('parcelas_cartao_1')
        total_parcelado_cartao_1 = request.POST.get('valor_parcelado_cartao_1')
        total_parcelado_cartao_1 = re.sub(r'^R\$|\.', '', total_parcelado_cartao_1).replace(',', '.') if total_parcelado_cartao_1 else 0.00
        faturamento_valor_parcelas = request.POST.get('valor_parcelas')
        faturamento_valor_parcelas = re.sub(r'^R\$|\.', '', faturamento_valor_parcelas).replace(',', '.') if faturamento_valor_parcelas else 0.00
        faturamento_quantidade_parcelas = request.POST.get('total_parcelas')
        total_pagamento2 = request.POST.get('total_2')
        total_pagamento2 = re.sub(r'^R\$|\.', '', total_pagamento2).replace(',', '.') if total_pagamento2 else 0.00
        forma_pagamento2 = request.POST.get('forma_pagamento2')
        parcelas_cartao_2 = request.POST.get('parcelas_cartao_2')
        total_parcelado_cartao_2 = request.POST.get('valor_parcelado_cartao_2')
        total_parcelado_cartao_2 = re.sub(r'^R\$|\.', '', total_parcelado_cartao_2).replace(',', '.') if total_parcelado_cartao_2 else 0.00
        faturamento_quantidade_parcelas_2 = request.POST.get('total_parcelas_2')
        faturamento_valor_parcelas_2 = request.POST.get('valor_parcelas_2')
        faturamento_valor_parcelas_2 = re.sub(r'^R\$|\.', '', faturamento_valor_parcelas_2).replace(',', '.') if faturamento_quantidade_parcelas_2 else 0.00

        pagamento_entrega = 0.00
        Contas_Receber.objects.filter (pedido=numero_pedido).delete()

        print (momento_pagamento_1)
        if momento_pagamento_1 == 'ANTECIPADO':

            status_pedido = 'Pendente Financeiro'
            datapg_antecipado = request.POST.get('data_antecipado')
            bancopg_antecipado = request.POST.get('banco_pa')
            data_vencimento = date.today()

            if 'CREDITO' in forma_pagamento1:
                pagamento_antecipado =float(total_parcelado_cartao_1)
                for parcela in range(1, int(parcelas_cartao_1) + 1):

                    valor_parcela = float(total_parcelado_cartao_1) / int(parcelas_cartao_1)


                    contas_receber = Contas_Receber(
                        pedido=numero_pedido,  # Substitua pelo valor correto
                        cliente=cliente_pedido,
                        descricao='Parcela {} - Crédito'.format(parcela),
                        data_vencimento=data_vencimento,  # Substitua pelo valor correto
                        data_pagamento=None,
                        tipo_pgto=forma_pagamento1,
                        numero_parcela=parcela,
                        total_parcelas=int(parcelas_cartao_1),
                        valor=valor_parcela,
                        status_conta='A Vencer',
                        momento_pagamento='ANTECIPADO'
                    )
                    contas_receber.save()
                    data_vencimento = (data_vencimento + timedelta(days=30))

            else:
                pagamento_antecipado =float(total_pagamento1)
                numero_parcela =1
                total_parcelas = 1
                data_vencimento = datetime.today() + timedelta(days=1)
                descricao = f"{cliente_pedido.nome} {numero_parcela} / {total_parcelas}"
                contas_receber = Contas_Receber(
                    pedido=numero_pedido,
                    cliente=cliente_pedido,
                    descricao=descricao,
                    data_vencimento=data_vencimento,
                    data_pagamento=None,
                    tipo_pgto=forma_pagamento1,
                    numero_parcela=1,
                    total_parcelas=1,
                    valor=total_pagamento1,
                    status_conta='A Vencer',
                    momento_pagamento='ANTECIPADO'
                )
                contas_receber.save()

        if momento_pagamento_1 == 'ENTREGA':
            data_vencimento = prazo_entrega
            status_pedido = 'Producao'
            data_vencimento = datetime.today() + timedelta(days=1)
            pagamento_entrega = 0.00
            if 'CREDITO' in forma_pagamento1:
                pagamento_entrega = round(float(total_parcelado_cartao_1), 2)
                for parcela in range(1, int(parcelas_cartao_1) + 1):
                    data_vencimento = (data_vencimento + timedelta(days=30))
                    print(data_vencimento)
                    valor_parcela = float(total_parcelado_cartao_1) / int(parcelas_cartao_1)
                    contas_receber = Contas_Receber(
                        pedido=numero_pedido,  # Substitua pelo valor correto
                        cliente = cliente_pedido,
                        descricao='Parcela {} - Crédito'.format(parcela),
                        data_vencimento=data_vencimento,  # Substitua pelo valor correto
                        data_pagamento=None,
                        tipo_pgto=forma_pagamento1,
                        numero_parcela=parcela,
                        total_parcelas=int(parcelas_cartao_1),
                        valor=valor_parcela,
                        status_conta='A Vencer',
                        momento_pagamento='ENTREGA'
                    )
                    contas_receber.save()
            else:
                numero_parcela = 1
                total_parcelas = 1
                descricao = f"{cliente_pedido.nome} {numero_parcela} / {total_parcelas}"
                pagamento_entrega = round(float(total_pagamento1), 2)
                contas_receber = Contas_Receber(
                    pedido=numero_pedido,  # Substitua pelo valor correto
                    cliente=cliente_pedido,
                    descricao=descricao,
                    data_vencimento=prazo_entrega,
                    data_pagamento=None,
                    tipo_pgto=forma_pagamento1,
                    numero_parcela=1,
                    total_parcelas=1,
                    valor=total_pagamento1,
                    status_conta='A Vencer',
                    momento_pagamento='ENTREGA'
                )
                contas_receber.save()

        if momento_pagamento_1 == 'FATURADO':
            status_pedido = 'Producao'

            valor_parcela = float(faturamento_valor_parcelas)
            print(valor_parcela)
            informacoes_parcelas = request.POST.get('informacoes_parcelas')
            for parcela in range(1, int(faturamento_quantidade_parcelas) + 1):



                contas_receber = Contas_Receber(
                    pedido=numero_pedido,
                    cliente = cliente_pedido,
                    descricao='{} - Parcela {} - Faturado'.format(cliente_pedido, parcela),
                    data_pagamento=None,
                    tipo_pgto=forma_pagamento1,
                    numero_parcela=parcela,
                    total_parcelas=int(faturamento_quantidade_parcelas),
                    valor=valor_parcela,
                    status_conta='Pendente Faturamento',
                    momento_pagamento='FATURADO'
                )

                contas_receber.save()


        if momento_pagamento_2 == 'ANTECIPADO':

            data_vencimento = datetime.today() + timedelta(days=1)
            datapg_antecipado = request.POST.get('data_antecipado_2')
            status_pedido = 'Pendente Financeiro'
            if 'CREDITO' in forma_pagamento2:
                pagamento_antecipado = float(total_parcelado_cartao_2)
                for parcela in range(1, int(parcelas_cartao_2) + 1):
                    data_vencimento = (data_vencimento + timedelta(days=30))
                    valor_parcela = float(total_parcelado_cartao_2) / int(parcelas_cartao_2)
                    contas_receber = Contas_Receber(
                        pedido=numero_pedido,
                        cliente = cliente_pedido, # Substitua pelo valor correto
                        descricao='Parcela {} - Crédito'.format(parcela),
                        data_vencimento=data_vencimento,
                        data_pagamento=None,
                        tipo_pgto=forma_pagamento2,
                        numero_parcela=parcela,
                        total_parcelas=int(parcelas_cartao_2),
                        valor=valor_parcela,
                        status_conta='A Vencer',
                        momento_pagamento='ANTECIPADO'
                    )
                    contas_receber.save()

            else:
                numero_parcela = 1
                total_parcelas = 1
                descricao = f"{cliente_pedido.nome} {numero_parcela} / {total_parcelas}"
                pagamento_antecipado = float(total_pagamento2)
                contas_receber = Contas_Receber(
                    pedido=numero_pedido,
                    cliente = cliente_pedido,# Substitua pelo valor correto
                    descricao=descricao,
                    data_vencimento=data_vencimento,
                    data_pagamento=None,
                    tipo_pgto=forma_pagamento2,
                    numero_parcela=1,
                    total_parcelas=1,
                    valor=total_pagamento2,
                    status_conta='A Vencer',
                    momento_pagamento='ANTECIPADO'
                )
                contas_receber.save()



        if momento_pagamento_2 == 'ENTREGA':

            data_vencimento = prazo_entrega

            if 'CREDITO' in forma_pagamento2:
                pagamento_entrega = round(float(total_parcelado_cartao_2), 2)
                data_vencimento = datetime.today() + timedelta(days=15)
                for parcela in range(1, int(parcelas_cartao_2) + 1):
                    data_vencimento = (data_vencimento + timedelta(days=30))
                    valor_parcela = float(total_parcelado_cartao_2) / int(parcelas_cartao_2)
                    contas_receber = Contas_Receber(
                        pedido=numero_pedido,
                        cliente = cliente_pedido,# Substitua pelo valor correto
                        descricao='Parcela {} - Crédito'.format(parcela),
                        data_vencimento=data_vencimento,  # Substitua pelo valor correto
                        data_pagamento=None,
                        tipo_pgto='CRÉDITO',
                        numero_parcela=parcela,
                        total_parcelas=int(parcelas_cartao_2),
                        valor=valor_parcela,
                        status_conta='A Vencer',
                        momento_pagamento='ENTREGA'
                    )
                    contas_receber.save()
            else:
                pagamento_entrega = round(float(total_pagamento2), 2)
                numero_parcela = 1
                total_parcelas = 1
                descricao = f"{cliente_pedido.nome} {numero_parcela} / {total_parcelas}"
                contas_receber = Contas_Receber(
                    pedido=numero_pedido,
                    cliente = cliente_pedido,# Substitua pelo valor correto
                    descricao=descricao,
                    data_vencimento=prazo_entrega,
                    data_pagamento=None,
                    tipo_pgto=forma_pagamento2,
                    numero_parcela=1,
                    total_parcelas=1,
                    valor=total_pagamento2,
                    status_conta='A Vencer',
                    momento_pagamento='ENTREGA'
                )
                contas_receber.save()

        if momento_pagamento_2 == 'FATURADO':


            valor_parcela = float(faturamento_valor_parcelas_2)

            for parcela in range(1, int(faturamento_quantidade_parcelas_2) + 1):


                contas_receber = Contas_Receber(
                    pedido=numero_pedido,
                    cliente = cliente_pedido,# Substitua pelo valor correto
                    descricao='{} - Parcela {} - Faturado'.format(cliente_pedido, parcela),
                    data_pagamento=None,
                    tipo_pgto=forma_pagamento2,
                    numero_parcela=parcela,
                    total_parcelas=int(faturamento_quantidade_parcelas_2),
                    valor=valor_parcela,
                    status_conta='Pendente Faturamento',
                    momento_pagamento='FATURADO'
                )

                contas_receber.save()

        # Obtém a data de hoje
        hoje = date.today()
        #CHECA PENDENCIA LOGISTICA
        itens = Itens_Pedido.objects.filter(pedido=numero_pedido)
        if prazo_entrega:
            prazo_entrega = datetime.strptime(prazo_entrega, '%Y-%m-%d').date()
        else:
            prazo_entrega = hoje + timedelta(days=15)



        # Itera sobre os itens do pedido
        if (tipo_pedido == 'PADRAO'):
            for item in itens:

                produto = item.produto

                # Calcula a data de entrega com base na data de hoje mais o prazo do produto
                data_entrega = hoje + timedelta(days=produto.prazo_fabricacao)

                # Verifica se a data de entrega é menor que o prazo de entrega
                if data_entrega > prazo_entrega:
                    if status_pedido == 'Pendente Financeiro':
                        pass
                    else:
                        status_pedido = 'Pendente Logistica'
        else:
            for item in itens:

                produto = item.produto_revenda

                # Calcula a data de entrega com base na data de hoje mais o prazo do produto
                data_entrega = hoje + timedelta(days=produto.prazo_fabricacao)

                # Verifica se a data de entrega é menor que o prazo de entrega
                if data_entrega > prazo_entrega:
                    if status_pedido == 'Pendente Financeiro':
                        pass
                    else:
                        status_pedido = 'Pendente Logistica'



        #CADASTRO DE ARQUIVO LAYOUT
        nome_arquivo = None
        arquivo = request.FILES.get('arquivo')

        empresa = Empresas.objects.filter(id=1).first()

        if arquivo:
            s3_client = boto3.client('s3', aws_access_key_id=empresa.aws_id,
                                     aws_secret_access_key=empresa.aws_chave)
            agora = datetime.now()
            nome_arquivo = f"{numero_pedido}{agora.strftime('%Y%m%d%H%M%S')}.pdf"

            s3_client.upload_fileobj(arquivo, 'storagesw', nome_arquivo)

        nfe_antecipado = request.POST.get('antecipado-nfe-opcao')

        flag_nfe_antecipada = 0
        if nfe_antecipado == 'sim':

            status_pedido = 'Nfe Antecipada'
            flag_nfe_antecipada = 1

        else:
            pass

        fundo = request.POST.get('fundo')
        borda = request.POST.get('borda')
        letras = request.POST.get('letras')
        logomarca = request.POST.get('logomarca')
        vendedor2_comissaov =0.00
        vendedor_comissaov =0.00
        if repasse > 0:
            total_pedido_parcial -= repasse
        vendedor_comissaov = total_pedido_parcial/100 * float(vendedor_comissao)
        if vendedor2:
            vendedor2_comissaov= total_pedido_parcial/100 * float(vendedor2_comissao)
            vendedor2_comissaov = round(vendedor2_comissaov, 2)
        else:
            pass
        vendedor_comissaov=round(vendedor_comissaov,2)

        observacao = request.POST.get('observacao')

        comprador_nome = request.POST.get('comprador_nome')
        comprador_telefone = request.POST.get('comprador_telefone')
        comprador_telefone= ''.join(filter(str.isdigit, comprador_telefone))
        comprador_email = request.POST.get('comprador_email')

        id_condicao = request.POST.get('condicoes_pagamento')
        condicao_pagamento = None
        if id_condicao:
            condicao_pagamento = Condicoes_Pagamentos.objects.filter(id=id_condicao).first()


        pedido_obj = Pedidos(
            numero=numero_pedido,
            datapedido=datapedido,
            valorTotal=total_pedido,
            valorTotalParcial=total_pedido_parcial,
            descontototal=desconto,
            cliente_pedido=cliente_pedido,
            prazo_entrega=prazo_entrega,
            pagamento_antecipado=pagamento_antecipado,
            datapg_antecipado= datapg_antecipado if datapg_antecipado else None ,
            bancopg_antecipado=bancopg_antecipado if bancopg_antecipado else None,
            Vendedor=vendedor,
            Vendedor2=vendedor2,
            vendedor_comissao=vendedor_comissao,
            vendedor_comissaov=vendedor_comissaov,
            vendedor2_comissao=vendedor2_comissao,
            vendedor2_comissaov=vendedor2_comissaov,
            cep_entrega=cep_entrega,
            uf_entrega=uf_entrega,
            cidade_entrega=cidade_entrega,
            bairro_entrega=bairro_entrega,
            endereco_entrega=endereco_entrega,
            numero_end_entrega=numero_end_entrega,
            flag_nfe_antecipada =flag_nfe_antecipada,
            complemento=complemento,
            ponto_referencia_entrega = ponto_referencia_entrega,
            empresa_pedido=empresa,
            zona_entrega=zona_entrega if zona_entrega else "Sem Zona",
            pagamento_entrega = pagamento_entrega if pagamento_entrega else 0,
            status=status_pedido,
            caminho_layout = f"https://storagesw.s3.amazonaws.com/{nome_arquivo}" if nome_arquivo else None,
            informacoes_adicionais = informacoes_complementares,
            operacao_nota = 1,
            frete_flag = frete_flag,
            flag_nfe = nfe_flag,
            tipo_pedido = tipo_do_pedido,
            informacoes_pagamento=informacoes_parcelas if informacoes_parcelas else None,
            retirada_flag = retirada_flag,
            valor_repasse = repasse,
            logomarca=logomarca,
            letra=letras,
            fundo=fundo,
            borda= borda,
            observacao= observacao,
            tipo_entrega = tipo_frete,
            flag_aberto=flag_pedido,
            bonificado = bonificado_pedido if bonificado_pedido else None,
            flag_impressao = 0,
            comprador_nome= comprador_nome,
            comprador_email = comprador_email,
            comprador_telefone = comprador_telefone,
            condicao_pagamento = condicao_pagamento if condicao_pagamento else None

        )

        pedido_obj.save()
        if int(empresa.ultimo_pedido_ref) >= int(numero_pedido):
            pass
        else:
            empresa.ultimo_pedido_ref = numero_pedido
            empresa.save()

        if frete_valor > 0:

            conta_obj = Contas_Pagar(
                pedido=numero_pedido,
                descricao="FRETE REF PEDIDO",
                valor=frete_valor,

            )
            conta_obj.save()
            atualiza_pedido = Pedidos.objects.get(numero=numero_pedido)
            atualiza_pedido.frete_valor = frete_valor
            atualiza_pedido.save()
        else:
            pass

        contas = Contas_Receber.objects.filter(pedido=numero_pedido)
        pedido = Pedidos.objects.filter(numero=numero_pedido).first()
        for conta in contas:
            conta.pedido_conta = pedido
            conta.save()

        ### VERIFICA SE TEM ITEM A SER CONFERIDO
        busca_item(numero_pedido)





        if tipo_do_pedido.descricao == 'PADRAO':
            compara_pedidos(numero_pedido)

        redirect_url = f'/geradocpedido/{numero_pedido}/'

        return redirect(redirect_url)




    return render(request, 'novopedido_super.html', context)

def pedido_save(request):

    data = json.loads(request.body)  # Obtém o valor do parâmetro 'data' da URL
    print (data)

    response_data = {
        'success': True,
        'message': 'Dados do formulário recebidos com sucesso!'
    }
    return JsonResponse(response_data)


def create_contas_pagar(request):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    empresas = Empresas.objects.all()
    if 'Financeiro' or 'Gerente' in colaborador.funcao:

        context = {

            'fornecedores': Fornecedor.objects.all(),
            'tipos': TipoPgto.objects.all(),
            'empresas': empresas,


        }
        if request.method =='POST':

            fornecedor = request.POST.get('fornecedor')
            fornecedor = Fornecedor.objects.get(nome=fornecedor)
            pedido = request.POST.get('numero_pedido')
            descricao = request.POST.get('descricao')
            notafiscal = request.POST.get('nota_fiscal')
            notafiscal_doc = request.FILES.get('file-notafiscal', None)
            empresa_nome = request.POST.get('empresa')
            empresa = Empresas.objects.filter(nome=empresa_nome).first()
            nome_arquivo=""

            empresaaws = Empresas.objects.filter(id=1).first()
            if notafiscal_doc:
                s3_client = boto3.client('s3', aws_access_key_id=empresaaws.aws_id,
                                         aws_secret_access_key=empresaaws.aws_chave)
                agora = datetime.now()
                fornecedor_arquivo = fornecedor.nome.strip()
                nome_arquivo = f"NF{notafiscal}{fornecedor_arquivo}{agora.strftime('%Y%m%d%H%M%S')}.pdf"
                s3_client.upload_fileobj(notafiscal_doc, 'storagesw', nome_arquivo)


            quantidade_parcelas = int(request.POST.get('quantidade_parcelas'))
            valor_total = float(request.POST.get('valorTotal'))
            valor_parcela = valor_total / quantidade_parcelas

            tipo_pagamento = request.POST.get('tipo_pagamento')
            tipo_pagamento = TipoPgto.objects.get(descricao=tipo_pagamento)
            outras_informacoes = request.POST.get('outras_informacoes')

            for i in range(1, quantidade_parcelas + 1):
                valor_parcela_i = request.POST.get(f'valor_parcela_{i}')
                data_vencimento_i = request.POST.get(f'data_vencimento_{i}')
                anexar_boleto_i = request.FILES.get(f'anexar_boleto_{i}')
                if anexar_boleto_i:
                    empresaaws = Empresas.objects.filter(id=1).first()
                    s3_client = boto3.client('s3', aws_access_key_id=empresaaws.aws_id,
                                             aws_secret_access_key=empresaaws.aws_chave)

                    agora = datetime.now()
                    fornecedor_arquivo = fornecedor.nome.strip()
                    nome_boleto = f"Boleto{i}{fornecedor_arquivo}{agora.strftime('%Y%m%d%H%M%S')}.pdf"
                    s3_client.upload_fileobj(anexar_boleto_i, 'storagesw', nome_boleto)

                # Crie o objeto Contas_Pagar referente a cada parcela
                # e salve-o no banco de dados
                parcela = Contas_Pagar(
                    fornecedor = fornecedor,
                    numero_parcela=i,
                    valor=valor_parcela_i,
                    tipo_pgto=tipo_pagamento,
                    data_vencimento=data_vencimento_i,
                    documento_caminho = f'https://storagesw.s3.amazonaws.com/{nome_boleto}' if anexar_boleto_i else None,
                    pedido = pedido if pedido else None,
                    descricao = descricao,
                    total_parcelas = quantidade_parcelas,
                    status = 'A Vencer',
                    notafiscal = notafiscal,
                    notafiscal_caminho = f'https://storagesw.s3.amazonaws.com/{nome_arquivo}' if nome_arquivo else None,
                    outras_informacoes = outras_informacoes,
                    empresa=empresa
                )
                parcela.save()

            return JsonResponse({'message': 'Pagamento registrado com sucesso!'})

        return render (request,'new_contas_pagar.html', context)
    else: return redirect('/painel/')



def verificar_inadimplencia(pedido):
    pedido = Pedidos.objects.filter(numero=pedido).first()
    if pedido:
        contas = Contas_Receber.objects.filter(pedido_conta=pedido)
        if contas:
            for conta in contas:
                if conta.status_conta != 'Pago':
                    return False

            return True

def verifica_nota_fiscal (pedido):
    pedido = Pedidos.objects.filter(numero=pedido).first()
    if pedido.numero_nfe is not None:
        return True
    else:
        return False



def ajusta_bonificado():
    contas_pagar = Contas_Pagar.objects.filter(Q (status='Aguardando')|Q(status='A Vencer'))
    for conta in contas_pagar:
        if conta.descricao == 'BONIFICACAO':
            quitacao=verificar_inadimplencia(conta.pedido)
            if quitacao is True:
                conta.status = 'A Vencer'
                tem_notafiscal = verifica_nota_fiscal (conta.pedido)
                if tem_notafiscal is True:
                    pedido = Pedidos.objects.filter(numero = conta.pedido).first()
                    desconto_nota = pedido.valor_nota * 0.10
                    conta.valor =  pedido.valor_repasse - desconto_nota
                    conta.outras_informacoes = 'BONIFICADO COM DESCONTO NOTA'

                    conta.save()
            else:
                conta.status = 'Aguardando'
                conta.save()





@login_required
def contas_pagar(request):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    if colaborador.empresa.flag_pagamento == 1:
        return redirect('/dologin/')
    elif not colaborador:
        return redirect('/dologin/')
    else:
        pass

    if 'Vendedor' in colaborador.funcao:
        return redirect('/painel/')

    if 'Financeiro' or 'Gerente' in colaborador.funcao:

        ajusta_bonificado()
        today = date.today()
        consulta_pagamentos = Contas_Pagar.objects.all().order_by('data_vencimento')
        for consulta in consulta_pagamentos:
            if consulta.data_vencimento:
                if consulta.data_vencimento < today and (not consulta.data_pagamento):
                    consulta.status = 'Vencida'
                    consulta.save()
                elif (not consulta.data_pagamento):
                    consulta.status = 'A Vencer'
                    consulta.save()
                else:
                    consulta.status = 'Pago'
                    consulta.save()
            else:
                pass

        order_dict = {
            'Vencida': 1,
            'A Vencer': 2,
            'Pago': 3,
            'Aguardando': 4,
        }
        consulta_pagamentos = consulta_pagamentos.annotate(
            order=Case(
                *[When(status=status, then=Value(order)) for status, order in order_dict.items()],
                default=Value(4),  # Define um valor padrão maior para os outros status
                output_field=IntegerField(),
            )
        ).order_by('order', 'status', 'data_vencimento')

        context = {

            'consultas': consulta_pagamentos,
            'bancos': Bancos.objects.values('nome').distinct(),
            'empresas': Empresas.objects.all()


        }
        if request.method == 'GET':
            numero_pedido = request.GET.get('numero_pedido_filter')
            datavencimento_inicial = request.GET.get('filter_vencimento_inicial')
            datavencimento_final = request.GET.get('filter_vencimento_final')
            datapagamento_inicial = request.GET.get('filter_data_pagamento_inicial')
            datapagamento_final = request.GET.get('filter_data_pagamento_final')
            empresa_pedido = request.GET.get('filter_empresa')
            status_conta = request.GET.get('status_pagamento_filter')
            forma_pagamento = request.GET.get('forma_pagamento_filter')
            nome_bonificado = request.GET.get('nome_bonificado_filter')
            pedidos_busca = Pedidos.objects.none()
            contas_busca = consulta_pagamentos

            if numero_pedido or nome_bonificado or status_conta or forma_pagamento or empresa_pedido or (
                    datavencimento_inicial and datavencimento_final) or (
                    datapagamento_inicial and datapagamento_final):

                pedidos_busca = Pedidos.objects.all()
                query = Q()
                if empresa_pedido:
                    empresa = Empresas.objects.filter(nome=empresa_pedido).first()
                    query &=Q(empresa=empresa)



                if nome_bonificado:
                    bonificado = Fornecedor.objects.filter(nome__icontains=nome_bonificado).first()
                    query &= Q(fornecedor=bonificado)
                    print('empresa encontrada é:', bonificado)
                    print(query)


                if numero_pedido:

                    query &= Q(pedido__icontains=numero_pedido)


                if datavencimento_inicial and datavencimento_final:
                    query &= Q(
                        data_vencimento__range=[datavencimento_inicial, datavencimento_final])

                if datapagamento_inicial and datapagamento_final:
                    query &= Q(
                        data_pagamento__range=[datapagamento_inicial, datapagamento_final])

                if forma_pagamento:
                    tipo_pagamento = TipoPgto.objects.filter(descricao=forma_pagamento).first()
                    query &= Q(tipo_pgto=tipo_pagamento)

                if status_conta:
                    query &= Q(status__icontains=status_conta)

                contas_busca = Contas_Pagar.objects.filter(query).order_by('data_vencimento')
                consulta_pagamentos = contas_busca.order_by('data_vencimento')

                context = {

                    'consultas': consulta_pagamentos,
                    'bancos': Bancos.objects.values('nome').distinct(),
                    'empresas': Empresas.objects.all()

                }


        return render(request, 'contas_pagar.html', context)
    else: return redirect ('/painel')


def editar_conta_pagar(request,id):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    if 'Financeiro' or 'Gerente' in colaborador.funcao:

        consulta = Contas_Pagar.objects.get(id=id)
        tipos_pagamento = TipoPgto.objects.all()
        status_consulta = Contas_Pagar.STATUS_CONTA
        fornecedores = Fornecedor.objects.all()
        empresas = Empresas.objects.all()

        context = {

            'consulta':consulta,
            'tipos':tipos_pagamento,
            'status_consultas': status_consulta,
            'fornecedores': fornecedores,
            'empresas': empresas
        }
        if request.method == 'POST':

            conta=Contas_Pagar.objects.get(id=id)


            # Obter os dados dos campos do formulário
            status_conta = request.POST.get('status_conta')
            if status_conta:
                conta.status_conta=status_conta
            empresa_nome = request.POST.get('empresa')
            conta.empresa = Empresas.objects.filter(nome=empresa_nome).first()
            data_vencimento = request.POST.get('data_vencimento')
            if data_vencimento:
                conta.data_vencimento = data_vencimento
            data_pagamento = request.POST.get('data_pagamento')
            if data_pagamento:
                conta.data_pagamento =data_pagamento
            tipo_pagamento = request.POST.get('tipo_pagamento')
            if tipo_pagamento:
                tipo_pagamento=TipoPgto.objects.filter(descricao=tipo_pagamento).first()

            conta.valor =float(request.POST.get('valor'))
            conta.descricao = request.POST.get('descricao')
            fornecedor = request.POST.get('fornecedor')
            conta.fornecedor = Fornecedor.objects.filter(nome=fornecedor).first()
            pedido = request.POST.get ('numero_pedido')
            if pedido:
                conta.pedido = pedido
            conta.outras_informacoes=request.POST.get('outras_informacoes')
            arquivo = request.FILES.get('arquivo')

            if arquivo:
                empresaaws = Empresas.objects.filter(id=1).first()
                s3_client = boto3.client('s3', aws_access_key_id=empresaaws.aws_id,
                                         aws_secret_access_key=empresaaws.aws_chave)
                agora = datetime.now()
                fornecedor_arquivo = fornecedor
                nome_arquivo = f"{fornecedor_arquivo}{agora.strftime('%Y%m%d%H%M%S')}.pdf" if arquivo else None
                s3_client.upload_fileobj(arquivo, 'storagesw', nome_arquivo)
                conta.documento_caminho = f'https://storagesw.s3.amazonaws.com/{nome_arquivo}'

            # Atualizar a lista de pagamentos do Contas_Pagar


            conta.save()
            return redirect('/contas_pagar/')

        return render(request, 'editar_contas_pagar.html', context)

    return redirect('/contas_pagar/')

@login_required()
def dashboard (request):
    
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)


    if 'Gerente' in colaborador.funcao:
        data_atual = date.today()

        pedidos_mes_atual = Pedidos.objects.filter(datapedido__month=data_atual.month, datapedido__year=data_atual.year)
        pedidos_mes_atual_grf = Pedidos.objects.filter(datapedido__month=data_atual.month, datapedido__year=data_atual.year, empresa_pedido__id=1)
        pedidos_mes_atual_tem = Pedidos.objects.filter(datapedido__month=data_atual.month,
                                                       datapedido__year=data_atual.year, empresa_pedido__id=2)

        quantidade_pedidos_mes = pedidos_mes_atual.count()
        lista_pedidos_mes = []

        for pedido in pedidos_mes_atual:

            pedido_dict = {
                'pedido': pedido,
                'cliente': pedido.cliente_pedido.nome,
                'total': format_currency(pedido.valorTotal,'BRL', locale='pt_BR'),
                'empresa': pedido.empresa_pedido,
                'status': pedido.status
            }
            lista_pedidos_mes.append(pedido_dict)
        print ('PEDIDOS MES LISTADOS')


        valor_total_pedidos_mes_grf = pedidos_mes_atual_grf.aggregate(total=Sum('valorTotal')).get('total', 0)
        valor_total_pedidos_mes_grf = format_currency(valor_total_pedidos_mes_grf, 'BRL',
                                                  locale='pt_BR') if valor_total_pedidos_mes_grf else 0.00
        valor_total_pedidos_mes_tem = pedidos_mes_atual_tem.aggregate(total=Sum('valorTotal')).get('total', 0)
        valor_total_pedidos_mes_tem = format_currency(valor_total_pedidos_mes_tem, 'BRL',
                                                  locale='pt_BR') if valor_total_pedidos_mes_tem else 0.00
        valor_total_pedidos_mes = pedidos_mes_atual.aggregate(total=Sum('valorTotal')).get('total', 0)
        valor_total_pedidos_mes = format_currency(valor_total_pedidos_mes, 'BRL', locale='pt_BR') if valor_total_pedidos_mes else 0.00

        print ('Valores Totais Listados')
        ######### Aniversariantes  #############################

        # Filtra os colaboradores pelo dia e mês de aniversário
        data_atual = datetime.now()

        # Filtrar os próximos 5 aniversariantes


        aniversariantes = Colaborador.objects.annotate(
            proximo_aniversario_mes=ExpressionWrapper(
                F('data_nascimento__month') - data_atual.month,
                output_field=fields.IntegerField()
            ),
            proximo_aniversario_dia=ExpressionWrapper(
                F('data_nascimento__day') - data_atual.day,
                output_field=fields.IntegerField()
            )
        ).filter(
            proximo_aniversario_mes__gte=0,
        ).order_by('proximo_aniversario_mes', 'proximo_aniversario_dia')

        print ('Aniversariantes')

        #CALCULA O TOTAL A RECEBER
        total_a_receber = Contas_Receber.objects.filter(
            Q(status_conta='Aberto') | Q(status_conta='Pendente Faturamento') | Q(status_conta='A Vencer') | Q(status_conta='Vencida')
        ).aggregate(total=Sum('valor'))
        total_aberto = total_a_receber['total']
        total_aberto = float(total_aberto) if total_aberto else 0
        total_aberto = format_currency(total_aberto, 'BRL', locale='pt_BR')
        totalvencido = Contas_Receber.objects.filter(status_conta='Vencida').aggregate(totalVencido=Sum('valor'))
        total_vencido = float (totalvencido['totalVencido']) if totalvencido['totalVencido'] else None
        total_vencido = format_currency(total_vencido, 'BRL', locale='pt_BR') if total_vencido else 0

        print ('Resumo Financeiro Efetuado')
        #CALCULO DE TOTAL DE CLIENTES
        numero_clientes = Pedidos.objects.values('cliente_pedido').distinct().count()

        #TOTAL DE PEDIDOS ENTREGUES
        pedidos_finalizados = Pedidos.objects.filter(status='Finalizado').aggregate(totalEntregue=Sum('valorTotal'))
        total_finalizado = (pedidos_finalizados['totalEntregue']) if pedidos_finalizados else 0
        if total_finalizado:
            total_finalizado = format_currency(total_finalizado, 'BRL', locale='pt_BR')
        else: total_finalizado=0.00

        ##############################################################################################################
        #CALCULA A DIFERENÇA SEMANAL A RECEBER
        # Obtém a data atual e a data de 7 dias atrás
        data_atual = datetime.now().date()
        data_7_dias_atras = data_atual - timedelta(days=7)

        # Filtra as contas dos últimos 7 dias e calcula o total do campo "valor"
        contas_ultimos_7_dias = Contas_Receber.objects.filter(data_vencimento__range=[data_7_dias_atras, data_atual])
        total_ultimos_7_dias = contas_ultimos_7_dias.aggregate(total=Sum('valor'))['total']

        # Obtém a data de 14 dias atrás
        data_14_dias_atras = data_atual - timedelta(days=14)

        # Filtra as contas dos 7 dias anteriores (entre 8 e 14 dias atrás) e calcula o total do campo "valor"
        contas_7_dias_anteriores = Contas_Receber.objects.filter(
            data_vencimento__range=[data_14_dias_atras, data_7_dias_atras])
        total_7_dias_anteriores = contas_7_dias_anteriores.aggregate(total=Sum('valor'))['total']
        total_ultimos_7_dias = total_ultimos_7_dias if total_ultimos_7_dias else 0
        # Calcula a diferença percentual
        diferenca_percentual = 0.0
        if total_7_dias_anteriores:
            diferenca_percentual = ((total_ultimos_7_dias - total_7_dias_anteriores) / total_7_dias_anteriores) * 100
            diferenca_percentual = round(diferenca_percentual)

        ##################################################################################################
        # CONSULTA CONTAS A PAGAR DO DIA
        dia_atual = date.today()
        contas_pagar_dia = Contas_Pagar.objects.filter(data_vencimento=dia_atual)
        ##################################################################################################
        # CALCULA % DE INADIMPLENCIA SOBRE AS FATURAS ABERTAS
        # Obtém a data atual e a data de 7 dias atrás
        contas_abertas_a_vencer = Contas_Receber.objects.filter(status_conta__in=['Aberto', 'A Vencer'])
        total_aberto_a_vencer = contas_abertas_a_vencer.aggregate(total=Sum('valor'))['total']

        # Obtém a soma total das contas 'Vencidas'
        contas_vencidas = Contas_Receber.objects.filter(status_conta='Vencida')
        total_vencidas = contas_vencidas.aggregate(total=Sum('valor'))['total']
        lista_contas_vencidas = []
        for conta in contas_vencidas:
            numero_pedido = conta.pedido

            pedidos_inadimplentes = Pedidos.objects.filter(numero=numero_pedido).first()
            if pedidos_inadimplentes:
                cliente = pedidos_inadimplentes.cliente_pedido.nome
                vencimento = conta.data_vencimento
                tipo_pgto = conta.tipo_pgto
                parcela = conta.numero_parcela
                valor = conta.valor
                valor = format_currency(valor, 'BRL', locale='pt_BR')
                total = pedidos_inadimplentes.valorTotal
                total = format_currency(total, 'BRL', locale='pt_BR')
                conta_dict = {
                    'pedido': numero_pedido,
                    'cliente': cliente,
                    'vencimento': vencimento,
                    'tipo_pgto': tipo_pgto,
                    'parcela': conta.numero_parcela,
                    'valor': valor,
                    'total': total
                }
                lista_contas_vencidas.append(conta_dict)

        else:
            pedidos_inadimplentes = None





        # Calcula a porcentagem de inadimplência
        porcentagem_inadimplencia = 0
        if total_aberto_a_vencer:

            if total_vencidas:
                porcentagem_inadimplencia = round((total_vencidas / total_aberto_a_vencer) * 100)
            else: porcentagem_inadimplencia = 0

        print ('Comparativo Financeiro Efetuado')
        ##################################################################################################
        ################## DIFERENÇA DE CLIENTES EM RELAÇÃO A SEMANA ANTERIOR ############################
        ##################################################################################################

        hoje = datetime.today()
        semana_anterior = hoje - timedelta(days=7)

        # Obter a quantidade de clientes diferentes atendidos na semana atual
        clientes_semana_atual = Pedidos.objects.filter(datapedido__gte=hoje.date()).values(
            'cliente_pedido').distinct().count()

        # Obter a quantidade de clientes diferentes atendidos na semana anterior
        clientes_semana_anterior = Pedidos.objects.filter(
            datapedido__range=(semana_anterior.date(), hoje.date() - timedelta(days=1))).values(
            'cliente_pedido').distinct().count()

        # Calcular a diferença em porcentagem
        diferenca_porcentagem_clientes = 0.0
        if clientes_semana_anterior > 0:
            diferenca_porcentagem_clientes = round(((clientes_semana_atual - clientes_semana_anterior) / clientes_semana_anterior) * 100)


        ##################################################################################################
        ################## DIFERENÇA DE PEDIDOS FINALIZADOS EM RELAÇÃO A SEMANA ANTERIOR ############################
        ##################################################################################################
        # Obter a data da semana anterior
        hoje = datetime.today()
        semana_anterior = hoje - timedelta(days=7)

        # Obter a quantidade de pedidos finalizados na semana atual
        pedidos_semana_atual = Pedidos.objects.filter(status='Finalizado', data_entrega_efetiva__gte=hoje.date()).count()

        # Obter a quantidade de pedidos finalizados na semana anterior
        pedidos_semana_anterior = Pedidos.objects.filter(status='Finalizado', data_entrega_efetiva__range=(
        semana_anterior.date(), hoje.date() - timedelta(days=1))).count()

        # Calcular a diferença em porcentagem
        diferenca_porcentagem = 0.0
        diferenca_pedidos = 0.0
        if pedidos_semana_anterior > 0:
            diferenca_pedidos = round( ((pedidos_semana_atual - pedidos_semana_anterior) / pedidos_semana_anterior) * 100)


        ##################################################################################################
        ################## CALCULO DE COMISSIONAMENTO E VENDAS POR VENDEDOR ############################
        ##################################################################################################
        lista = get_lista_comissao(request)

        vendedores = lista['vendedores']
        comissionamento = lista['comissionamento']

        pedidos_status = get_pedidos_status(request)

        nome_usuario = request.user.username

        print('Todas Consultas Efetuadas')

        context = {
            'nome': nome_usuario,
            'total_aberto': total_aberto,
            'total_vencido': total_vencido,
            'numero_clientes': numero_clientes,
            'total_finalizado': total_finalizado,
            'diferenca_percentual': diferenca_percentual if diferenca_percentual else 0,
            'porcentagem_inadimplencia': porcentagem_inadimplencia if porcentagem_inadimplencia else 0,
            'diferenca_clientes': diferenca_porcentagem_clientes if diferenca_porcentagem_clientes else 0,
            'diferenca_pedidos': diferenca_pedidos if diferenca_pedidos else 0,
            'vendedores': vendedores,
            'comissionamento': comissionamento,
            'pedidos_status': pedidos_status,
            'pedidos_mes_atual': pedidos_mes_atual,
            'quantidade_pedidos_mes': quantidade_pedidos_mes,
            'pedidos_mes': lista_pedidos_mes,
            'Valor_total_pedidos_mes': valor_total_pedidos_mes,
            'aniversariantes': aniversariantes if aniversariantes else None,
            'pedidos_inadimplentes': pedidos_inadimplentes if pedidos_inadimplentes else None,
            'contas_vencidas': lista_contas_vencidas if lista_contas_vencidas else None,
            'contas_pagar_dia': contas_pagar_dia if contas_pagar_dia else None,
            'total_mes_grf':valor_total_pedidos_mes_grf,
            'total_mes_tem':valor_total_pedidos_mes_tem

        }

        return render(request,'dashboard/dashboard.html', context)
    else:
        return redirect('/painel/')

def registrar_entrega(request, numero_pedido):

    pedido = Pedidos.objects.filter(numero=numero_pedido).first()
    hora_atual = datetime.now()
    context={

        'pedido':pedido,
        'hora_atual': hora_atual

    }
    if request.method == 'POST':

        registro = numero_pedido
        data_entrega = request.POST.get ('data_entrega')
        nome_usuario = request.user.username
        colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
        nome_recebedor = request.POST.get ('nome_recebedor')
        documento_recebedor = request.POST.get ('documento_recebedor')
        documento_recebedor = ''.join(filter(str.isdigit, documento_recebedor))
        pedido = Pedidos.objects.filter(numero=numero_pedido).first()


        registro = Registra (

            numero_registro=registro,
            data_entrega=data_entrega,
            nome_recebedor=nome_recebedor,
            documento_recebedor=documento_recebedor,
            colaborador=colaborador,
            pedido=pedido,

        )
        registro.save()

        pedido.status = 'Entregue'
        pedido.data_entrega_efetiva = data_entrega
        pedido.save()
        guia = Guia_Reparo.objects.filter(codigo=pedido.numero).first()
        if guia:
            guia.status = Status_Reparo.objects.filter(descricao='Reparo Entregue').first()
            guia.data_entrega = data_entrega
            guia.usuario_entrega = colaborador
            guia.data_alteracao = datetime.now()
            guia.save()

        return redirect('/entrega/')


    return render(request,'registra_entrega.html', context)

def rejeitar_entrega (request, numero_pedido):
    pedido = Pedidos.objects.filter(numero=numero_pedido).first()
    context ={

        'pedido':pedido
    }
    if request.method == 'POST':

        registro = numero_pedido
        data_rejeicao = request.POST.get('data_rejeicao')
        nome_usuario = request.user.username
        colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
        nome_recebedor = request.POST.get ('nome_recebedor')
        documento_recebedor = request.POST.get ('documento_recebedor')
        observacao_rejeicao = request.POST.get('observacao_rejeicao')
        documento_recebedor = ''.join(filter(str.isdigit, documento_recebedor))
        pedido = Pedidos.objects.filter(numero=numero_pedido).first()


        registro = Registra (

            numero_registro=registro,
            data_rejeicao=data_rejeicao,
            nome_recebedor=nome_recebedor,
            documento_recebedor=documento_recebedor,
            colaborador=colaborador,
            pedido=pedido,
            observacao_rejeicao = observacao_rejeicao

        )
        registro.save()

        pedido.status = 'Producao'
        pedido.flag_entrega_rejeitada = 1
        pedido.observacao_entrega_rejeitada = observacao_rejeicao
        pedido.save()

        return redirect('/entrega/')
    return render(request, 'rejeita_entrega.html', context)




def registrar_coleta(request, numero_pedido):
    pedido = Pedidos.objects.filter(numero=numero_pedido).first()
    hora_atual = datetime.now()
    context = {

        'pedido': pedido,
        'hora_atual': hora_atual

    }
    if request.method == 'POST':

        data_entrega = request.POST.get('data_entrega')
        nome_usuario = request.user.username
        colaborador = Colaborador.objects.get(usuario__username=nome_usuario)


        pedido = Pedidos.objects.filter(numero=numero_pedido).first()
        if pedido.status == 'Em Coleta Reparo':
            pedido_reparo = Reparos.objects.filter(pedido=pedido).first()
            pedido.status = 'Producao Reparo'
            pedido_reparo.usuario_coleta = colaborador
            pedido_reparo.data_coleta = data_entrega
            pedido_reparo.save()
            pedido.save()

            guia = Guia_Reparo.objects.filter(codigo=pedido.numero).first()
            if guia:
                guia.status = Status_Reparo.objects.filter(descricao = 'Producao Reparo').first()
                guia.usuario_coleta = colaborador
                guia.data_coleta = data_entrega
                guia.data_alteracao = datetime.now()
                guia.save()
            else:
                pass


            return HttpResponse('Pedido Enviado Para Produção')

    return render(request, 'registra_coleta.html', context)


def get_lista_comissao_old(request):
    # Obtendo a lista de vendedores únicos
    vendedores = Pedidos.objects.exclude(
        Q(Vendedor__isnull=True) | Q(Vendedor=''),
        Q(Vendedor2__isnull=True) | Q(Vendedor2='')
    ).values_list('Vendedor', 'Vendedor2').distinct()


    # Obtendo a soma das comissões para o Vendedor1
    comissoes_vendedor1 = Pedidos.objects.filter(vendedor_comissao__isnull=False).annotate(
        comissao=(F('valorTotal')-F('acerto'))* F('vendedor_comissao') / 100
    ).values('Vendedor').annotate(
        total_comissao=Sum('comissao')
    )

    # Obtendo a soma das comissões para o Vendedor2
    comissoes_vendedor2 = Pedidos.objects.filter(vendedor2_comissao__isnull=False).annotate(
        comissao=F('valorTotal') * F('vendedor2_comissao') / 100
    ).values('Vendedor2').annotate(
        total_comissao=Sum('comissao')
    )

    # Combinando as listas de comissões para os vendedores
    comissoes_totais = list(chain(comissoes_vendedor1, comissoes_vendedor2))
    comissoes_por_vendedor = []

    for comissao in comissoes_totais:
        vendedor = comissao['Vendedor'] if 'Vendedor' in comissao else comissao['Vendedor2']
        total_comissao = comissao['total_comissao']
        total_comissao = format_currency(total_comissao, 'BRL', locale='pt_BR')

        comissoes_por_vendedor.append({
            'vendedor': vendedor,
            'comissao': total_comissao,
        })

    dados = {
        'vendedores': vendedores,
        'comissionamento': comissoes_por_vendedor,
    }


    return (dados)

def get_lista_comissao(request):

    colaboradores_ativos =  Colaborador.objects.filter (status__descricao='ATIVO')

    vendedores =colaboradores_ativos.filter(funcao__in=['Vendedor Interno', 'Vendedor Externo', 'Gerente']).values_list(
        'nome', flat=True)


    # Obtendo a soma das comissões para cada vendedor por mês
    comissoes_por_vendedor = []

    mes_atual = timezone.now().month  # Mês atual

    for vendedor in vendedores:
        comissoes_vendedor1 = Pedidos.objects.filter(
            Vendedor=vendedor,
            vendedor_comissao__isnull=False,
            datapedido__month=mes_atual
        ).annotate(
            comissao=F('valorTotal') * F('vendedor_comissao') / 100
        ).aggregate(
            total_comissao=Sum('comissao')
        )

        comissoes_vendedor2 = Pedidos.objects.filter(
            Vendedor2=vendedor,
            vendedor2_comissao__isnull=False,
            datapedido__month=mes_atual
        ).annotate(
            comissao=F('valorTotal') * F('vendedor2_comissao') / 100
        ).aggregate(
            total_comissao=Sum('comissao')
        )

        total_comissao = (comissoes_vendedor1['total_comissao'] or 0) + (comissoes_vendedor2['total_comissao'] or 0)
        total_comissao = format_currency(total_comissao, 'BRL', locale='pt_BR')

        comissoes_por_vendedor.append({
            'vendedor': vendedor,
            'comissao': total_comissao,
        })

    dados = {
        'vendedores': vendedores,
        'comissionamento': comissoes_por_vendedor,
    }

    return dados

def get_pedidos_status(request):
    abertos = Pedidos.objects.filter(status='Aberto').count()
    producao = Pedidos.objects.filter(status='Producao').count()
    pendente_financeiro = Pedidos.objects.filter(status='PENDENTE_FINANCEIRO').count()
    liberado_entrega = Pedidos.objects.filter(status='Liberado para Entrega').count()
    em_entrega = Pedidos.objects.filter(status='Em Rota de Entrega').count()
    finalizados = Pedidos.objects.filter(status='Em Rota de Entrega').count()

    context = {

        'abertos': abertos if abertos else 0,
        'producao': producao if producao else 0,
        'pendentef': pendente_financeiro if pendente_financeiro else 0,
        'liberado_entrega': liberado_entrega if liberado_entrega else 0,
        'em_entrega':em_entrega if em_entrega else 0,
        'finalizados': finalizados if finalizados else 0,

    }
    return context

@login_required
def comissao (request):
    colaboradores_ativos =  Colaborador.objects.filter (status__descricao='ATIVO')

    vendedores = colaboradores_ativos.filter(funcao__in=['Vendedor Interno', 'Vendedor Externo', 'Gerente']).values_list(
        'nome', flat=True)
    mes_atual = datetime.now().month
    ano_atual = datetime.now().year

    pedidos = Pedidos.objects.filter(datapedido__month=mes_atual, datapedido__year=ano_atual)

    meses = {
        1: 'Janeiro',
        2: 'Fevereiro',
        3: 'Março',
        4: 'Abril',
        5: 'Maio',
        6: 'Junho',
        7: 'Julho',
        8: 'Agosto',
        9: 'Setembro',
        10: 'Outubro',
        11: 'Novembro',
        12: 'Dezembro'
    }

    colaboradores = Colaborador.objects.all()
    vendedores_comissoes = []

    for colaborador in colaboradores:
        vendedor = colaborador.nome
        comissao = 0
        valor_total = 0
        quantidade_pedidos = 0
        total_vendido = 0

        vendedor1_pedidos = Pedidos.objects.filter(Vendedor=colaborador)
        vendedor2_pedidos = Pedidos.objects.filter(Vendedor2=colaborador)

        if vendedor1_pedidos.exists():
            comissao += vendedor1_pedidos.aggregate(comissao_total=Sum('vendedor_comissaov'))[
                'comissao_total']
            valor_total += vendedor1_pedidos.aggregate(total_vendido=Sum('valorTotal'))['total_vendido']
            quantidade_pedidos += vendedor1_pedidos.values('numero').distinct().count()

        if vendedor2_pedidos.exists():
            comissao += \
            vendedor2_pedidos.aggregate(comissao_total=Sum('vendedor2_comissaov'))[
                'comissao_total']
            valor_total += vendedor2_pedidos.aggregate(total_vendido=Sum('valorTotal'))['total_vendido']
            quantidade_pedidos += vendedor2_pedidos.values('numero').distinct().count()

        if quantidade_pedidos >= 1:

            vendedores_comissoes.append({
                'vendedor': vendedor,
                'comissao_total': format_currency(comissao, 'BRL', locale='pt_BR'),
                'total_vendido': format_currency(valor_total, 'BRL', locale='pt_BR'),
                'quantidade_pedidos': quantidade_pedidos,
            })


    context = {
        'nome': request.user.username,
        'vendedores': vendedores,
        'pedidos': pedidos,
        'meses': meses.values(),
        'vendedores_comissoes': vendedores_comissoes

    }


    if request.method == 'POST':
        vendedor = request.POST.get('vendedor')
        #ano_atual = datetime.now().year
        ano_atual = request.POST.get('ano')

        if vendedor:
            mes = request.POST.get('mes')

            meses_dict = {
                'janeiro': 1,
                'fevereiro': 2,
                'março': 3,
                'abril': 4,
                'maio': 5,
                'junho': 6,
                'julho': 7,
                'agosto': 8,
                'setembro': 9,
                'outubro': 10,
                'novembro': 11,
                'dezembro': 12
            }
            mes_numero = meses_dict.get(mes.lower())
            total = 0.00
            if mes_numero is None:
                # Se o mês não for válido, retorne uma resposta de erro ou faça o tratamento adequado.
                # Por exemplo, redirecione para uma página de erro.
                return HttpResponse('Mês inválido')
            dados = Pedidos.objects.filter(
                Q(Vendedor__icontains=vendedor, datapedido__month=mes_numero,datapedido__year=ano_atual ) |
                Q(Vendedor2__icontains=vendedor, datapedido__month=mes_numero, datapedido__year=ano_atual)
            )
            resultado = []
            total = 0
            comissao = 0
            valor_repasse = 0.00
            totalVendido = 0.00
            for pedido in dados:
                if pedido.Vendedor == vendedor:
                    comissao = pedido.vendedor_comissao
                    total += (pedido.vendedor_comissaov)
                    totalVendido += pedido.valorTotal
                    comissao_pedido = (pedido.vendedor_comissaov)
                elif pedido.Vendedor2 == vendedor:
                    if pedido.vendedor2_comissao >=0:
                        comissao = pedido.vendedor2_comissao
                        total += (pedido.vendedor2_comissaov)
                        comissao_pedido = (pedido.vendedor2_comissaov)
                        totalVendido += pedido.valorTotal
                    else:
                        comissao = 0
                        comissao_pedido = 0

                else:
                    comissao = 0
                    comissao_pedido = 0



                resultado.append({
                    'numero': pedido.numero,
                    'status': pedido.status,
                    'valorTotal': format_currency(pedido.valorTotal,'BRL', locale='pt_BR'),
                    'dataPedido': pedido.datapedido.strftime('%d/%m/%Y'),
                    'cliente': pedido.cliente_pedido.nome,
                    'comissao': comissao,
                    'comissao_pedido': format_currency(comissao_pedido,'BRL', locale='pt_BR')


                })

            total=format_currency(total, 'BRL', locale='pt_BR')
            totalVendido = format_currency(totalVendido, 'BRL', locale='pt_BR')
            context = {
                'nome': request.user.username,
                'resultados': resultado,
                'total': total,
                'vendedor':vendedor,
                'vendedores': vendedores,
                'meses':meses.values(),
                'vendedores_comissoes': vendedores_comissoes,
                'totalVendido': totalVendido


            }
            # Retorne a resposta como um JSON
            return render(request, 'comissoes.html', context)



    return render(request,'comissoes.html',context)


def busca_comissao(request, vendedor, mes):

    # Realize a busca do vendedor pelo nome em ambos os campos Vendedor e Vendedor2
    dados = Pedidos.objects.filter(
        Q(Vendedor__icontains=vendedor, datapedido__month=mes) |
        Q(Vendedor2__icontains=vendedor, datapedido__month=mes)
    )

    # Crie uma lista com os dados relevantes, incluindo a comissão correspondente
    resultado = []
    for pedido in dados:
        if pedido.Vendedor == vendedor:
            comissao = pedido.vendedor_comissao
        elif pedido.Vendedor2 == vendedor:
            comissao = pedido.vendedor2_comissao
        else:
            comissao = None

        resultado.append({
            'numero': pedido.numero,
            'valorTotal': pedido.valorTotal,
            'dataPedido': pedido.datapedido.strftime('%d/%m/%Y'),
            'cliente': pedido.cliente_pedido.nome,
            'comissao': comissao,

        })

    context = {

        'resultados':resultado,


    }
    # Retorne a resposta como um JSON
    return render(request,'comissoes.html',context)

def layout_pedido(request):

    if request.method == 'POST':


        numero_pedido= request.POST.get('numero_pedido')
        pedido = Pedidos.objects.get(numero=numero_pedido)
        arquivo = request.FILES['arquivo'].file
        empresaaws = Empresas.objects.filter(id=1).first()
        s3_client = boto3.client('s3', aws_access_key_id=empresaaws.aws_id,
                                 aws_secret_access_key=empresaaws.aws_chave)
        agora = datetime.now()
        nome_arquivo = f"{numero_pedido}{agora.strftime('%Y%m%d%H%M%S')}.pdf"

        s3_client.upload_fileobj(arquivo, 'storagesw', nome_arquivo)

        try:
            layout_existente = Layout.objects.get(numero_pedido=pedido)
            # Deletar o layout existente
            layout_existente.delete()
        except Layout.DoesNotExist:
            pass
        pedido.caminho_layout = f"https://storagesw.s3.amazonaws.com/{nome_arquivo}"
        pedido.save()
        layout = Layout(
            numero_pedido=pedido,
            arquivo_path = f'https://storagesw.s3.amazonaws.com/{nome_arquivo}',
            arquivo_nome=nome_arquivo
        )
        layout.save()
        return redirect ('/meuspedidos/')


    return render (request, 'cadastra_layout.html')
def mensagem_sucesso_layout(request):

    return render (request,'mensagem_sucesso.html')

def producao(request):

    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)

    pedidos = Pedidos.objects.filter(status__icontains ='Producao')
    total_producao = pedidos.count()


    data_atual = date.today()
    dia_anterior = data_atual - timedelta(days=1)
    pedidos_atrasados = Pedidos.objects.filter(prazo_entrega__lt=dia_anterior, status__icontains ='Producao')
    total_atrasados = pedidos_atrasados.count()

    prazo_entrega_limite_7dias = data_atual + timedelta(days=7)
    pedidos_faltando_7dias = pedidos.filter(prazo_entrega__lte=prazo_entrega_limite_7dias, prazo_entrega__gte=data_atual)
    t_pedidos_faltando_7dias = pedidos_faltando_7dias.count()


    prazo_entrega_limite_3dias = data_atual + timedelta(days=3)
    pedidos_faltando_3dias = pedidos.filter(prazo_entrega__lte=prazo_entrega_limite_3dias, prazo_entrega__gte=data_atual)
    t_pedidos_faltando_3dias = pedidos_faltando_3dias.count()



    context = {

        'pedidos': pedidos,
        'total_producao': total_producao,
        'total_atrasado': total_atrasados,
        't_pedidos_faltando_3dias': t_pedidos_faltando_3dias,
        't_pedidos_faltando_7dias': t_pedidos_faltando_7dias,
        'nome':nome_usuario,
        'funcao': colaborador.funcao

    }



    return render (request, 'producao.html', context)

def painel_expedicao(request):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    if colaborador.empresa.flag_pagamento == 1:
        return redirect('/dologin/')
    elif not colaborador:
        return redirect('/dologin/')
    else:
        pass



    pedidos_em_expedicao = Pedidos.objects.filter(status__icontains='Expedicao').order_by('numero')
    total_em_expedicao = pedidos_em_expedicao.count()
    total_liberados =  Pedidos.objects.filter(Q(status='Liberado para Entrega')|Q(status='Liberado Coleta')|Q(status='Liberado Retirada Cliente')).count()
    total_rota = Pedidos.objects.filter(status='Em Rota de Entrega').count()
    total_financeiro = Pedidos.objects.filter(status='Financeiro').count()
    total_retirada = Pedidos.objects.filter(status='Expedicao', tipo_entrega='retirada').count()
    zonas = set()
    bairros = set()
    cidades = set()

    pedidos = Pedidos.objects.filter(status='Producao')
    total_producao = Pedidos.objects.filter(status='Producao').count()

    data_atual = date.today()
    dia_anterior = data_atual - timedelta(days=1)
    pedidos_atrasados = Pedidos.objects.filter(prazo_entrega__lt=dia_anterior, status='Producao')
    total_atrasados = pedidos_atrasados.count()

    prazo_entrega_limite_7dias = data_atual + timedelta(days=7)
    pedidos_faltando_7dias = Pedidos.objects.filter(prazo_entrega__lte=prazo_entrega_limite_7dias,
                                                     prazo_entrega__gte=data_atual, status='Producao')

    t_pedidos_faltando_7dias = pedidos_faltando_7dias.count()
    prazo_entrega_limite_3dias = data_atual + timedelta(days=3)
    pedidos_faltando_3dias = Pedidos.objects.filter(prazo_entrega__lte=prazo_entrega_limite_3dias,
                                                     prazo_entrega__gte=data_atual, status='Producao')
    t_pedidos_faltando_3dias = pedidos_faltando_3dias.count()


    pedidos_liberados = pedidos_em_expedicao.filter(status='Expedicao')
    for pedido in pedidos_liberados:
        pedido.bairro_entrega = pedido.bairro_entrega.upper()
        pedido.cidade_entrega = pedido.cidade_entrega.upper()
        testa_rota = Rota.objects.filter(pedido=pedido).first()
        if testa_rota is None:
            if pedido.tipo_entrega == 'propria' and 'RJ' in pedido.uf_entrega:

                regiao = Bairro.objects.filter(nome=pedido.bairro_entrega, cidade=pedido.cidade_entrega).first()
                if regiao:

                    rota = Rota(

                        pedido = pedido,
                        regiao = regiao.regiao

                    )
                    rota.save()
                else:
                    regiao = Bairro.objects.filter(cidade=pedido.cidade_entrega).first()
                    if regiao:
                        rota = Rota(

                            pedido=pedido,
                            regiao=regiao.regiao

                        )
                        rota.save()

            elif pedido.tipo_entrega == 'Transportadora' or pedido.uf_entrega != 'RJ':
                rota = Rota(

                    pedido = pedido,
                    regiao = 'transportadora'

                )
                rota.save()
            elif pedido.tipo_entrega == 'retirada':
                rota = Rota(

                    pedido=pedido,
                    regiao='retirada'

                )
                rota.save()

            else:
                rota = Rota(

                    pedido=pedido,
                    regiao='transportadora'

                )
                rota.save()
        else:
            pass

    rota_1 = 0
    rota_2 = 0
    rota_3 = 0
    rota_4 = 0
    rota_5 = 0
    rota_6 = 0
    rota_7 = 0
    rota_8 = 0
    rota_transportadora = 0

    pedidos_rota = Rota.objects.all()
    for pedido in pedidos_rota:
        if pedido.pedido.status != 'Expedicao':
            Rota.objects.filter(pedido__numero=pedido.pedido.numero).delete()
        else:
            if pedido.regiao == '1':
                rota_1 += 1
            elif pedido.regiao == '2':
                rota_2 += 1
            elif pedido.regiao == '3':
                rota_3 += 1
            elif pedido.regiao == '4':
                rota_4 += 1
            elif pedido.regiao == '5':
                rota_5 += 1
            elif pedido.regiao == '6':
                rota_6 += 1
            elif pedido.regiao == '7':
                rota_7 += 1
            elif pedido.regiao == '8':
                rota_8 += 1
            elif pedido.regiao == 'transportadora':
                rota_transportadora += 1





    for p in pedidos_em_expedicao:
        zona = p.zona_entrega
        cidade = p.cidade_entrega
        bairro = p.bairro_entrega

        if zona not in zonas:
            if zona:
                zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)
        if bairro not in bairros:
            bairros.add(bairro)


    context = {

        'pedidos': pedidos_em_expedicao,
        'total_expedicao': total_em_expedicao,
        'total_liberados': total_liberados,
        'total_rota': total_rota,
        'zonas': zonas,
        'cidades': cidades,
        'bairros': bairros,
        'colaborador': colaborador,
        'total_producao': total_producao,
        'total_atrasado': total_atrasados,
        't_pedidos_faltando_3dias': t_pedidos_faltando_3dias,
        't_pedidos_faltando_7dias': t_pedidos_faltando_7dias,
        'nome': nome_usuario,
        'funcao': colaborador.funcao,
        'pedidos_producao': pedidos,
        'total_rota_1': rota_1,
        'total_rota_2': rota_2,
        'total_rota_3': rota_3,
        'total_rota_4': rota_4,
        'total_rota_5': rota_5,
        'total_rota_6': rota_6,
        'total_rota_7': rota_7,
        'total_rota_8': rota_8,
        'total_rota_transportadora': rota_transportadora,
        'total_financeiro': total_financeiro,
        'total_retirada': total_retirada
    }

    return render(request, 'expedicao.html', context)

def expedicao_total (request):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)

    pedidos_em_expedicao = Pedidos.objects.filter(status__icontains='Expedicao').order_by('numero')
    zonas = set()
    cidades = set()
    status_l = set()

    for pedido in pedidos_em_expedicao:
        zona = pedido.zona_entrega
        cidade = pedido.cidade_entrega
        status = pedido.status

        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)
        if status not in status_l:
            status_l.add(status)


    context = {

        'pedidos': pedidos_em_expedicao,
        'zonas':zonas,
        'cidades':cidades,
        'status_choices': status_l,
        'colaborador': colaborador


    }

    return render(request, 'expedicao_total.html', context)

def expedicao_financeiro (request):
    pedidos_em_expedicao = Pedidos.objects.filter(status='Financeiro').order_by('numero')
    zonas = set()
    cidades = set()
    status_l = set()

    for pedido in pedidos_em_expedicao:
        zona = pedido.zona_entrega
        cidade = pedido.cidade_entrega
        status = pedido.status

        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)
        if status not in status_l:
            status_l.add(status)


    context = {

        'pedidos': pedidos_em_expedicao,
        'zonas':zonas,
        'cidades':cidades,
        'status_choices': status_l


    }

    return render(request, 'expedicao_total.html', context)

def expedicao_liberados (request):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    pedidos_em_expedicao = Pedidos.objects.filter(Q(status='Liberado para Entrega') | Q(status='Liberado Coleta')
                                                  | Q(status='Liberado Retirada Cliente')).order_by('numero')
    zonas = set()
    cidades = set()
    status_l = set()



    for pedido in pedidos_em_expedicao:
        zona = pedido.zona_entrega
        cidade = pedido.cidade_entrega
        status = pedido.status

        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)
        if status not in status_l:
            status_l.add(status)


    context = {

        'pedidos': pedidos_em_expedicao,
        'zonas':zonas,
        'cidades':cidades,
        'status_choices': status_l,
        'colaborador':colaborador


    }

    return render(request, 'expedicao_total.html', context)


def expedicao_rota (request):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    pedidos_em_expedicao = Pedidos.objects.filter(Q(status='Em Rota de Entrega')|Q(status='Em Coleta Reparo')).order_by('numero')
    zonas = set()
    cidades = set()
    status_l = set()



    for pedido in pedidos_em_expedicao:
        zona = pedido.zona_entrega
        cidade = pedido.cidade_entrega
        status = pedido.status

        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)
        if status not in status_l:
            status_l.add(status)


    context = {

        'pedidos': pedidos_em_expedicao,
        'zonas':zonas,
        'cidades':cidades,
        'status_choices': status_l,
        'colaborador': colaborador


    }

    return render(request, 'expedicao_total.html', context)

def producao_total (request):
    pedidos_em_producao = Pedidos.objects.filter(status__icontains='Producao')
    zonas = set()
    cidades = set()



    for pedido in pedidos_em_producao:
        zona = pedido.zona_entrega
        cidade = pedido.cidade_entrega

        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)


    context = {

        'pedidos': pedidos_em_producao,
        'zonas':zonas,
        'cidades':cidades,


    }

    return render(request, 'producao_total.html', context)

def producao_7 (request):


    data_atual = date.today()
    prazo_entrega_limite_7dias = data_atual + timedelta(days=7)
    pedidos_em_producao = Pedidos.objects.filter(prazo_entrega__lte=prazo_entrega_limite_7dias, prazo_entrega__gte=data_atual, status__icontains='Producao')


    pedidos = pedidos_em_producao.order_by('numero')

    zonas = set()
    cidades = set()

    for pedido in pedidos:
        zona = pedido.zona_entrega
        cidade = pedido.cidade_entrega

        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)


    context = {

        'pedidos': pedidos,
        'zonas':zonas,
        'cidades':cidades,



    }

    return render(request, 'producao_total.html', context)
def producao_3 (request):


    data_atual = date.today()
    prazo_entrega_limite_3dias = data_atual + timedelta(days=3)
    pedidos_em_producao = Pedidos.objects.filter(prazo_entrega__lte=prazo_entrega_limite_3dias, prazo_entrega__gte=data_atual, status__icontains='Producao')



    pedidos = pedidos_em_producao.order_by('numero')

    zonas = set()

    cidades = set()

    for pedido in pedidos:
        zona = pedido.zona_entrega
        cidade = pedido.cidade_entrega
        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)


    context = {

        'pedidos': pedidos,
        'zonas':zonas,
        'cidades':cidades,


    }

    return render(request, 'producao_total.html', context)

def producao_atrasados (request):
    data_atual = date.today()
    dia_anterior = data_atual - timedelta(days=1)
    pedidos_atrasados = Pedidos.objects.filter(prazo_entrega__lt=dia_anterior, status='Producao')



    pedidos = pedidos_atrasados.order_by('numero')

    zonas = set()
    cidades = set()

    for pedido in pedidos:
        zona = pedido.zona_entrega
        cidade = pedido.cidade_entrega
        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)


    context = {

        'pedidos': pedidos,
        'zonas':zonas,
        'cidades':cidades,

    }

    return render(request, 'producao_total.html', context)

def filtrar_producao(request):
    # Obter os dados da requisição POST
    data = json.loads(request.body)


    # Filtrar os pedidos com base nos critérios
    pedidos = Pedidos.objects.all()
    filtro_q = Q()
    if data['pedido']:
        filtro_q &= Q(numero__icontains=data['pedido'])

    if data['zona']:
        filtro_q &= Q(zona_entrega=data['zona'])

    if data['cidade']:
        filtro_q &= Q(cidade_entrega=data['cidade'])

    if data['dataInicial'] and data['dataFinal']:
        filtro_q &= Q(prazo_entrega__range=[data['dataInicial'], data['dataFinal']])

    pedidos = pedidos.filter(filtro_q)
    pedidos = pedidos.filter(Q(status='Producao') | Q(status='Expedicao') | Q(status='Liberado Para Entrega') |
                    Q(status='Liberado Para Entrega') | Q(status='Em Rota de Entrega'))
    # Criar uma lista de dicionários com os dados dos pedidos
    pedidos_data = []
    for pedido in pedidos:
        pedido_data = {
            'numero': pedido.numero,
            'cliente_pedido': pedido.cliente_pedido.nome,
            'prazo_entrega': pedido.prazo_entrega.strftime('%d/%m/%Y'),
            'Vendedor': pedido.Vendedor,
            'zona_entrega': pedido.zona_entrega,
            'bairro_entrega': pedido.bairro_entrega,
            'cidade_entrega': pedido.cidade_entrega,
            'uf_entrega': pedido.uf_entrega,
            'status': pedido.status

        }
        pedidos_data.append(pedido_data)

    # Retornar os dados como resposta JSON
    return JsonResponse({'pedidos': pedidos_data})

def filtrar_expedicao(request):
    # Obter os dados da requisição POST
    data = json.loads(request.body)

    # Filtrar os pedidos com base nos critérios
    pedidos = Pedidos.objects.filter(Q(status='Expedicao') | Q(status='Liberado para Entrega') | Q(status='Liberado Coleta')
                                                  | Q(status='Liberado Retirada Cliente')).order_by('status')
    filtro_q = Q()
    if data['pedido']:
        filtro_q &= Q(numero__icontains=data['pedido'])

    if data['status']:
        filtro_q &= Q(status=data['status'])


    if data['zona']:
        filtro_q &= Q(zona_entrega=data['zona'])

    if data['cidade']:
        filtro_q &= Q(cidade_entrega=data['cidade'])

    if data['dataInicial'] and data['dataFinal']:
        filtro_q &= Q(prazo_entrega__range=[data['dataInicial'], data['dataFinal']])

    pedidos = pedidos.filter(filtro_q)

    # Criar uma lista de dicionários com os dados dos pedidos
    pedidos_data = []
    for pedido in pedidos:
        pedido_data = {
            'numero': pedido.numero,
            'cliente_pedido': pedido.cliente_pedido.nome,
            'prazo_entrega': pedido.prazo_entrega.strftime('%d/%m/%Y'),
            'Vendedor': pedido.Vendedor,
            'zona_entrega': pedido.zona_entrega,
            'bairro_entrega': pedido.bairro_entrega,
            'cidade_entrega': pedido.cidade_entrega,
            'uf_entrega': pedido.uf_entrega,
            'status': pedido.status

        }
        pedidos_data.append(pedido_data)


    # Retornar os dados como resposta JSON
    return JsonResponse({'pedidos': pedidos_data})

def regiao_1 (request):
    pedidos_regiao_1 = Rota.objects.filter(regiao='1')
    zonas = set()
    cidades = set()
    status = set()



    for pedido in pedidos_regiao_1:
        zona = pedido.pedido.zona_entrega
        cidade = pedido.pedido.cidade_entrega
        statusp = pedido.pedido.status

        if statusp not in status:
            status.add(statusp)
        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)


    context = {

        'pedidos': pedidos_regiao_1,
        'zonas':zonas,
        'cidades':cidades,
        'status_choices':status


    }

    return render(request, 'expedicao_regioes.html', context)

def regiao_2 (request):
    pedidos_regiao_2 = Rota.objects.filter(regiao='2')
    zonas = set()
    cidades = set()
    status = set()



    for pedido in pedidos_regiao_2:
        zona = pedido.pedido.zona_entrega
        cidade = pedido.pedido.cidade_entrega
        statusp = pedido.pedido.status

        if statusp not in status:
            status.add(statusp)
        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)


    context = {

        'pedidos': pedidos_regiao_2,
        'zonas':zonas,
        'cidades':cidades,
        'status_choices':status


    }

    return render(request, 'expedicao_regioes.html', context)

def regiao_3 (request):
    pedidos_regiao = Rota.objects.filter(regiao='3')
    zonas = set()
    cidades = set()
    status = set()



    for pedido in pedidos_regiao:
        zona = pedido.pedido.zona_entrega
        cidade = pedido.pedido.cidade_entrega
        statusp = pedido.pedido.status

        if statusp not in status:
            status.add(statusp)
        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)


    context = {

        'pedidos': pedidos_regiao,
        'zonas':zonas,
        'cidades':cidades,
        'status_choices':status


    }

    return render(request, 'expedicao_regioes.html', context)

def regiao_4 (request):
    pedidos_regiao = Rota.objects.filter(regiao='4')
    zonas = set()
    cidades = set()
    status = set()



    for pedido in pedidos_regiao:
        zona = pedido.pedido.zona_entrega
        cidade = pedido.pedido.cidade_entrega
        statusp = pedido.pedido.status

        if statusp not in status:
            status.add(statusp)
        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)


    context = {

        'pedidos': pedidos_regiao,
        'zonas':zonas,
        'cidades':cidades,
        'status_choices':status


    }

    return render(request, 'expedicao_regioes.html', context)

def regiao_5 (request):
    pedidos_regiao = Rota.objects.filter(regiao='5')
    zonas = set()
    cidades = set()
    status = set()



    for pedido in pedidos_regiao:
        zona = pedido.pedido.zona_entrega
        cidade = pedido.pedido.cidade_entrega
        statusp = pedido.pedido.status

        if statusp not in status:
            status.add(statusp)
        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)


    context = {

        'pedidos': pedidos_regiao,
        'zonas':zonas,
        'cidades':cidades,
        'status_choices':status


    }

    return render(request, 'expedicao_regioes.html', context)

def regiao_6 (request):
    pedidos_regiao = Rota.objects.filter(regiao='6')
    zonas = set()
    cidades = set()
    status = set()



    for pedido in pedidos_regiao:
        zona = pedido.pedido.zona_entrega
        cidade = pedido.pedido.cidade_entrega
        statusp = pedido.pedido.status

        if statusp not in status:
            status.add(statusp)
        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)


    context = {

        'pedidos': pedidos_regiao,
        'zonas':zonas,
        'cidades':cidades,
        'status_choices':status


    }

    return render(request, 'expedicao_regioes.html', context)

def regiao_7 (request):
    pedidos_regiao = Rota.objects.filter(regiao='7')
    zonas = set()
    cidades = set()
    status = set()



    for pedido in pedidos_regiao:
        zona = pedido.pedido.zona_entrega
        cidade = pedido.pedido.cidade_entrega
        statusp = pedido.pedido.status

        if statusp not in status:
            status.add(statusp)
        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)


    context = {

        'pedidos': pedidos_regiao,
        'zonas':zonas,
        'cidades':cidades,
        'status_choices':status


    }

    return render(request, 'expedicao_regioes.html', context)

def regiao_8 (request):
    pedidos_regiao = Rota.objects.filter(regiao='8')
    zonas = set()
    cidades = set()
    status = set()



    for pedido in pedidos_regiao:
        zona = pedido.pedido.zona_entrega
        cidade = pedido.pedido.cidade_entrega
        statusp = pedido.pedido.status

        if statusp not in status:
            status.add(statusp)
        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)


    context = {

        'pedidos': pedidos_regiao,
        'zonas':zonas,
        'cidades':cidades,
        'status_choices':status


    }

    return render(request, 'expedicao_regioes.html', context)

def regiao_transportadora (request):
    pedidos_regiao = Rota.objects.filter(regiao='transportadora')
    zonas = set()
    cidades = set()
    status = set()



    for pedido in pedidos_regiao:
        zona = pedido.pedido.zona_entrega
        cidade = pedido.pedido.cidade_entrega
        statusp = pedido.pedido.status

        if statusp not in status:
            status.add(statusp)
        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)


    context = {

        'pedidos': pedidos_regiao,
        'zonas':zonas,
        'cidades':cidades,
        'status_choices':status


    }

    return render(request, 'expedicao_regioes.html', context)

def expedicao_retirada (request):
    pedidos_regiao = Rota.objects.filter(regiao='retirada')
    zonas = set()
    cidades = set()
    status = set()



    for pedido in pedidos_regiao:
        zona = pedido.pedido.zona_entrega
        cidade = pedido.pedido.cidade_entrega
        statusp = pedido.pedido.status

        if statusp not in status:
            status.add(statusp)
        if zona not in zonas:
            zonas.add(zona)
        if cidade not in cidades:
            cidades.add(cidade)


    context = {

        'pedidos': pedidos_regiao,
        'zonas':zonas,
        'cidades':cidades,
        'status_choices':status


    }

    return render(request, 'expedicao_regioes.html', context)

def filtrar_regiao(request):
    # Obter os dados da requisição POST
    data = json.loads(request.body)

    # Filtrar os pedidos com base nos critérios
    pedidos = Pedidos.objects.filter(Q(status='Expedicao') | Q(status='Liberado para Entrega') | Q(status='Liberado Coleta')
                                                  | Q(status='Liberado Retirada Cliente')).order_by('status')
    filtro_q = Q()
    if data['pedido']:
        filtro_q &= Q(numero__icontains=data['pedido'])

    if data['status']:
        filtro_q &= Q(status=data['status'])


    if data['zona']:
        filtro_q &= Q(zona_entrega=data['zona'])

    if data['cidade']:
        filtro_q &= Q(cidade_entrega=data['cidade'])

    if data['dataInicial'] and data['dataFinal']:
        filtro_q &= Q(prazo_entrega__range=[data['dataInicial'], data['dataFinal']])

    pedidos = pedidos.filter(filtro_q)

    # Criar uma lista de dicionários com os dados dos pedidos
    pedidos_data = []
    for pedido in pedidos:
        pedido_data = {
            'numero': pedido.numero,
            'cliente_pedido': pedido.cliente_pedido.nome,
            'prazo_entrega': pedido.prazo_entrega.strftime('%d/%m/%Y'),
            'Vendedor': pedido.Vendedor,
            'zona_entrega': pedido.zona_entrega,
            'bairro_entrega': pedido.bairro_entrega,
            'cidade_entrega': pedido.cidade_entrega,
            'uf_entrega': pedido.uf_entrega,
            'status': pedido.status

        }
        pedidos_data.append(pedido_data)


    # Retornar os dados como resposta JSON
    return JsonResponse({'pedidos': pedidos_data})

def select_regiao_entrega(request):
    # Obter os dados da requisição POST
    if request.method == 'POST':
        try:
            data = json.loads(request.body)  # Carrega o JSON do corpo da requisição
            pedidos_marcados = data['pedidos_marcados']

            # Processar cada número de pedido
            for numero_pedido in pedidos_marcados:
                # Aqui você pode realizar as operações que desejar com o número do pedido

                pedido = Pedidos.objects.filter(numero=numero_pedido).first()
                if pedido.status == 'Expedicao':

                    pedido.status = 'Financeiro'
                    pedido.save()
                else:
                    pass

            return JsonResponse({'message': 'Pedidos recebidos com sucesso!'})
        except json.JSONDecodeError:
                return JsonResponse({'message': 'Erro ao processar os pedidos'}, status=400)

        return render(request, 'sua_template.html')


def get_resumo_comissao_mes(request):
    data = json.loads(request.body)  # Obtém o valor do parâmetro 'data' da URL
    data_mes = data.get('mes')
    mes_atual = datetime.now().month
    ano_atual = datetime.now().year
    meses_dict = {
        'janeiro': 1,
        'fevereiro': 2,
        'marco': 3,
        'abril': 4,
        'maio': 5,
        'junho': 6,
        'julho': 7,
        'agosto': 8,
        'setembro': 9,
        'outubro': 10,
        'novembro': 11,
        'dezembro': 12
    }
    data_mes = meses_dict.get(data_mes.lower())



    if data_mes:
        # Se um mês específico for fornecido na solicitação, use-o para filtrar os pedidos
        mes = int(data_mes)
    else:
        # Caso contrário, use o mês atual
        mes = mes_atual

    colaboradores = Colaborador.objects.all()
    vendedores_comissoes = []

    for colaborador in colaboradores:
        vendedor = colaborador.nome
        comissao = 0
        valor_total = 0
        quantidade_pedidos = 0

        vendedor1_pedidos = Pedidos.objects.filter(Vendedor=colaborador, datapedido__month=mes, datapedido__year=ano_atual)
        vendedor2_pedidos = Pedidos.objects.filter(Vendedor2=colaborador, datapedido__month=mes, datapedido__year=ano_atual)

        if vendedor1_pedidos.exists():
            comissao += vendedor1_pedidos.aggregate(comissao_total=Sum('vendedor_comissaov'))[
                'comissao_total']
            valor_total += vendedor1_pedidos.aggregate(total_vendido=Sum('valorTotal'))['total_vendido']
            quantidade_pedidos += vendedor1_pedidos.values('numero').distinct().count()

        if vendedor2_pedidos.exists():
            comissao += vendedor2_pedidos.aggregate(comissao_total=Sum('vendedor2_comissaov'))[
                'comissao_total']
            valor_total += vendedor2_pedidos.aggregate(total_vendido=Sum('valorTotal'))['total_vendido']
            quantidade_pedidos += vendedor2_pedidos.values('numero').distinct().count()

        if quantidade_pedidos >= 1:
            vendedores_comissoes.append({
                'vendedor': vendedor,
                'comissao_total': format_currency(comissao, 'BRL', locale='pt_BR'),
                'total_vendido': format_currency(valor_total, 'BRL', locale='pt_BR'),
                'quantidade_pedidos': quantidade_pedidos,
            })
    print(vendedores_comissoes)


    return JsonResponse(vendedores_comissoes, safe=False)


def itens_edit_nfe (request, numero_pedido):
    pedido = Pedidos.objects.filter(numero=numero_pedido).first()
    itens = Itens_Pedido.objects.filter(pedido=numero_pedido)
    context = {
        'pedido':pedido,
        'itens':itens
    }
    if request.method == 'POST':
        nomes_itens = request.POST.getlist('nfe_nome_item')
        quantidades = request.POST.getlist('nf_qtd_item')
        precos = request.POST.getlist('nfe_preco_item')
        totais_itens = request.POST.getlist('nfe_item_total')
        for item, nome, quantidade, preco, total_item in zip(itens, nomes_itens, quantidades, precos, totais_itens):
            item.nome_nf = nome
            item.quantidade_nf = quantidade
            item.preco_nf = preco
            total_item =  re.sub(r'^R\$|\.', '', total_item).replace(',', '.')
            item.total_item_nf = total_item
            item.flag_editado = 1
            item.save()

        return render(request, 'editar_itens_nf.html', context)


    return render(request,'editar_itens_nf.html', context)


def emitir_nota(request):
    data = json.loads(request.body)

    pedido_numero = data.get('numero_pedido')
    inscricao_estadual = data.get('inscricao')
    pedido = Pedidos.objects.get(numero=pedido_numero)
    if pedido.cliente_pedido.inscricao_estadual:
        pass
    else:
        pedido.cliente_pedido.inscricao_estadual = inscricao_estadual
    informacoes_adicionais = data.get('informacoes_complementares')
    if informacoes_adicionais:
        pedido.informacoes_adicionais = informacoes_adicionais

    valor_desconto_item = 0
    if pedido.descontototal > 0:
        total_itens = len(Itens_Pedido.objects.filter(pedido=pedido_numero, tipo='PADRAO'))
        valor_desconto_total = pedido.descontototal
        valor_desconto_item = valor_desconto_total / total_itens



    print('Inscricao Estadual:', pedido.cliente_pedido.inscricao_estadual)

    if pedido.cliente_pedido.nome_contato:
        if pedido.cliente_pedido.inscricao_estadual is None:
            return JsonResponse({"error": "Falta cadastro da inscrição estadual do CNPJ"})
        else:
            pass
    ref = {
        "ref": (str(pedido.empresa_pedido.cnpj) + str(pedido.numero)).zfill(44)
    }
   # url = "https://api.focusnfe.com.br/"
    #if (pedido.empresa_pedido.nome == 'TEM INDUSTRIA E COMERCIO DE TAPETES LTDA'):
     #   token = "OK0V8A4AKMOi5ywOFOUljMR8Ew4S3Imf"
    #else: token = "Dzrt7d64Eq30parVmw1xT8swfZ2uMejS"

    token = pedido.empresa_pedido.token_nota
    url = pedido.empresa_pedido.url_nota
    # Construir os dados para envio à API de emissão de nota fiscal
    # Dados da Nota Fiscal
    nfe = {}
    itens = {}
    # Obtém a data e hora atual no fuso horário local (UTC)
    data_hora_utc = datetime.utcnow()

    # Calcula o deslocamento para o fuso horário brasileiro (UTC-3)
    deslocamento_horario = timedelta(hours=-3)

    # Adiciona o deslocamento ao horário UTC para obter a data e hora no fuso horário brasileiro
    data_hora_brasileira = data_hora_utc + deslocamento_horario

    # Formata a data e hora no formato ISO (string)
    data_hora_formatada_brasileira = data_hora_brasileira.isoformat()


    nfe["natureza_operacao"] = "Venda"
    nfe["forma_pagamento"] = "0"
    nfe["data_emissao"] = data_hora_formatada_brasileira
    nfe["tipo_documento"] = "1"
    if pedido.uf_entrega == 'RJ' and pedido.cliente_pedido.estado == 'RJ':
        nfe["local_destino"] = "1"
    else:
        nfe["local_destino"] = "2"
    nfe["finalidade_emissao"] = "1"
    nfe["consumidor_final"] = "1"
    nfe["presenca_comprador"] = "0"
    nfe["cnpj_emitente"] = pedido.empresa_pedido.cnpj
    nfe["logradouro_emitente"] = pedido.empresa_pedido.endereco
    nfe["numero_emitente"] = pedido.empresa_pedido.endereco_numero
    nfe["bairro_emitente"] = pedido.empresa_pedido.bairro.nome
    nfe["municipio_emitente"] = pedido.empresa_pedido.cidade.nome
    nfe["uf_emitente"] = "RJ"
    nfe["cep_emitente"] = pedido.empresa_pedido.cep
    nfe["telefone_emitente"] = pedido.empresa_pedido.telefone
    nfe["inscricao_estadual_emitente"] = pedido.empresa_pedido.inscricao_estadual
    nfe["regime_tributario_emitente"] = "1"
    if len(pedido.cliente_pedido.documento) >= 12:
        nfe["cnpj_destinatario"] = pedido.cliente_pedido.documento
    else:
        nfe["cpf_destinatario"] = pedido.cliente_pedido.documento
    nfe["nome_destinatario"] = pedido.cliente_pedido.nome
    nfe["logradouro_destinatario"] = pedido.cliente_pedido.endereco
    nfe["numero_destinatario"] = pedido.cliente_pedido.numero_endereco
    nfe["bairro_destinatario"] = pedido.cliente_pedido.bairro
    nfe["municipio_destinatario"] = pedido.cliente_pedido.cidade
    nfe["uf_destinatario"] = pedido.cliente_pedido.estado
    nfe["cep_destinatario"] = pedido.cliente_pedido.CEP
    nfe["complemento_destinatario"] = pedido.cliente_pedido.complemento
    if(pedido.cliente_pedido.CEP != pedido.cep_entrega):
        nfe["logradouro_entrega"]= pedido.endereco_entrega
        nfe["numero_entrega"] = pedido.numero_end_entrega
        nfe["complemento_entrega"] = pedido.complemento
        nfe["bairro_entrega"] = pedido.bairro_entrega
        nfe["municipio_entrega"] = pedido.cidade_entrega
        nfe["uf_entrega"] = pedido.uf_entrega
        nfe["cep_entrega"] = pedido.cep_entrega

    if (pedido.cliente_pedido.nome_contato):
        if (pedido.cliente_pedido.inscricao_estadual == 'ISENTO'):
            nfe["indicador_inscricao_estadual_destinatario"] = "2"

        elif (pedido.cliente_pedido.inscricao_estadual):
            nfe["indicador_inscricao_estadual_destinatario"] = "1"
            nfe["inscricao_estadual_destinatario"]= pedido.cliente_pedido.inscricao_estadual

    else:
        nfe["indicador_inscricao_estadual_destinatario"] = "2"
    nfe["icms_base_calculo"] = "0"
    nfe["icms_valor_total"] = "0"
    nfe["icms_valor_total_desonerado"] = "0"
    nfe["icms_base_calculo_st"] = "0"
    nfe["icms_valor_total_st"] = "0"
    nfe["valor_produtos"] = pedido.valorTotal
    if pedido.frete_valor > 0:
        nfe["valor_frete"] = pedido.frete_valor
    else:
        nfe["valor_frete"] = "0"
    nfe["valor_seguro"] = "0"
    nfe["valor_desconto"] = "0"
    nfe["valor_total_ii"] = "0"
    nfe["valor_ipi"] = "0"
    nfe["valor_pis"] = "0"
    nfe["valor_cofins"] = "0"
    nfe["valor_outras_despesas"] = "0"
    nfe["valor_total"] = pedido.valorTotal
    nfe["modalidade_frete"] = "0"

    if pedido.frete_fornecedor:
        nfe["volumes"] = []
        volumes = {}
        volumes["quantidade"] = pedido.frete_volume
        volumes["peso_bruto"] = pedido.frete_peso
        volumes["peso_liquido"] = pedido.frete_peso
        nfe["volumes"].append(volumes)
        nfe["nome_transportador"] = pedido.frete_fornecedor.nome
        if pedido.frete_fornecedor.documento:
            nfe["cnpj_transportador"] = pedido.frete_fornecedor.documento
    if pedido.informacoes_adicionais:
        nfe["informacoes_adicionais_contribuinte"] = pedido.informacoes_adicionais
    else:
        nfe["informacoes_adicionais_contribuinte"] = None

    itens_pedido = Itens_Pedido.objects.filter(Q( pedido=pedido.numero, tipo='PADRAO') | Q(pedido=pedido.numero, tipo='ADICIONAL'))
    numero_item = 1
    nfe["items"] = []  # Inicializa a lista de itens
    for item in itens_pedido:
        produto = item.produto
        itens = {}
        itens["numero_item"] = numero_item
        itens["codigo_produto"] = produto.codigo
        itens["descricao"] = f"{item.nome} ({item.comprimento} x {item.largura})"
        if pedido.uf_entrega == 'RJ' and pedido.cliente_pedido.estado == 'RJ':
            itens["cfop"] = "5101"
        else:
            itens["cfop"] = "6101"
        itens["unidade_comercial"] = "UN"
        itens["quantidade_comercial"] = item.quantidade
        itens["valor_unitario_comercial"] = item.preco - item.desconto
        itens["valor_bruto"] = item.total_item
        itens["valor_desconto"] = 0.00
        itens["unidade_tributavel"] = "UN"
        itens["codigo_ncm"] = produto.ncm
        itens["quantidade_tributavel"] = item.quantidade
        itens["valor_unitario_tributavel"] = item.preco - item.desconto
        itens["inclui_no_total"] = "1"
        itens["icms_origem"] = "0"
        itens["icms_situacao_tributaria"] = "400"
        itens["pis_situacao_tributaria"] = "99"
        itens["cofins_situacao_tributaria"] = "99"
        nfe["items"].append(itens)
        numero_item += 1







    url = "https://api.focusnfe.com.br/v2/nfe"
    ref = {

        "ref": str(pedido_numero) + str(pedido.operacao_nota),
        "token": token
    }
    print(url,nfe)
    ref = urllib.parse.urlencode(ref)



    r = requests.post(url, params=ref, data=json.dumps(nfe), auth=(token, ""))
    print(r.status_code, r.text)
    response = r.json()

    if 'status' in response:
        status = response['status']

        if r.status_code == 202:
            while status == 'processando_autorizacao':
                json_response = r.json()
                ref = json_response['ref']
                pedido.referencia_nfe = ref

                url_consulta = 'https://api.focusnfe.com.br/v2/nfe/'
                completa = 'completa=1'
                response = requests.get(url_consulta+ref, params=completa, auth=(token, "") )
                print(response.status_code, response.text)
                resposta = response.json()
                status = resposta['status']

        if response.status_code == 200:
            json_response = response.json()
            status_nfe = json_response['status']
            if status_nfe == 'autorizado':
                xml = json_response['caminho_xml_nota_fiscal']
                nfe = json_response['caminho_danfe']
                nfe_numero = json_response['numero']
                url_focus = 'https://api.focusnfe.com.br'
                pedido.caminho_xml = url_focus+xml
                pedido.caminho_nfe = url_focus+nfe
                pedido.numero_nfe = nfe_numero
                pedido.valor_nota = pedido.valorTotal
                hoje = date.today()
                pedido.dataEmissao_nfe=hoje
                if pedido.flag_nfe_antecipada >0:
                    pedido.status = 'Producao'
                else:
                    pass

                pedido.save()
                response_data = {
                    'nfe': pedido.caminho_nfe,
                    'xml': pedido.caminho_xml
                }
                return JsonResponse(response_data)
            else:

                json_response = response.json()
                error_code = json_response['status_sefaz']
                error_message = json_response['mensagem_sefaz']

                response_data = {
                    'code': error_code,
                    'error': error_message
                }

                return JsonResponse(response_data, status=response.status_code)
        else:

            json_response = response.json()
            error_code = json_response['codigo']
            error_message = json_response['mensagem']

            response_data = {
                'code': error_code,
                'error': error_message
            }

            return JsonResponse(response_data, status=response.status_code)
    else:

        error_code = response['codigo']
        error_message = response['mensagem']

        response_data = {
            'code': error_code,
            'error': error_message
        }
        return JsonResponse(response_data)



    # Mostra na tela o código HTTP da requisição e a mensagem de retorno da API

def cancelar_nota(request):
    data = json.loads(request.body)
    pedido_numero = data.get('numero_pedido')
    pedido = Pedidos.objects.filter(numero=pedido_numero).first()
    if pedido:
        ref = pedido.referencia_nfe
        url = 'https://api.focusnfe.com.br/v2/nfe/'
        #if (pedido.empresa_pedido.nome == 'TEM INDUSTRIA E COMERCIO DE TAPETES LTDA'):
        token = str (pedido.empresa_pedido.token_nota)
        #else: token = "Dzrt7d64Eq30parVmw1xT8swfZ2uMejS"
        justificativa = {}
        justificativa["justificativa"] = "Erro na emissão referente a nota fiscal"
        r = requests.delete(url + ref, data=json.dumps(justificativa), auth=(token, ""))
        print(r.status_code, r.text)
        response_data = r.json()
        status = response_data['status_sefaz']
        if status == '135':
            xml = response_data['caminho_xml_cancelamento']
            pedido.caminho_nfe = ""
            pedido.referencia_nfe = ""
            url = "https://api.focusnfe.com.br"
            pedido.caminho_xml = url+xml
            pedido.numero_nfe = None
            pedido.operacao_nota += 1
            pedido.save()


        return JsonResponse(response_data)

    else:
        print ('Pedido Não Encontrado')
        return HttpResponse('Pedido Não Encontrado')



@csrf_exempt
@login_required
def entregas (request):
    data_atual = datetime.now().date()
    pedidos_em_entrega = Pedidos.objects.filter(
        Q(status='Em Rota de Entrega') |Q(status='EM COLETA REPARO') | Q(status='Entregue', data_entrega_efetiva__date=data_atual)
    ).order_by('status')
    print(data_atual)


    context = {

        'pedidos': pedidos_em_entrega

    }
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    if 'Logistica' or 'Gerente' in colaborador.funcao:

        return render(request, 'painel_entregas.html', context)

    else:
        return redirect('painel')


def get_comissao(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        numero_pedido = data['numero_pedido']
        vendedor1 = data['vendedor1']
        vendedor2 = data['vendedor2']
        documento_cliente = data['cliente']
        documento = ''.join(filter(str.isdigit, documento_cliente))
        cliente = Cliente.objects.get(documento=documento)
        tipo_pedido = cliente.tipo_cliente.descricao
        itens = data['itens']
        vendedor = Colaborador.objects.get(nome=vendedor1)
        vendedor_comissao =0
        menor_comissao = 100
        ultimo_quantidade=0
        ultimo_largura=0
        ultimo_comprimento=0
        if vendedor2:
            vendedor2 = Colaborador.objects.get(nome=vendedor2)
            menor_comissao2 = 100
            vendedor2_comissao = 0
        else:
            menor_comissao2 = 0
            vendedor2 = None

        pedido_total = 0.00

        if tipo_pedido == 'PADRAO':


            for item in itens:
                quantidade = item['quantidade']
                quantidade = int(quantidade)
                comprimento = item['comprimento']
                comprimento = float(comprimento)
                comprimento = round(comprimento, 2)
                largura = item['largura']
                largura = float (largura)
                largura=round(largura,2)
                valorUnitario = item['valorUnitario']
                valorUnitario = float (valorUnitario)
                valorUnitario = round(valorUnitario,2)
                desconto = item['desconto']
                tipo_produto = item['tipo_produto']

                if (tipo_produto == 'PADRAO' or tipo_produto == 'ADICIONAL'):

                    if desconto:
                        desconto = float(desconto)
                    else:
                        desconto =0.00
                    produto = Produtos.objects.filter(descricao=item['produto']).first()

                    item_valor_total = quantidade * valorUnitario - desconto
                    item_valor_un = valorUnitario - desconto
                    pedido_total += round((item_valor_total), 2)

                    if produto.medida_padrao == 'SIM':
                        if produto.tipo_medida == 'm2':
                            if 'MPC' in produto.descricao:
                                item_medida = largura * float (produto.tamanho_padrao)

                            else:
                                item_medida = comprimento * float (produto.tamanho_padrao)
                        elif produto.tipo_medida == 'pc':
                            item_medida = quantidade * float (produto.tamanho_padrao)
                        else:
                            item_medida = comprimento

                    else:

                        if (produto.tipo_medida == 'linear'):
                            item_medida = comprimento
                        else:
                            item_medida = comprimento * largura

                    item_medida = round(item_medida,3)


                    item_valor_base = ( item_valor_un + desconto ) / item_medida
                    item_valor_base = round(item_valor_base,2)

                    print (item_valor_base, item_valor_un, item_medida)



                    if(vendedor.funcao == 'Vendedor Interno' or vendedor.funcao == 'Gerente'):
                        base_comissao=[
                            (produto.preco_base_comissao1, produto.percentual_comissao1),
                            (produto.preco_base_comissao2, produto.percentual_comissao2),
                            (produto.preco_base_comissao3, produto.percentual_comissao3_interno),
                            (produto.preco_base_comissao4, produto.percentual_comissao4_interno),
                            (produto.preco_base_comissao5, produto.percentual_comissao5_interno),
                        ]

                        for base, percentual in base_comissao:
                            if item_valor_base >= base:
                                vendedor_comissao = percentual

                        menor_comissao = vendedor_comissao if menor_comissao>= vendedor_comissao else menor_comissao

                    if (vendedor.funcao == 'Vendedor Externo'):
                        base_comissao = [
                            (produto.preco_base_comissao1, produto.percentual_comissao1),
                            (produto.preco_base_comissao2, produto.percentual_comissao2),
                            (produto.preco_base_comissao3, produto.percentual_comissao3_externo),
                            (produto.preco_base_comissao4, produto.percentual_comissao4_externo),
                            (produto.preco_base_comissao5, produto.percentual_comissao5_externo),
                        ]

                        for base, percentual in base_comissao:
                            if item.valor_base >= base:
                                vendedor_comissao = percentual

                        menor_comissao = vendedor_comissao if menor_comissao >= vendedor_comissao else menor_comissao

                    if vendedor2:

                        if (vendedor2.funcao == 'Vendedor Interno' or vendedor2.funcao == 'Gerente'):
                            base_comissao = [
                                (produto.preco_base_comissao1, produto.percentual_comissao1),
                                (produto.preco_base_comissao2, produto.percentual_comissao2),
                                (produto.preco_base_comissao3, produto.percentual_comissao3_interno),
                                (produto.preco_base_comissao4, produto.percentual_comissao4_interno),
                                (produto.preco_base_comissao5, produto.percentual_comissao5_interno),
                            ]

                            for base, percentual in base_comissao:
                                if item_valor_base >= base:
                                    vendedor2_comissao = percentual

                            menor_comissao2 = vendedor2_comissao if menor_comissao2 >= vendedor2_comissao else menor_comissao2

                        if (vendedor2.funcao == 'Vendedor Externo'):
                            base_comissao = [
                                (produto.preco_base_comissao1, produto.percentual_comissao1),
                                (produto.preco_base_comissao2, produto.percentual_comissao2),
                                (produto.preco_base_comissao3, produto.percentual_comissao3_externo),
                                (produto.preco_base_comissao4, produto.percentual_comissao4_externo),
                                (produto.preco_base_comissao5, produto.percentual_comissao5_externo),
                            ]

                            for base, percentual in base_comissao:
                                if item_valor_base >= base:
                                    vendedor2_comissao = percentual

                            menor_comissao2 = vendedor2_comissao if menor_comissao2 >= vendedor2_comissao else menor_comissao2





                if (tipo_produto == 'BRINDE'):
                    pass
        elif tipo_pedido == 'REVENDA':
            revenda = Revenda.objects.get(cliente=cliente)
            nivel_revenda = revenda.nivel



            pedido_total = 0
            for item in itens:
                quantidade = item['quantidade']
                quantidade = int(quantidade)
                comprimento = item['comprimento']
                comprimento = float(comprimento)
                largura = item['largura']
                largura = float (largura)
                valorUnitario = item['valorUnitario']
                valorUnitario = float (valorUnitario)
                desconto = item['desconto']
                tipo_produto = item['tipo_produto']
                preco_minimo = 999999
                if (tipo_produto == 'PADRAO'):

                    if desconto:
                        desconto = float(desconto)
                    else:
                        desconto =0.00
                    produto = Produtos_Revenda.objects.filter(descricao=item['produto']).first()
                    print(produto)
                    item_valor_total = quantidade * valorUnitario - desconto
                    item_valor_unitario = valorUnitario - desconto
                    pedido_total += round((item_valor_total), 2)

                    if produto.medida_padrao == 'SIM':
                        if produto.tipo_medida == 'm2':
                            if 'MPC' in produto.descricao:
                                item_medida = largura * float (produto.tamanho_padrao)

                            else:
                                item_medida = comprimento * float (produto.tamanho_padrao)
                        elif produto.tipo_medida == 'pc':
                            item_medida = quantidade * float (produto.tamanho_padrao)
                        else:
                            item_medida = comprimento

                    else:

                        if (produto.tipo_medida == 'linear'):
                            item_medida = comprimento
                        else:
                            item_medida = comprimento * largura

                    item_medida = round(item_medida,3)
                    item_valor_base = item_valor_unitario / item_medida
                    item_valor_base = round(item_valor_base, 2)
                    print(item_valor_base, item_valor_total, item_medida)

                    if nivel_revenda == 1:
                        preco_minimo = produto.preco_base_1
                    elif nivel_revenda == 2:
                        print(produto.preco_base_2)
                        preco_minimo = produto.preco_base_2
                    elif nivel_revenda == 3:
                        preco_minimo = produto.preco_base_3
                    elif nivel_revenda == 4:
                        preco_minimo = produto.preco_base_4


                    if item_valor_base >= preco_minimo:
                        vendedor_comissao = produto.comissao
                        menor_comissao = vendedor_comissao
                    else:
                        vendor_comissao = 0.00
                        menor_comissao = 0



                    if vendedor2:
                        menor_comissao2 = menor_comissao



            if (tipo_produto == 'BRINDE'):
                    pass

        response_data = {
            'comissaoPercentual1': menor_comissao if menor_comissao else 0.00,
            'totalComissao1': pedido_total*menor_comissao/100 if menor_comissao else 0.00,
            'comissaoPercentual2': menor_comissao2 if menor_comissao2 else 0.00,
            'totalComissao2': pedido_total*menor_comissao2/100 if menor_comissao2 else 0.00,
        }
        print (response_data)
        pedido = Temp_Pedidos.objects.filter(pedido=numero_pedido).first()
        if pedido:
            pedido.vendedor1 = vendedor
            pedido.comissao1 = menor_comissao if menor_comissao else 0.00
            if vendedor2:
                pedido.vendedor2 = vendedor2
                pedido.comissao2 = menor_comissao2 if menor_comissao2 else 0.00
            pedido.total_parcial = pedido_total
            pedido.save()

        return JsonResponse(response_data)

def get_price(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        print (data)
        vendedor1 = data['vendedor1']
        vcomissao1 = float(data['comissao1'])
        adicional = float(data['adicional'])
        adicional_item = 0.00
        documento_cliente = data['documento']
        documento = ''.join(filter(str.isdigit, documento_cliente))
        cliente = Cliente.objects.filter(documento=documento).first()
        tipo_pedido = cliente.tipo_cliente.descricao
        vendedor = Colaborador.objects.get(nome=vendedor1)
        preco_unitario = 0.00
        preco_tabela_metro = 0.00
        quantidade = data['quantidade']
        quantidade = int(quantidade)


        if tipo_pedido == 'PADRAO':

            tipo_produto = data['tipo_produto']
            produto = Produtos.objects.filter(descricao=data['produto']).first()

            if tipo_produto == 'PADRAO' or tipo_produto == 'ADICIONAL':
                if 'Interno' or 'Gerente' in vendedor.funcao:

                    if vcomissao1 == 1:
                        preco_tabela_metro = produto.preco_base_comissao1

                    elif vcomissao1 == 3:
                        preco_tabela_metro = produto.preco_base_comissao2

                    elif vcomissao1 == 5:
                        preco_tabela_metro = produto.preco_base_comissao3
                    elif vcomissao1 == 6:
                        preco_tabela_metro = produto.preco_base_comissao4
                    elif vcomissao1 == 7:
                        preco_tabela_metro = produto.preco_base_comissao5
                    else:
                        return JsonResponse({'status': 'error', 'message': 'Valor do Item não encontrado na tabela!'})
                    preco_tabela_metro = round(preco_tabela_metro,2)
                    if 'MPC' in produto.descricao:
                        comprimento = produto.tamanho_padrao
                        largura = data['largura']
                        largura = float(largura)
                        largura = round(largura, 3)
                    elif 'MPL' in produto.descricao:
                        largura = produto.tamanho_padrao
                        comprimento = data['comprimento']
                        comprimento = float(comprimento)
                        comprimento = round(comprimento, 3)
                    else:
                        comprimento = data['comprimento']
                        comprimento = float(comprimento)
                        comprimento = round(comprimento, 3)
                        largura = data['largura']
                        largura = float(largura)
                        largura = round(largura, 3)

                    if produto.tipo_medida == 'm2':

                        m2 = comprimento * largura
                        m2 = round(m2,3)
                        preco_unitario = preco_tabela_metro * m2
                        print(preco_unitario)
                        preco_unitario = int(round(preco_unitario, 2))
                        print('Nome do Produto: ', produto.descricao,'Preço Tabela:',preco_tabela_metro,'Metro Quadrado: ',m2 ,'Preço Unitario:',preco_unitario )

                    elif produto.tipo_medida =='linear':
                        linear = comprimento
                        preco_unitario = linear * preco_tabela_metro
                        preco_unitario = round(preco_unitario,2)
                        print('Nome do Produto: ', produto.descricao, 'Preço Tabela:', preco_tabela_metro,
                              'Metro Linear: ', linear, 'Preço Unitario:', preco_unitario)
                    elif produto.tipo_medida == 'pc':
                        pc = 1
                        preco_unitario = pc * preco_tabela_metro
                        preco_unitario = (round(preco_unitario, 2))
                        print('Nome do Produto: ', produto.descricao, 'Preço Tabela:', preco_tabela_metro,
                              'Peça: ', pc, 'Preço Unitario:', preco_unitario)

                    preco_unitario = int(preco_unitario)
                    total = (quantidade * (preco_unitario + adicional))

                    total = format_currency(total,'BRL', locale='pt_BR')
                    response_data = {
                        "valor_unitario": "R$ {:.2f}".format(preco_unitario),
                        "total": total
                    }
                    return JsonResponse(response_data)
                elif tipo_produto == 'BRINDE':

                    valor_unitario = 0.00
                    total = 0.00
                    response_data = {
                        "valor_unitario": "R$ {:.2f}".format(valor_unitario),
                        "total": "R$ {:.2f}".format(total)
                    }
                    return JsonResponse(response_data)
                else:
                    return JsonResponse({'status': 'error', 'message': 'Tipo Item Não Encontrado'})


        elif tipo_pedido =='REVENDA':
            revenda = Revenda.objects.filter(cliente=cliente).first()
            nivel_revenda = revenda.nivel
            tipo_produto = data['tipo_produto']
            produto = Produtos_Revenda.objects.filter(descricao=data['produto']).first()
            preco_tabela_metro = 0.00

            if tipo_produto == 'BRINDE':
                valor_unitario = 0.00
                total = 0.00
                total = format_currency(total, 'BRL', locale='pt_BR')
                response_data = {
                    "valor_unitario": "R$ {:.2f}".format(valor_unitario),
                    "total": total
                }
                return JsonResponse(response_data)
            elif tipo_produto == 'ADICIONAL' or tipo_produto == 'PADRAO':
                if nivel_revenda == 1:
                    preco_tabela_metro = produto.preco_base_1
                elif nivel_revenda == 2:
                    preco_tabela_metro = produto.preco_base_2
                elif nivel_revenda == 3:
                    preco_tabela_metro = produto.preco_base_3
                elif nivel_revenda == 4:
                    preco_tabela_metro = produto.preco_base_4
                else:
                    return JsonResponse({'status': 'error', 'message': 'Valor do Item não encontrado na tabela de Revenda!'})

                if 'MPC' in produto.descricao:
                    comprimento = produto.tamanho_padrao
                    largura = data['largura']
                    largura = float(largura)
                    largura = round(largura, 2)
                elif 'MPL' in produto.descricao:
                    largura = produto.tamanho_padrao
                    comprimento = data['comprimento']
                    comprimento = float(comprimento)
                    comprimento = round(comprimento, 2)
                else:
                    comprimento = data['comprimento']
                    comprimento = float(comprimento)
                    comprimento = round(comprimento, 2)
                    largura = data['largura']
                    largura = float(largura)
                    largura = round(largura, 2)

                if produto.tipo_medida == 'm2':

                    m2 = comprimento * largura
                    preco_unitario = preco_tabela_metro * m2
                    preco_unitario = round(preco_unitario, 2)
                    print('Nome do Produto: ', produto.descricao, 'Preço Tabela:', preco_tabela_metro, 'Metro Quadrado: ',
                          m2, 'Preço Unitario:', preco_unitario)

                elif produto.tipo_medida == 'linear':
                    linear = comprimento
                    preco_unitario = linear * preco_tabela_metro
                    preco_unitario = round(preco_unitario, 2)
                    print('Nome do Produto: ', produto.descricao, 'Preço Tabela:', preco_tabela_metro,
                          'Metro Linear: ', linear, 'Preço Unitario:', preco_unitario)
                elif produto.tipo_medida == 'pc':
                    pc = 1
                    preco_unitario = pc * preco_tabela_metro
                    preco_unitario = round(preco_unitario, 2)
                    print('Nome do Produto: ', produto.descricao, 'Preço Tabela:', preco_tabela_metro,
                          'Peça: ', pc, 'Preço Unitario:', preco_unitario)

                else:
                    return JsonResponse(
                        {'status': 'error', 'message': 'Tipo do Produto Inválido'})

                preco_unitario = int(preco_unitario)
                total = (quantidade * (preco_unitario + adicional))
                total = format_currency(total, 'BRL', locale='pt_BR')

                response_data = {
                    "valor_unitario": "R$ {:.2f}".format(preco_unitario),
                    "total": total
                }
                return JsonResponse(response_data)

        else:
            return JsonResponse(
                {'status': 'error', 'message': 'Tipo do Pedido Inválido'})

    else:
        return JsonResponse(
            {'status': 'error', 'message': 'Método Inválido!'})

def ajusta_comissao(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        numero_pedido = data.get('numero_pedido')
        print(numero_pedido)
        vendedor = data.get('vendedor')
        comissao = data.get('comissao')
        justificativa = data.get('justificativa')
        pedido = Pedidos.objects.get(numero=numero_pedido)
        pedido.justificativa_comissao = justificativa
        if pedido.Vendedor == vendedor:
            pedido.vendedor_comissao = comissao
            pedido.vendedor_comissaov = pedido.valorTotalParcial * float(comissao) / 100
        elif pedido.Vendedor2 == vendedor:
            pedido.vendedor2_comissao = comissao
            pedido.vendedor2_comissaov = pedido.valorTotalParcial * float(comissao) / 100
        pedido.save()
        # Exemplo de resposta JSON
        response_data = {'status': 'success', 'message': 'Pedido atualizado com sucesso'}
        return JsonResponse(response_data)

        # Tratar outros métodos HTTP, se necessário
    return JsonResponse({'status': 'error', 'message': 'Método não permitido'})

def transportadora (request):
    if request.method == 'POST':
        data = json.loads(request.body)
        numero_pedido = data.get('numero_pedido')
        transportadora = data.get('transportadora')
        valor_frete = data.get('valor_frete')
        peso = data.get('peso')
        volumes = data.get('volumes')
        conta_pagar_opcao = data.get('conta_pagar')
        pedido = Pedidos.objects.get(numero=numero_pedido)
        pedido.frete_valor = valor_frete
        pedido.frete_volume = volumes
        pedido.frete_peso = peso
        fornecedor = Fornecedor.objects.get(nome=transportadora)
        pedido.frete_fornecedor = fornecedor

        if conta_pagar_opcao == 'sim':
            conta = Contas_Pagar (

                pedido = numero_pedido,
                descricao = 'FRETE TRANSPORTADORA',
                numero_parcela = 1,
                total_parcelas = 1,
                fornecedor = fornecedor,
                valor = valor_frete
            )
            conta.save()
        pedido.frete_flag = 0
        pedido.save()
        response_data = '{"message": "Pedido Atualizado Com Sucesso!"}'

    try:
        response_data_dict = json.loads(response_data)
        return JsonResponse(response_data_dict)
    except json.JSONDecodeError as e:
        # Trate o erro caso a string JSON seja inválida
        error_message = "Erro ao decodificar JSON: " + str(e)
        return JsonResponse({"error": error_message}, status=400)

def registra_recebimento (request):
    if request.method == 'POST':
        data = json.loads(request.body)
        numero_pedido = data.get('numeroPedido')
        banco_id = data.get('bancoId')
        conta_id = data.get('contaID')
        valor_recebido = float(data.get('valorPago'))
        data_pagamento = data.get('dataPagamento')
        forma_pagamento = data.get('tipoPagamento')
        conta = Contas_Receber.objects.filter(id=conta_id).first()
        banco = Bancos.objects.filter(id=banco_id).first()
        conta.status_conta = 'Pago'
        conta.data_pagamento = data_pagamento
        conta.tipo_pgto = forma_pagamento
        conta.banco = banco
        if valor_recebido >= conta.valor:

            conta.valor = valor_recebido
            conta.save()

        else:
            return JsonResponse({"error": 'Valor Não Confere'}, status=400)

        response_data = '{"message": "Pedido Atualizado Com Sucesso!"}'

    try:
        response_data_dict = json.loads(response_data)
        return JsonResponse(response_data_dict)
    except json.JSONDecodeError as e:
        # Trate o erro caso a string JSON seja inválida
        error_message = "Erro ao decodificar JSON: " + str(e)
        return JsonResponse({"error": error_message}, status=400)

def registra_pagamento(request):
    if request.method == 'POST':
        id_conta = request.POST.get('id_conta')
        data_pagamento = request.POST.get('data_pagamento')
        arquivo_comprovante = request.FILES.get('comprovante')
        banco_nome = request.POST.get('banco')
        print(id_conta)
        print(banco_nome)
        conta = Contas_Pagar.objects.filter(id=id_conta).first()
        print(conta)
        banco = Bancos.objects.filter(nome=banco_nome, empresa=conta.empresa).first()
        print(banco)

        if banco:

            if arquivo_comprovante:
                empresaaws = Empresas.objects.filter(id=1).first()
                s3_client = boto3.client('s3', aws_access_key_id=empresaaws.aws_id,
                                         aws_secret_access_key=empresaaws.aws_chave)
                agora = datetime.now()
                nome_arquivo = f"COMPROVANTE{agora.strftime('%Y%m%d%H%M%S')}.pdf"

                s3_client.upload_fileobj(arquivo_comprovante, 'storagesw', nome_arquivo)


                conta.comprovante_caminho = f'https://storagesw.s3.amazonaws.com/{nome_arquivo}'
                conta.status = 'Pago'
                conta.data_pagamento = data_pagamento
                conta.banco = banco
                conta.save()
                return JsonResponse({'mensagem': 'Pagamento registrado com sucesso!'})
        else:
            return JsonResponse({'mensagem': 'Banco Não Encontrado, ou Banco não existe para a empresa do pagamento!'})


def Teste():
    pedido_numero = '121124021319402'
    pedido = Pedidos.objects.get(numero=pedido_numero)
    id_conta = 22459
    conta = Contas_Receber.objects.get(id=id_conta, status_conta='A Vencer')

    url_cliente = "https://api.asaas.com/v3/customers"

    if pedido.comprador_email:
        email = pedido.comprador_email
    else:
        email = pedido.cliente_pedido.email

    payload = {
        "name": pedido.cliente_pedido.nome,
        "cpfCnpj": pedido.cliente_pedido.documento,
        "email": email,
        "phone": pedido.cliente_pedido.telefone2,
        "mobilePhone": pedido.cliente_pedido.telefone1,
        "address": pedido.cliente_pedido.endereco,
        "addressNumber": pedido.cliente_pedido.numero_endereco,
        "complement": pedido.cliente_pedido.complemento,
        "postalCode": pedido.cliente_pedido.CEP,
        "province": pedido.cliente_pedido.bairro,
        "externalReference": pedido.cliente_pedido.id,
        "notificationDisabled": False
    }
    print(payload)

    if pedido.empresa_pedido.nome == 'TEM INDUSTRIA E COMERCIO DE TAPETES LTDA':
        headers = {
            "accept": "application/json",
            "content-type": "application/json",
            "access_token": "$aact_YTU5YTE0M2M2N2I4MTliNzk0YTI5N2U5MzdjNWZmNDQ6OjAwMDAwMDAwMDAwMDAzMjk5NjA6OiRhYWNoXzkxM2IzMTAyLTViMDQtNDFlYy1hMTg5LTJhMmY3ODMxZjJjMw=="
        }

    else:
        headers = {
            "accept": "application/json",
            "content-type": "application/json",
            "access_token": "$aact_YTU5YTE0M2M2N2I4MTliNzk0YTI5N2U5MzdjNWZmNDQ6OjAwMDAwMDAwMDAwMDAzMzI5Njk6OiRhYWNoX2ViZWQ5N2E3LTRhNTAtNGIxMC1iM2NhLTJjODk3NWJkOTVhNg=="
        }

    response = requests.post(url_cliente, json=payload, headers=headers)
    json_response = response.json()
    print(json_response)
    if response.status_code == 200:

        # Converter a resposta em um objeto JSON usando o módulo json do Python
        json_response = response.json()
        print(json_response)
        if json_response['id']:
            # Obter o valor do campo "id"
            id_cliente = json_response['id']
            # Use o valor do "id" como necessário
            print(id_cliente)  # Deve imprimir o ID do cliente retornado pela API
        else:
            print(f"Erro na resposta da API: {response.status_code} - {response.text}")



    else:
        # Caso ocorra algum erro na resposta, você pode tratá-lo aqui
        print(f"Erro na resposta da API: {response.status_code} - {response.text}")

    url_boleto = "https://asaas.com/api/v3/payments"
    due_date_str = conta.data_vencimento.strftime('%Y-%m-%d')

    payload = {
        "billingType": "BOLETO",
        "interest": {"value": 1},
        "fine": {"value": 1},
        "value": conta.valor,
        "dueDate": due_date_str,
        "description": pedido_numero,
        "customer": id_cliente,
        "postalService": False,
        "externalReference": pedido_numero,

    }

    response = requests.post(url_boleto, json=payload, headers=headers)
    print(response.text)
    if response.status_code == 200:
        json_response = response.json()
        bank_slip_url = json_response['bankSlipUrl']
        cobranca_id = json_response['id']
        conta.caminho_boleto = bank_slip_url
        conta.id_cobranca = cobranca_id
        conta.tipo_pgto = 'BOLETO'

        conta.save()

        response_data = {
            'bankSlipUrl': bank_slip_url
        }

        # Retorna a resposta em formato JSON
    return JsonResponse(response_data)


@csrf_exempt
def gerar_boleto(request):
    data = json.loads(request.body)
    print(data)
    pedido_numero = data.get('numero_pedido')
    id_conta = data.get('id_conta')

    pedido = Pedidos.objects.get(numero=pedido_numero)
    conta = Contas_Receber.objects.get(id=id_conta, status_conta='A Vencer')

    url_cliente = "https://api.asaas.com/v3/customers"

    if pedido.comprador_email:
        email = pedido.comprador_email
    else:
        email = pedido.cliente_pedido.email

    payload = {
        "name": pedido.cliente_pedido.nome,
        "cpfCnpj": pedido.cliente_pedido.documento,
        "email": email,
        "phone": pedido.cliente_pedido.telefone2,
        "mobilePhone": pedido.cliente_pedido.telefone1,
        "address": pedido.cliente_pedido.endereco,
        "addressNumber": pedido.cliente_pedido.numero_endereco,
        "complement": pedido.cliente_pedido.complemento,
        "postalCode": pedido.cliente_pedido.CEP,
        "province": pedido.cliente_pedido.bairro,
        "externalReference": pedido.cliente_pedido.id,
        "notificationDisabled": False
    }
    print(payload)


    if pedido.empresa_pedido.nome == 'TEM INDUSTRIA E COMERCIO DE TAPETES LTDA':
        headers = {
            "accept": "application/json",
            "content-type": "application/json",
            "access_token": "$aact_YTU5YTE0M2M2N2I4MTliNzk0YTI5N2U5MzdjNWZmNDQ6OjAwMDAwMDAwMDAwMDAzMjk5NjA6OiRhYWNoXzkxM2IzMTAyLTViMDQtNDFlYy1hMTg5LTJhMmY3ODMxZjJjMw=="
        }

    else:
        headers = {
            "accept": "application/json",
            "content-type": "application/json",
            "access_token": "$aact_YTU5YTE0M2M2N2I4MTliNzk0YTI5N2U5MzdjNWZmNDQ6OjAwMDAwMDAwMDAwMDAzMzI5Njk6OiRhYWNoX2ViZWQ5N2E3LTRhNTAtNGIxMC1iM2NhLTJjODk3NWJkOTVhNg=="
        }

    response = requests.post(url_cliente, json=payload, headers=headers)

    if response.status_code == 200:
        # Converter a resposta em um objeto JSON usando o módulo json do Python
        json_response = response.json()
        print (json_response)
        if json_response['id']:
        # Obter o valor do campo "id"
            id_cliente = json_response['id']
            # Use o valor do "id" como necessário
            print(id_cliente)  # Deve imprimir o ID do cliente retornado pela API
        else:
            print(f"Erro na resposta da API: {response.status_code} - {response.text}")



    else:
        # Caso ocorra algum erro na resposta, você pode tratá-lo aqui
        print(f"Erro na resposta da API: {response.status_code} - {response.text}")


    #url_boleto = "https://asaas.com/api/v3/payments"
    url_boleto ="https://api.asaas.com/v3/payments"
    due_date_str = conta.data_vencimento.strftime('%Y-%m-%d')

    payload = {
        "billingType": "BOLETO",
        "interest": {"value": 1},
        "fine": {"value": 1},
        "value": conta.valor,
        "dueDate": due_date_str,
        "description": pedido_numero,
        "customer": id_cliente,
        "postalService": False,
        "externalReference": pedido_numero,

    }

    response = requests.post(url_boleto, json=payload, headers=headers)
    print(response.text)
    if response.status_code == 200:
        json_response = response.json()
        bank_slip_url = json_response['bankSlipUrl']
        cobranca_id = json_response['id']
        conta.caminho_boleto = bank_slip_url
        conta.id_cobranca = cobranca_id
        conta.tipo_pgto='BOLETO'


        conta.save()

        response_data = {
            'bankSlipUrl': bank_slip_url
        }

        # Retorna a resposta em formato JSON
    return JsonResponse(response_data)


def delete_boleto(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        id = data['id_conta']
        conta = Contas_Receber.objects.filter(id_cobranca=id).first()
        numero_pedido = conta.pedido
        pedido = Pedidos.objects.get(numero=numero_pedido)

        url= "https://asaas.com/api/v3/payments/"
        url = url+id
        print(url)
        if pedido.empresa_pedido.nome == 'TEM INDUSTRIA E COMERCIO DE TAPETES LTDA':
            headers = {
                "accept": "application/json",
                "access_token": "$aact_YTU5YTE0M2M2N2I4MTliNzk0YTI5N2U5MzdjNWZmNDQ6OjAwMDAwMDAwMDAwMDAzMjk5NjA6OiRhYWNoXzkxM2IzMTAyLTViMDQtNDFlYy1hMTg5LTJhMmY3ODMxZjJjMw=="
            }
        else:
            headers = {
                "accept": "application/json",
                "content-type": "application/json",
                "access_token": "$aact_YTU5YTE0M2M2N2I4MTliNzk0YTI5N2U5MzdjNWZmNDQ6OjAwMDAwMDAwMDAwMDAzMzI5Njk6OiRhYWNoX2ViZWQ5N2E3LTRhNTAtNGIxMC1iM2NhLTJjODk3NWJkOTVhNg=="
            }

        response = requests.delete(url, headers=headers)

        if response.status_code == 200:
            conta.id_cobranca=''
            conta.caminho_boleto=''
            conta.save()

            return JsonResponse({'message': 'Boleto excluído com sucesso.'}, status=200)
        else:
            return JsonResponse({'error': 'Erro ao excluir boleto.'}, status=400)

def expedicao_frete(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        numero_pedido = data.get('numero_pedido')
        peso = data.get('peso')
        volume = data.get('volume')

        pedido = Pedidos.objects.get(numero=numero_pedido)
        pedido.frete_peso = peso
        pedido.frete_volume = volume
        pedido.save()
        return JsonResponse({'mensagem': 'Pedido Atualizado'})



@csrf_exempt
def webhook_view(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        print (data)
        event_type = data.get('event')


        payment_id = data.get('payment', {}).get('id')
        # customer_id = data['payment']['customer']
        description = data.get('payment', {}).get('description')
        data_vencimento = data.get ('payment', {}).get('dueDate')
        conta = Contas_Receber.objects.filter(id_cobranca=payment_id).first()

        if conta:

            if event_type == 'PAYMENT_CREATED':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_AWAITING_RISK_ANALYSIS':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_APPROVED_BY_RISK_ANALYSIS':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_REPROVED_BY_RISK_ANALYSIS':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_UPDATED':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_CONFIRMED':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.status_conta = 'Pago'
                conta.data_pagamento = date.today()
                conta.save()

            elif event_type == 'PAYMENT_RECEIVED':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.data_pagamento = date.today()
                conta.status_conta = 'Pago'
                conta.save()

            elif event_type == 'PAYMENT_ANTICIPATED':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.data_pagamento = date.today()
                conta.status_conta = 'Pago'
                conta.save()

            elif event_type == 'PAYMENT_OVERDUE':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.status_conta = 'Vencida'
                conta.data_pagamento = None

                conta.save()

            elif event_type == 'PAYMENT_DELETED':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_RESTORED':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_REFUNDED':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_REFUND_IN_PROGRESS':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_RECEIVED_IN_CASH_UNDONE':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_CHARGEBACK_REQUESTED':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_CHARGEBACK_DISPUTE':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_AWAITING_CHARGEBACK_REVERSAL':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_DUNNING_RECEIVED':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_DUNNING_REQUESTED':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_BANK_SLIP_VIEWED':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            elif event_type == 'PAYMENT_CHECKOUT_VIEWED':
                conta.status_cobranca = data.get('payment', {}).get('status')
                conta.save()

            else:
                pass

            return HttpResponse("Webhook received successfully", status=200)
        else:

            if isinstance(description, int):
                conta = Contas_Receber.objects.filter(pedido=description, data_vencimento = data_vencimento).first()
                if conta:

                    if event_type == 'PAYMENT_CREATED':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_AWAITING_RISK_ANALYSIS':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_APPROVED_BY_RISK_ANALYSIS':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_REPROVED_BY_RISK_ANALYSIS':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_UPDATED':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_CONFIRMED':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.status_conta = 'Pago'
                        conta.data_pagamento = date.today()
                        conta.save()

                    elif event_type == 'PAYMENT_RECEIVED':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.data_pagamento = date.today()
                        conta.status_conta = 'Pago'
                        conta.save()

                    elif event_type == 'PAYMENT_ANTICIPATED':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.data_pagamento = date.today()
                        conta.status_conta = 'Pago'
                        conta.save()

                    elif event_type == 'PAYMENT_OVERDUE':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.status_conta = 'Vencida'
                        conta.data_pagamento = None

                        conta.save()

                    elif event_type == 'PAYMENT_DELETED':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_RESTORED':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_REFUNDED':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_REFUND_IN_PROGRESS':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_RECEIVED_IN_CASH_UNDONE':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_CHARGEBACK_REQUESTED':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_CHARGEBACK_DISPUTE':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_AWAITING_CHARGEBACK_REVERSAL':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_DUNNING_RECEIVED':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_DUNNING_REQUESTED':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_BANK_SLIP_VIEWED':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                    elif event_type == 'PAYMENT_CHECKOUT_VIEWED':
                        conta.status_cobranca = data.get('payment', {}).get('status')
                        conta.save()

                else:
                    pass

                return HttpResponse("Webhook received successfully", status=200)
            else:
                return HttpResponse("Conta Não Encontrada!", status=200)

    else:
        return HttpResponse("Method Not Allowed", status=405)




@login_required
def nova_guia_reparo(request, numero_pedido):
    usuario_nome = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario_nome).first()
    if request.method == 'POST':
        itens = request.POST.getlist("items")
        observacoes = request.POST.getlist("observacoes")
        checks = request.POST.getlist("checks")

        pedido = Pedidos.objects.filter(numero=numero_pedido).first()
        pedido.flag_reparo = 1
        nova_guia = Guia_Reparo()
        nova_guia.codigo = pedido.numero
        status = Status_Reparo.objects.filter(descricao='Expedicao Reparo').first()
        nova_guia.status = status
        nova_guia.pedido = pedido
        nova_guia.data_criacao = datetime.now()
        nova_guia.data_alteracao = datetime.now()
        nova_guia.save()
        nova_guia.observacao = 'Observação: '
        observacoes_verificadas=[]
        for observacao in observacoes:
            if len(observacao.strip()) == 0:
                pass
            else:
                observacoes_verificadas.append(observacao)
        for check, observacao in zip(checks, observacoes_verificadas ):

            if check and observacao:

                try:
                    novo_reparo = Reparos()
                    novo_reparo.guia = nova_guia
                    novo_reparo.pedido = pedido
                    novo_reparo.item = Itens_Pedido.objects.filter(pedido=pedido.numero, id=check).first()
                    novo_reparo.observacao = observacao
                    novo_reparo.save()
                    nova_guia.observacao += f"\n{observacao}"
                    nova_guia.data_alteracao = datetime.now()
                    nova_guia.save()

                except Exception as e:  # Use Exception e capture o erro para depuração, se necessário
                    print(f"Erro ao salvar o reparo: {e}")
                    nova_guia.delete()  # Deleta a guia apenas em caso de erro

                    return HttpResponse('Erro ao Salvar Reparo')

            else:
                nova_guia.delete()
                return HttpResponse('Erro ao salvar: Faltaram Informações')
        pedido.status = 'Expedicao Reparo'
        pedido.save()
        return HttpResponse('Nova Guia Gerada Com Sucesso!\n Acesse Módulo Guias de Reparo e Consulte pelo Numero Do Pedido')
    if 'Gerente' or 'Vendedor' in colaborador.funcao:

        pedido = Pedidos.objects.filter(numero=numero_pedido).first()
        itens = Itens_Pedido.objects.filter(pedido=pedido.numero)


        context = {

            'pedido': pedido,
            'itens': itens

        }
        return render (request, 'guiareparo/new_guia.html', context)

    else:
        HttpResponse('Usuario sem permissão')
@login_required
def delete_guia(request):

    if request.method == 'POST':
        id_guia = request.POST.get('id_guia')
        guia = Guia_Reparo.objects.filter(id=id_guia).first()
        if guia:
            reparos = Reparos.objects.filter(guia=guia)
            if reparos:
                reparos.delete()
            guia.delete()
            return JsonResponse({'mensagem': 'Guia deletada com sucesso!'})
        else:
           return JsonResponse({'mensagem': 'Falha ao deletar a Guia! Guia não Encontrada!'})






    # Se a requisição não for do tipo POST, você pode retornar um erro
    return JsonResponse({'mensagem': 'Método inválido.'}, status=400)
def solicita_reparo(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        print(data)
        numero_pedido = data.get('numero_pedido')
        observacao = data.get('observacao')
        nome_usuario = request.user.username
        colaborador = Colaborador.objects.filter(usuario__username=nome_usuario).first()
        status = 'Expedicao Reparo'
        print(colaborador)
        pedido = Pedidos.objects.filter(numero=numero_pedido).first()
        pedido_reparo = Reparos(
            pedido = pedido,
            observacao = observacao
        )
        pedido_reparo.save()
        pedido.status = status
        pedido.save()
        return JsonResponse({'status': 'success'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

@login_required
def painel_guias(request):

    colaborador = Colaborador.objects.filter(usuario__username=request.user.username).first()
    if 'Gerente' in colaborador.funcao or 'Vendedor' in colaborador.funcao:
        if colaborador.funcao == 'Gerente':
            guias = Guia_Reparo.objects.all()

        else:
            guias = Guia_Reparo.objects.filter(Q(pedido__Vendedor=colaborador.nome)|Q(pedido__Vendedor2=colaborador.nome))

        context = {

            'guias': guias
        }

        return render (request, 'guiareparo/painel_reparos.html', context)

    else:
        return HttpResponse('Sem Autorização')

def atualiza_status(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        print(data)
        numero_pedido = data.get('numero_pedido')
        nome_usuario = request.user.username
        colaborador = Colaborador.objects.filter(usuario__username=nome_usuario).first()
        status = ''
        print(colaborador)
        pedido = Pedidos.objects.filter(numero=numero_pedido).first()
        status_anterior = pedido.status
        print(pedido.status)
        if 'Gerente' or 'Financeiro' in colaborador.funcao:
            if data.get('status'):
                status = data.get('status')

                pedido.status = status
                pedido.save()
                if pedido and colaborador and status_anterior:
                    novo_log(pedido.numero,colaborador,f'{pedido.numero} Alteração de Status {status_anterior} para {pedido.status}')
                return JsonResponse({'status': 'success'})

        if pedido.status == 'Nfe Antecipada':

            pedido.status = 'Pendente Financeiro'
            pedido.save()
            if pedido and colaborador and status_anterior:
                novo_log(pedido.numero, colaborador,
                         f'{pedido.numero} Alteração de Status {status_anterior}  para {pedido.status}')


            return JsonResponse({'status': 'success'})

        elif pedido.status == 'Confirma os Valores':

            pedido.status = 'Producao'
            pedido.save()
            if pedido and colaborador and status_anterior:
                novo_log(pedido.numero, colaborador,
                         f'{pedido.numero} Alteração de Status {status_anterior}  para {pedido.status}')
            return JsonResponse({'status': 'success'})

        elif pedido.status == 'Pendente Financeiro':
            pedido.status = 'Producao'
            pedido.save()
            if pedido and colaborador and status_anterior:
                novo_log(pedido.numero, colaborador,
                         f'{pedido.numero} Alteração de Status {status_anterior}  para {pedido.status}')
            return JsonResponse({'status': 'success'})

        elif pedido.status == 'Pendente Logistica':
            pedido.status = 'Producao'
            pedido.save()
            return JsonResponse({'status': 'success'})

        elif 'Producao' in pedido.status:
            pedido.status = 'Expedicao'
            pedido.save()
            Producao.objects.filter(pedido=pedido).delete()
            guia = Guia_Reparo.objects.filter(codigo=pedido.numero).first()
            if guia:
                guia.status = Status_Reparo.objects.filter(descricao='Expedicao Reparo Concluido').first()
                guia.data_alteracao=datetime.now()
                guia.save()
            return JsonResponse({'status': 'success'})

        elif pedido.status == 'Expedicao':
            pedido.status = 'Financeiro'
            pedido.save()
            if pedido and colaborador and status_anterior:
                novo_log(pedido.numero, colaborador,
                         f'{pedido.numero} Alteração de Status {status_anterior}  para {pedido.status}')
            guia = Guia_Reparo.objects.filter(codigo=pedido.numero).first()
            if guia:
                pedido.status = 'Liberado para Entrega'
                pedido.save()
                guia.status = Status_Reparo.objects.filter(descricao='Reparo Rota de Entrega').first()
                guia.data_alteracao = datetime.now()
                guia.save()

            return JsonResponse({'status': 'success'})

        elif pedido.status == 'Expedicao Reparo':
            pedido.status = 'Em Coleta Reparo'
            pedido.save()
            guia = Guia_Reparo.objects.filter(codigo=pedido.numero).first()
            if pedido and colaborador and status_anterior:
                novo_log(pedido.numero, colaborador,
                         f'{pedido.numero} Alteração de Status {status_anterior}  para {pedido.status}')
            if guia:
                guia.status = Status_Reparo.objects.filter(descricao='Em Coleta Reparo').first()
                guia.data_alteracao = datetime.now()
                guia.save()
            return JsonResponse({'status': 'success'})

        elif pedido.status == 'Em Coleta Reparo':
            pedido.status = 'Producao Reparo'
            pedido.save()
            guia = Guia_Reparo.objects.filter(codigo=pedido.numero).first()
            if guia:
                guia.status = Status_Reparo.objects.filter(descricao='Producao Reparo').first()
                guia.data_alteracao = datetime.now()
                guia.save()
            return JsonResponse({'status': 'success'})

        elif pedido.status == 'Producao Reparo':
            pedido.status = 'Expedicao Reparo Concluido'
            pedido.save()
            if pedido and colaborador and status_anterior:
                novo_log(pedido.numero, colaborador,
                         f'{pedido.numero} Alteração de Status {status_anterior}  para {pedido.status}')
            guia = Guia_Reparo.objects.filter(codigo=pedido.numero).first()
            if guia:
                guia.status = Status_Reparo.objects.filter(descricao='Expedicao Reparo Concluido').first()
                guia.data_alteracao = datetime.now()
                guia.save()
            return JsonResponse({'status': 'success'})

        elif   pedido.status == 'Expedicao Reparo Concluido':
               pedido.status = 'Em Rota de Entrega'
               pedido.save()
               if pedido and colaborador and status_anterior:
                   novo_log(pedido.numero, colaborador,
                            f'{pedido.numero} Alteração de Status {status_anterior}  para {pedido.status}')
               guia = Guia_Reparo.objects.filter(codigo=pedido.numero).first()
               if guia:
                   guia.status = Status_Reparo.objects.filter(descricao='Reparo Rota de Entrega').first()
                   guia.data_alteracao = datetime.now()
                   guia.save()
               return JsonResponse({'status': 'success'})

        elif pedido.status == 'Financeiro':
            pedido.status = 'Liberado para Entrega'
            pedido.save()
            if pedido and colaborador and status_anterior:
                novo_log(pedido.numero, colaborador,
                         f'{pedido.numero} Alteração de Status {status_anterior}  para {pedido.status}')
            return JsonResponse({'status': 'success'})

        elif pedido.status == 'Liberado para Entrega':
            if pedido.tipo_entrega == 'retirada':
                pedido.status = 'Liberado Retirada Cliente'
            elif pedido.tipo_entrega == 'transportadora':
                pedido.status = 'Liberado Coleta'
            else:
                pedido.status = 'Em Rota de Entrega'

            if pedido and colaborador and status_anterior:
                novo_log(pedido.numero, colaborador,
                         f'{pedido.numero} Alteração de Status {status_anterior}  para {pedido.status}')
            pedido.save()
            return JsonResponse({'status': 'success'})
        elif pedido.status == 'Liberado Coleta':
            pedido.status = 'Em Rota Transportadora'
            pedido.save()
            return JsonResponse({'status': 'success'})
        elif pedido.status == 'Em Rota de Entrega' or pedido.status == 'Retirada Cliente' or pedido.status == 'Em Rota Transportadora':
            pedido.status = 'Entregue'
            pedido.save()
            if pedido and colaborador and status_anterior:
                novo_log(pedido.numero, colaborador,
                         f'{pedido.numero} Alteração de Status {status_anterior}  para {pedido.status}')
            guia = Guia_Reparo.objects.filter(codigo=pedido.numero).first()
            if guia:
                guia.status = Status_Reparo.objects.filter(descricao='Reparo Entregue').first()
                guia.data_alteracao = datetime.now()
                guia.save()
            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

def edita_status(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        numero_pedido = data.get('numero_pedido')
        pedido = Pedidos.objects.get(numero=numero_pedido)
        banco_nome = data.get('banco')
        empresa_nome = data.get('empresa')
        banco = Bancos.objects.filter(nome=banco_nome, empresa__nome=empresa_nome).first()
        if banco is None:
            return JsonResponse({'status': 'error', 'message': 'Banco Não Encontrado Com a Respectiva Empresa!'}, status=405)
        else:
            if pedido.status == 'Pendente Financeiro' or pedido.status == 'PENDENTE FINANCEIRO':
                valor = float(data.get('valor'))
                data = data.get('data')
                contas = Contas_Receber.objects.filter(pedido=numero_pedido,cliente=pedido.cliente_pedido,momento_pagamento='ANTECIPADO')
                for conta in contas:
                    totalconta = conta.total_parcelas*conta.valor
                    if totalconta == valor:
                        conta.data_pagamento = data
                        conta.status_conta = 'Pago'
                        conta.banco=banco
                        conta.save()
                    else:
                        return JsonResponse({'status': 'error', 'message': 'Conta Não Confere com o Contas a Receber!'}, status=405)

                itens = Itens_Pedido.objects.filter(pedido=numero_pedido)
                hoje = date.today()

                # Itera sobre os itens do pedido
                if (pedido.tipo_pedido.descricao == 'PADRAO'):
                    for item in itens:

                        produto = item.produto

                        # Calcula a data de entrega com base na data de hoje mais o prazo do produto
                        data_entrega = hoje + timedelta(days=produto.prazo_fabricacao)
                        print(data_entrega)
                        # Verifica se a data de entrega é menor que o prazo de entrega
                        if data_entrega > pedido.prazo_entrega:

                            pedido.status = 'Pendente Logistica'
                            pedido.save()

                        else:
                            pedido.status = 'Producao'
                            pedido.save()

                    return JsonResponse({'status': 'success'})

                else:
                    pedido.status = 'Producao'
                    pedido.save()

                    return JsonResponse({'status': 'success'})

            else:
                return JsonResponse({'status': 'error', 'message': 'Pedido Com Status {{pedido.status}}'}, status=405)

    else:
          return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

def pedido_create_aberto(request):
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    colaboradores_ativos = Colaborador.objects.filter(status__descricao='ATIVO')
    vendedores = colaboradores_ativos.filter(Q(funcao='Vendedor Interno') | Q(funcao='Vendedor Externo')|Q(funcao='Gerente'))
    clientes = Cliente.objects.all()
    produtos = Produtos.objects.all()
    produtos_adicionais = Produtos.objects.filter(produto_adicional = 'SIM')
    empresas = Empresas.objects.all()
    bairros = Bairro.objects.all()
    tipos_pagamentos = TipoPgto.objects.all()
    fornecedores = Fornecedor.objects.filter(tipo_fornecedor='1')
    bonificados = Fornecedor.objects.filter(tipo_fornecedor='8')
    tipos_pedidos = Tipo_Pedido.objects.all()
    produtos_revenda = Produtos_Revenda.objects.all()
    lista_produtos = []


    for produto in produtos:
        lista_produtos.append(produto)
    for produto in produtos_revenda:
        lista_produtos.append(produto)

    revendas = Revenda.objects.all()
    zona_entrega = 'Sem Zona'
    momentos = [
        'Antecipado',
        'Entrega',
        'Faturado'
    ]
    cep_entrega = ''
    uf_entrega = ''
    cidade_entrega = ''
    bairro_entrega = ''
    endereco_entrega = ''
    numero_end_entrega = '0'
    complemento = ''
    total_avista ='0'
    total_parcelado = '0'
    desconto ='0'

    context = {
        'vendedor': colaborador.nome,
        'vendedores': vendedores,
        'clientes': clientes,
        'produtos': lista_produtos,
        'produtos_adicionais': produtos_adicionais,
        'bairros': bairros,
        'fornecedores': fornecedores,
        'empresas': empresas,
        'tipos_pagamentos':tipos_pagamentos,
        'momentos': momentos,
        'colaborador':colaborador,
        'revendas':revendas,
        'produtos_revenda':produtos_revenda,
        'tipos_pedidos':tipos_pedidos,
        'lista_produtos': lista_produtos,
        'bonificados': bonificados

    }

    if request.method == 'POST':

        pagamento_antecipado = 0.00

        numero_pedido = request.POST['numero_pedido']
        print(numero_pedido)
        flag_pedido = request.POST.get('flag_pedido')
        if flag_pedido is None:
            flag_pedido = 0
        datapg_antecipado = None
        pedido_temp = Temp_Pedidos.objects.filter(pedido=numero_pedido).first()
        datapedido = request.POST.get('data-pedido')
        prazo_entrega = request.POST.get('data_entrega')

        if 'Vendedor or Gerente' in colaborador.funcao:
            vendedor = colaborador.nome
        else:

             vendedor = request.POST.get('vendedor1')

        vendedor2 = request.POST.get('vendedor2')

        total_pedido_parcial = 0.00

        ponto_referencia_entrega = request.POST.get('pontoreferencia-opcional')
        documento = request.POST.get('documento')
        documento_cliente = ''.join(filter(str.isdigit, documento))
        cliente_pedido = Cliente.objects.filter(documento=documento_cliente).first()


        empresa = request.POST.get('empresa_pedido')
        empresa = Empresas.objects.get(nome=empresa) if empresa else colaborador.empresa


        tipo_pedido = request.POST.get('tipo_pedido')
        tipo_do_pedido = Tipo_Pedido.objects.filter(descricao=tipo_pedido).first()

        itens = Itens_Pedido.objects.filter(pedido=numero_pedido)
        for item in itens:
            total_pedido_parcial += item.total_item


        tipo_frete = request.POST.get('opcao_frete')
        nfe_opcao = request.POST.get('pedido-nfe-opcao')
        retirada_flag = 0
        if nfe_opcao == 'sim':
            nfe_flag = 1
        else:
            nfe_flag = 0


        if tipo_frete == 'Transportadora':
            frete_flag = 1
        else:
            frete_flag = 0
            if tipo_frete == 'retirada':
                retirada_flag = 1
            else:
                retirada_flag = 0



        endereco_opcao = request.POST.get('endereco-opcao')
        frete_valor = request.POST.get('frete')
        frete_valor = re.sub(r'^R\$|\.', '', frete_valor).replace(',', '.')
        frete_valor = round(float(frete_valor), 2) if frete_valor else 0
        status_pedido = 'Aberto'

        numero_pedido_str = str(numero_pedido)
        last_5_digits = numero_pedido_str[-5:]


        informacoes_complementares = request.POST.get('informacoes_complementares')

        if cliente_pedido:
            cliente_id = cliente_pedido
        else:
            cliente_id = None  # Trate o caso em que o cliente não é encontrado



        if endereco_opcao == 'cadastro':
            if cliente_pedido:
                cep_entrega = cliente_pedido.CEP
                uf_entrega = cliente_pedido.estado
                cidade_entrega = cliente_pedido.cidade
                bairro_entrega = cliente_pedido.bairro
                endereco_entrega = cliente_pedido.endereco
                numero_end_entrega = cliente_pedido.numero_endereco
                complemento = cliente_pedido.complemento
                ponto_referencia_entrega = request.POST.get ('ponto-referencia-cliente')
                if (cidade_entrega == 'Rio de Janeiro' or cidade_entrega == 'RIO DE JANEIRO'):
                    localidade = Bairro.objects.filter(nome=bairro_entrega, cidade=cidade_entrega).first()
                    if localidade:
                        zona_entrega = localidade.zona if localidade.zona else None


        elif endereco_opcao == 'opcional':
            cep_entrega = request.POST.get('cep-entrega-opcional')
            bairro_entrega = request.POST.get('bairro-entrega-opcional')
            cidade_entrega = request.POST.get('cidade-entrega-opcional')
            uf_entrega = request.POST.get('uf-entrega-opcional')
            endereco_entrega = request.POST.get('endereco-entrega-opcional')
            numero_end_entrega = request.POST.get('numero-endereco-opcional')
            complemento = request.POST.get('complemento-endereco-opcional')
            ponto_referencia_entrega = request.POST.get('pontoreferencia-opcional')
            if (cidade_entrega == 'Rio de Janeiro'):
                localidade = Bairro.objects.get(nome=bairro_entrega, cidade=cidade_entrega)
                zona_entrega = localidade.zona if localidade.zona else None

        print (total_pedido_parcial)
        desconto = 0.00
        repasse = request.POST.get('acerto')
        repasse=re.sub(r'^R\$|\.', '', repasse).replace(',', '.') if repasse else None
        repasse = float(repasse)
        desconto = float(desconto) if desconto else 0
        if desconto > 0:
            total_pedido =round(float(total_pedido_parcial - desconto), 2)
        else:
            total_pedido = float(total_pedido_parcial)
            total_pedido = round(total_pedido,2)
        print(total_pedido)
        bonificado_pedido = ''
        if repasse > 0:
            repasse = float(repasse)
            conta_pagar = Contas_Pagar()
            total_pagar = total_pedido - total_pedido_parcial
            conta_pagar.status = 'A Vencer'
            conta_pagar.pedido = numero_pedido
            bonificado = request.POST.get('bonificado')
            conta_pagar.fornecedor = Fornecedor.objects.filter(nome=bonificado).first()
            bonificado_pedido = conta_pagar.fornecedor
            prazo_entrega_datetime = datetime.strptime(prazo_entrega, '%Y-%m-%d')
            data_vencimento = prazo_entrega_datetime + timedelta(days=30)
            conta_pagar.data_vencimento = data_vencimento
            conta_pagar.numero_parcela=1
            conta_pagar.total_parcelas =1
            conta_pagar.valor = repasse
            conta_pagar.descricao = "BONIFICACAO"
            conta_pagar.save()

        itens = Itens_Pedido.objects.filter(pedido=numero_pedido)
        menor_comissao = 100
        vendedor_comissao = 0
        vendedor = Colaborador.objects.get(nome=vendedor)


        vendedor_comissao = request.POST.get('comissao1')
        vendedor2_comissao = 0
        if vendedor2:
            vendedor2 = Colaborador.objects.get(nome=vendedor2)
            if vendedor_comissao =='1':
                vendedor2_comissao = 1
            elif vendedor_comissao =='3':
                vendedor2_comissao = 3
            elif vendedor_comissao == '5':
                vendedor2_comissao = 7
            elif vendedor_comissao == '6':
                vendedor2_comissao = 10
            elif vendedor_comissao == '7':
                vendedor2_comissao = 15




        informacoes_parcelas = ""
        momento_pagamento_1 = request.POST.get('momento-pagamento1')
        momento_pagamento_2 = request.POST.get('momento-pagamento2') if request.POST.get('momento-pagamento2') else None
        total_pagamento1 = request.POST.get('total_1')
        total_pagamento1 = re.sub(r'^R\$|\.', '', total_pagamento1).replace(',', '.') if total_pagamento1 else 0.00
        forma_pagamento1 = request.POST.get('forma_pagamento1')
        parcelas_cartao_1 = request.POST.get('parcelas_cartao_1')
        total_parcelado_cartao_1 = request.POST.get('valor_parcelado_cartao_1')
        total_parcelado_cartao_1 = re.sub(r'^R\$|\.', '', total_parcelado_cartao_1).replace(',', '.') if total_parcelado_cartao_1 else 0.00
        faturamento_valor_parcelas = request.POST.get('valor_parcelas')
        faturamento_valor_parcelas = re.sub(r'^R\$|\.', '', faturamento_valor_parcelas).replace(',', '.') if faturamento_valor_parcelas else 0.00
        faturamento_quantidade_parcelas = request.POST.get('total_parcelas')
        total_pagamento2 = request.POST.get('total_2')
        total_pagamento2 = re.sub(r'^R\$|\.', '', total_pagamento2).replace(',', '.') if total_pagamento2 else 0.00
        forma_pagamento2 = request.POST.get('forma_pagamento2')
        parcelas_cartao_2 = request.POST.get('parcelas_cartao_2')
        total_parcelado_cartao_2 = request.POST.get('valor_parcelado_cartao_2')
        total_parcelado_cartao_2 = re.sub(r'^R\$|\.', '', total_parcelado_cartao_2).replace(',', '.') if total_parcelado_cartao_2 else 0.00
        faturamento_quantidade_parcelas_2 = request.POST.get('total_parcelas_2')
        faturamento_valor_parcelas_2 = request.POST.get('valor_parcelas_2')
        faturamento_valor_parcelas_2 = re.sub(r'^R\$|\.', '', faturamento_valor_parcelas_2).replace(',', '.') if faturamento_quantidade_parcelas_2 else 0.00
        bancopg_antecipado=""
        pagamento_entrega = 0.00
        Contas_Receber.objects.filter (pedido=numero_pedido).delete()

        if momento_pagamento_1 == 'ANTECIPADO':
            data_vencimento = datetime.today() + timedelta(days=1)
            status_pedido = 'Pendente Financeiro'
            datapg_antecipado = request.POST.get('data_antecipado')
            bancopg_antecipado = request.POST.get('banco_pa')

            if forma_pagamento1 == 'CARTAO CREDITO':
                pagamento_antecipado =float(total_parcelado_cartao_1)
                for parcela in range(1, int(parcelas_cartao_1) + 1):
                    data_vencimento = (data_vencimento + timedelta(days=30))
                    valor_parcela = float(total_parcelado_cartao_1) / int(parcelas_cartao_1)


                    contas_receber = Contas_Receber(
                        pedido=numero_pedido,  # Substitua pelo valor correto
                        cliente=cliente_pedido,
                        descricao='Parcela {} - Crédito'.format(parcela),
                        data_vencimento=data_vencimento,  # Substitua pelo valor correto
                        data_pagamento=None,
                        tipo_pgto=forma_pagamento1,
                        numero_parcela=parcela,
                        total_parcelas=int(parcelas_cartao_1),
                        valor=valor_parcela,
                        status_conta='A Vencer',
                        momento_pagamento='ANTECIPADO'
                    )
                    contas_receber.save()

            else:
                pagamento_antecipado =float(total_pagamento1)
                numero_parcela =1
                total_parcelas = 1
                descricao = f"{cliente_pedido.nome} {numero_parcela} / {total_parcelas}"
                contas_receber = Contas_Receber(
                    pedido=numero_pedido,
                    cliente=cliente_pedido,
                    descricao=descricao,
                    data_vencimento=data_vencimento,
                    data_pagamento=None,
                    tipo_pgto=forma_pagamento1,
                    numero_parcela=1,
                    total_parcelas=1,
                    valor=total_pagamento1,
                    status_conta='A Vencer',
                    momento_pagamento='ANTECIPADO'
                )
                contas_receber.save()

        if momento_pagamento_1 == 'ENTREGA':
            data_vencimento = prazo_entrega
            status_pedido = 'Producao'
            data_vencimento = datetime.today() + timedelta(days=1)
            pagamento_entrega = 0.00
            if forma_pagamento1 == 'CARTAO CREDITO':
                pagamento_entrega = round(float(total_parcelado_cartao_1), 2)
                for parcela in range(1, int(parcelas_cartao_1) + 1):
                    data_vencimento = (data_vencimento + timedelta(days=30))
                    print(data_vencimento)
                    valor_parcela = float(total_parcelado_cartao_1) / int(parcelas_cartao_1)
                    contas_receber = Contas_Receber(
                        pedido=numero_pedido,  # Substitua pelo valor correto
                        cliente = cliente_pedido,
                        descricao='Parcela {} - Crédito'.format(parcela),
                        data_vencimento=data_vencimento,  # Substitua pelo valor correto
                        data_pagamento=None,
                        tipo_pgto=forma_pagamento1,
                        numero_parcela=parcela,
                        total_parcelas=int(parcelas_cartao_1),
                        valor=valor_parcela,
                        status_conta='Aberto',
                        momento_pagamento='ENTREGA'
                    )
                    contas_receber.save()
            else:
                numero_parcela = 1
                total_parcelas = 1
                descricao = f"{cliente_pedido.nome} {numero_parcela} / {total_parcelas}"
                pagamento_entrega = round(float(total_pagamento1), 2)
                data_vencimento = (data_vencimento + timedelta(days=15))
                contas_receber = Contas_Receber(
                    pedido=numero_pedido,  # Substitua pelo valor correto
                    cliente=cliente_pedido,
                    descricao=descricao,
                    data_vencimento=data_vencimento,
                    data_pagamento=None,
                    tipo_pgto=forma_pagamento1,
                    numero_parcela=1,
                    total_parcelas=1,
                    valor=total_pagamento1,
                    status_conta='Aberto',
                    momento_pagamento='ENTREGA'
                )
                contas_receber.save()

        if momento_pagamento_1 == 'FATURADO':
            status_pedido = 'Producao'

            valor_parcela = float(faturamento_valor_parcelas)
            print(valor_parcela)
            informacoes_parcelas = request.POST.get('informacoes_parcelas')
            for parcela in range(1, int(faturamento_quantidade_parcelas) + 1):



                contas_receber = Contas_Receber(
                    pedido=numero_pedido,
                    cliente = cliente_pedido,
                    descricao='{} - Parcela {} - Faturado'.format(cliente_pedido, parcela),
                    data_pagamento=None,
                    tipo_pgto=forma_pagamento1,
                    numero_parcela=parcela,
                    total_parcelas=int(faturamento_quantidade_parcelas),
                    valor=valor_parcela,
                    status_conta='Pendente Faturamento',
                    momento_pagamento='FATURADO'
                )

                contas_receber.save()


        if momento_pagamento_2 == 'ANTECIPADO':


            data_vencimento = datetime.today() + timedelta(days=1)
            datapg_antecipado = request.POST.get('data_antecipado_2')
            status_pedido = 'Pendente Financeiro'
            if forma_pagamento2 == 'CARTAO CREDITO':
                pagamento_antecipado = float(total_parcelado_cartao_2)
                for parcela in range(1, int(parcelas_cartao_2) + 1):
                    data_vencimento = (data_vencimento + timedelta(days=30))
                    valor_parcela = float(total_parcelado_cartao_2) / int(parcelas_cartao_2)
                    contas_receber = Contas_Receber(
                        pedido=numero_pedido,
                        cliente = cliente_pedido, # Substitua pelo valor correto
                        descricao='Parcela {} - Crédito'.format(parcela),
                        data_vencimento=data_vencimento,
                        data_pagamento=None,
                        tipo_pgto=forma_pagamento2,
                        numero_parcela=parcela,
                        total_parcelas=int(parcelas_cartao_2),
                        valor=valor_parcela,
                        status_conta='A Vencer',
                        momento_pagamento='ANTECIPADO'
                    )
                    contas_receber.save()

            else:
                numero_parcela = 1
                total_parcelas = 1
                descricao = f"{cliente_pedido.nome} {numero_parcela} / {total_parcelas}"
                pagamento_antecipado = float(total_pagamento2)
                contas_receber = Contas_Receber(
                    pedido=numero_pedido,
                    cliente = cliente_pedido,# Substitua pelo valor correto
                    descricao=descricao,
                    data_vencimento=data_vencimento,
                    data_pagamento=None,
                    tipo_pgto=forma_pagamento2,
                    numero_parcela=1,
                    total_parcelas=1,
                    valor=total_pagamento2,
                    status_conta='A Vencer',
                    momento_pagamento='ANTECIPADO'
                )
                contas_receber.save()



        if momento_pagamento_2 == 'ENTREGA':

            data_vencimento = prazo_entrega

            pagamento_entrega = round(float(total_parcelado_cartao_2),2)
            if forma_pagamento2 == 'CARTAO CREDITO':
                data_vencimento = datetime.today() + timedelta(days=15)
                for parcela in range(1, int(parcelas_cartao_2) + 1):
                    data_vencimento = (data_vencimento + timedelta(days=30))
                    valor_parcela = float(total_parcelado_cartao_2) / int(parcelas_cartao_2)
                    contas_receber = Contas_Receber(
                        pedido=numero_pedido,
                        cliente = cliente_pedido,# Substitua pelo valor correto
                        descricao='Parcela {} - Crédito'.format(parcela),
                        data_vencimento=data_vencimento,  # Substitua pelo valor correto
                        data_pagamento=None,
                        tipo_pgto='CRÉDITO',
                        numero_parcela=parcela,
                        total_parcelas=int(parcelas_cartao_2),
                        valor=valor_parcela,
                        status_conta='Aberto',
                        momento_pagamento='ENTREGA'
                    )
                    contas_receber.save()
            else:
                numero_parcela = 1
                total_parcelas = 1
                descricao = f"{cliente_pedido.nome} {numero_parcela} / {total_parcelas}"
                contas_receber = Contas_Receber(
                    pedido=numero_pedido,
                    cliente = cliente_pedido,# Substitua pelo valor correto
                    descricao=descricao,
                    data_vencimento=prazo_entrega,
                    data_pagamento=None,
                    tipo_pgto=forma_pagamento2,
                    numero_parcela=1,
                    total_parcelas=1,
                    valor=total_pagamento2,
                    status_conta='Aberto',
                    momento_pagamento='ENTREGA'
                )
                contas_receber.save()

        if momento_pagamento_2 == 'FATURADO':


            valor_parcela = float(faturamento_valor_parcelas_2)

            for parcela in range(1, int(faturamento_quantidade_parcelas_2) + 1):


                contas_receber = Contas_Receber(
                    pedido=numero_pedido,
                    cliente = cliente_pedido,# Substitua pelo valor correto
                    descricao='{} - Parcela {} - Faturado'.format(cliente_pedido, parcela),
                    data_pagamento=None,
                    tipo_pgto=forma_pagamento2,
                    numero_parcela=parcela,
                    total_parcelas=int(faturamento_quantidade_parcelas_2),
                    valor=valor_parcela,
                    status_conta='Pendente Faturamento',
                    momento_pagamento='FATURADO'
                )

                contas_receber.save()

        # Obtém a data de hoje
        hoje = date.today()
        #CHECA PENDENCIA LOGISTICA
        itens = Itens_Pedido.objects.filter(pedido=numero_pedido)
        if prazo_entrega:
            prazo_entrega = datetime.strptime(prazo_entrega, '%Y-%m-%d').date()
        else:
            prazo_entrega = hoje + timedelta(days=15)



        # Itera sobre os itens do pedido
        if (tipo_pedido == 'PADRAO'):
            for item in itens:

                produto = item.produto

                # Calcula a data de entrega com base na data de hoje mais o prazo do produto
                data_entrega = hoje + timedelta(days=produto.prazo_fabricacao)

                # Verifica se a data de entrega é menor que o prazo de entrega
                if data_entrega > prazo_entrega:
                    if status_pedido == 'Pendente Financeiro':
                        pass
                    else:
                        status_pedido = 'Pendente Logistica'
        else:
            for item in itens:

                produto = item.produto_revenda

                # Calcula a data de entrega com base na data de hoje mais o prazo do produto
                data_entrega = hoje + timedelta(days=produto.prazo_fabricacao)

                # Verifica se a data de entrega é menor que o prazo de entrega
                if data_entrega > prazo_entrega:
                    if status_pedido == 'Pendente Financeiro':
                        pass
                    else:
                        status_pedido = 'Pendente Logistica'



        #CADASTRO DE ARQUIVO LAYOUT
        nome_arquivo = None
        arquivo = request.FILES.get('arquivo')
        if arquivo:
            empresaaws = Empresas.objects.filter(id=1).first()
            s3_client = boto3.client('s3', aws_access_key_id=empresaaws.aws_id,
                                     aws_secret_access_key=empresaaws.aws_chave)
            agora = datetime.now()
            nome_arquivo = f"{numero_pedido}{agora.strftime('%Y%m%d%H%M%S')}.pdf"

            s3_client.upload_fileobj(arquivo, 'storagesw', nome_arquivo)

        nfe_antecipado = request.POST.get('antecipado-nfe-opcao')


        if nfe_antecipado == 'sim':

            status_pedido = 'Nfe Antecipada'
        else:
            pass

        fundo = request.POST.get('fundo')
        borda = request.POST.get('borda')
        letras = request.POST.get('letras')
        logomarca = request.POST.get('logomarca')

        vendedor2_comissaov = 0.00
        vendedor_comissaov = 0.00
        vendedor_comissaov = total_pedido_parcial / 100 * float(vendedor_comissao)
        if vendedor2:
            vendedor2_comissaov = total_pedido_parcial / 100 * float(vendedor2_comissao)
            vendedor2_comissaov = round(vendedor2_comissaov, 2)
        else:
            pass
        vendedor_comissaov = round(vendedor_comissaov,2)

        comprador_nome = request.POST.get('comprador_nome')
        comprador_telefone = request.POST.get('comprador_telefone')
        comprador_telefone = ''.join(filter(str.isdigit, comprador_telefone))
        comprador_email = request.POST.get('comprador_email')
        observacao = request.POST.get ('observacao')



        pedido_obj = Pedidos(
            numero=numero_pedido,
            datapedido=datapedido,
            valorTotal=total_pedido,
            valorTotalParcial=total_pedido_parcial,
            descontototal=desconto,
            cliente_pedido=cliente_pedido,
            prazo_entrega=prazo_entrega,
            pagamento_antecipado=pagamento_antecipado,
            datapg_antecipado= datapg_antecipado if datapg_antecipado else None,
            bancopg_antecipado=bancopg_antecipado if bancopg_antecipado else None,
            Vendedor=vendedor,
            Vendedor2=vendedor2,
            vendedor_comissao=vendedor_comissao,
            vendedor_comissaov = vendedor_comissaov,
            vendedor2_comissaov = vendedor2_comissaov,
            vendedor2_comissao=vendedor2_comissao,
            cep_entrega=cep_entrega,
            uf_entrega=uf_entrega,
            cidade_entrega=cidade_entrega,
            bairro_entrega=bairro_entrega,
            endereco_entrega=endereco_entrega,
            numero_end_entrega=numero_end_entrega,
            complemento=complemento,
            ponto_referencia_entrega = ponto_referencia_entrega,
            empresa_pedido=empresa,
            zona_entrega=zona_entrega if zona_entrega else "Sem Zona",
            pagamento_entrega = pagamento_entrega if pagamento_entrega else 0,
            status='Confirma os Valores',
            caminho_layout = f"https://storagesw.s3.amazonaws.com/{nome_arquivo}" if nome_arquivo else None,
            informacoes_adicionais = informacoes_complementares,
            operacao_nota = 1,
            frete_flag = frete_flag,
            flag_nfe = nfe_flag,
            tipo_pedido = tipo_do_pedido,
            informacoes_pagamento=informacoes_parcelas if informacoes_parcelas else None,
            retirada_flag = retirada_flag,
            valor_repasse = repasse,
            logomarca=logomarca,
            letra=letras,
            fundo=fundo,
            borda= borda,
            tipo_entrega = tipo_frete,
            flag_aberto = flag_pedido,
            bonificado = bonificado_pedido if bonificado_pedido else None,
            flag_impressao=0,
            comprador_nome=comprador_nome,
            comprador_email=comprador_email,
            comprador_telefone=comprador_telefone,
            observacao = observacao

        )

        pedido_obj.save()

        if tipo_do_pedido.descricao == 'PADRAO':
            compara_pedidos(numero_pedido)

        conta_receber = Contas_Receber.objects.filter(pedido=numero_pedido)
        for conta in conta_receber:
            pedido = Pedidos.objects.filter(numero=numero_pedido).first()
            conta.pedido_conta = pedido
            conta.save()
        if int(empresa.ultimo_pedido_ref) >= int(numero_pedido):
            pass
        else:
            empresa.ultimo_pedido_ref = numero_pedido
            empresa.save()

        if frete_valor > 0:

            conta_obj = Contas_Pagar(
                pedido=numero_pedido,
                descricao="FRETE REF PEDIDO",
                valor=frete_valor,

            )
            conta_obj.save()
            atualiza_pedido = Pedidos.objects.get(numero=numero_pedido)
            atualiza_pedido.frete_valor = frete_valor
            atualiza_pedido.save()
        else:
            pass

        redirect_url = f'/geradocpedido/{numero_pedido}/'

        return redirect(redirect_url)




    return render(request, 'novopedido_super_aberto.html', context)

def save_aberto_itens(request):


    if request.method == 'POST':
        print('entrou')
        items = json.loads(request.POST.get('items'))

        tipo_pedido = request.POST.get('tipopedido')
        print(tipo_pedido)
        frete = request.POST.get('frete')
        quantidade_itens = 0
        quantidade_itens = len(items)
        print('total de itens:', quantidade_itens)
        frete_item = 0.00
        if frete:
            frete = float(frete)
            frete_item = frete / quantidade_itens
            frete_item = round (frete_item,2)

        print ('frete_item:', frete_item)
        pedido = request.POST.get('pedido')

        lista_itens = Itens_Pedido.objects.filter(pedido=pedido).filter()
        if lista_itens:
            return JsonResponse({'error': 'Já existe um pedido com o mesmo número'})




        try:
            for item_data in items:


                item = Itens_Pedido()
                item.pedido = pedido
                item.observacao_item = item_data['observacao']

                item.quantidade = (item_data['quantidade'])
                item.quantidade=int(item.quantidade)

                item.nome = item_data['produto']
                print('item_nome:',item.nome)
                item.tipo = item_data['tipo_produto']

                item.comprimento =item_data['comprimento']
                item.comprimento= float(item.comprimento)
                item.comprimento = round(item.comprimento,3)
                adicional = 0.00
                adicional = item_data['adicional'] if item_data['adicional'] else 0.00

                if tipo_pedido == 'PADRAO':
                    print ('PEDIDO PADRAO')
                    produto = Produtos.objects.filter(descricao=item.nome).first()
                    if produto is None:
                        return JsonResponse({'error': f'Produto Não Encontrado: {item.nome}'})
                    item.produto_id = produto.id
                else:
                    produto = Produtos_Revenda.objects.filter(descricao=item.nome).first()
                    if produto is None:
                        return JsonResponse({'error': f'Produto Não Encontrado: {item.nome}'})
                    item.produto_revenda_id = produto.id


                if 'MPC' in produto.descricao:
                    item.comprimento = produto.tamanho_padrao
                    item.largura = item_data['largura']
                    item.largura = float(item.largura)
                    item.largura = round(item.largura, 3)
                elif 'MPL' in produto.descricao:
                    item.largura = produto.tamanho_padrao
                    item.comprimento = item_data['comprimento']
                    item.comprimento = float(item.comprimento)
                    item.comprimento = round(item.comprimento, 3)
                else:
                    item.comprimento = item_data['comprimento']
                    item.largura = item_data['largura']

                if (produto.tipo_medida == 'linear'):
                    item.largura = 1
                    item.medida_final = item.largura * item.comprimento
                elif (produto.tipo_medida == 'm2') :

                    item.medida_final = item.largura * item.comprimento
                elif (produto.tipo_medida == 'pc'):
                    item.medida_final = item.quantidade

                item.medida_final = round((item.medida_final),3)
                print(item.nome, item.medida_final)

                item.preco = item_data['preco']
                item.preco = round(float(item.preco),2)
                print(item.preco)

                item.valor_base = (item.preco + adicional) / item.medida_final
                item.valor_base = round((item.valor_base), 2)

                item.preco = (frete_item/item.quantidade) + adicional + item.preco
                item.preco = round(item.preco,2)
                item.total_item = (item.quantidade * item.preco)
                item.total_item = round((item.total_item), 2)
                print(item.comprimento, item.largura, item.nome)
                item.largura = "{:.3f}".format(item.largura)
                item.comprimento = "{:.3f}".format(item.comprimento)

                item.save()


                quantidade_itens = quantidade_itens - 1

                if frete > 0:
                    frete = frete - frete_item
                    if quantidade_itens > 0:
                        frete_item = frete / quantidade_itens
                        frete_item = round(frete_item, 2)


            return JsonResponse({'message': 'Dados recebidos com sucesso'})
        except Exception as e:
            return JsonResponse({'error': 'Erro ao processar os itens', 'message': str(e)}, status=400)


def nota_aberta (request):
    data = json.loads(request.body)
    items = data['itens']
    print(items)
    valorTotal = 0

    # Itere pelos itens e some os valores da chave 'total_item'
    for item in items:
        valorTotal += float(item['total_item'])

    pedido_numero = data.get('numero_pedido')


    pedido = Pedidos.objects.filter(numero=pedido_numero).first()

    informacoes_adicionais = data.get('informacoes_complementares')


    if pedido.cliente_pedido.nome_contato:
        if pedido.cliente_pedido.inscricao_estadual is None:
            return JsonResponse({"error": "Falta cadastro da inscrição estadual do CNPJ"})
        else:
            pass
    ref = {
        "ref": (str(pedido.empresa_pedido.cnpj) + str(pedido.numero)).zfill(44)
    }
    url = "https://api.focusnfe.com.br/"
    if (pedido.empresa_pedido.nome == 'TEM INDUSTRIA E COMERCIO DE TAPETES LTDA'):
        token = "OK0V8A4AKMOi5ywOFOUljMR8Ew4S3Imf"
    else: token = "Dzrt7d64Eq30parVmw1xT8swfZ2uMejS"

    # Construir os dados para envio à API de emissão de nota fiscal
    # Dados da Nota Fiscal
    nfe = {}
    itens = {}
    # Obtém a data e hora atual no fuso horário local (UTC)
    data_hora_utc = datetime.utcnow()

    # Calcula o deslocamento para o fuso horário brasileiro (UTC-3)
    deslocamento_horario = timedelta(hours=-3)

    # Adiciona o deslocamento ao horário UTC para obter a data e hora no fuso horário brasileiro
    data_hora_brasileira = data_hora_utc + deslocamento_horario

    # Formata a data e hora no formato ISO (string)
    data_hora_formatada_brasileira = data_hora_brasileira.isoformat()


    nfe["natureza_operacao"] = "Venda"
    nfe["forma_pagamento"] = "0"
    nfe["data_emissao"] = data_hora_formatada_brasileira
    nfe["tipo_documento"] = "1"
    if pedido.uf_entrega == 'RJ'  and pedido.cliente_pedido.estado == 'RJ':
        nfe["local_destino"] = "1"
    else:
        nfe["local_destino"] = "2"
    nfe["finalidade_emissao"] = "1"
    nfe["consumidor_final"] = "1"
    nfe["presenca_comprador"] = "0"
    nfe["cnpj_emitente"] = pedido.empresa_pedido.cnpj
    nfe["logradouro_emitente"] = pedido.empresa_pedido.endereco
    nfe["numero_emitente"] = pedido.empresa_pedido.endereco_numero
    nfe["bairro_emitente"] = pedido.empresa_pedido.bairro.nome
    nfe["municipio_emitente"] = pedido.empresa_pedido.cidade.nome
    nfe["uf_emitente"] = "RJ"
    nfe["cep_emitente"] = pedido.empresa_pedido.cep
    nfe["telefone_emitente"] = pedido.empresa_pedido.telefone
    nfe["inscricao_estadual_emitente"] = pedido.empresa_pedido.inscricao_estadual
    nfe["regime_tributario_emitente"] = "1"
    if len(pedido.cliente_pedido.documento) >= 12:
        nfe["cnpj_destinatario"] = pedido.cliente_pedido.documento
    else:
        nfe["cpf_destinatario"] = pedido.cliente_pedido.documento
    nfe["nome_destinatario"] = pedido.cliente_pedido.nome
    nfe["logradouro_destinatario"] = pedido.cliente_pedido.endereco
    nfe["numero_destinatario"] = pedido.cliente_pedido.numero_endereco
    nfe["bairro_destinatario"] = pedido.cliente_pedido.bairro
    nfe["municipio_destinatario"] = pedido.cliente_pedido.cidade
    nfe["uf_destinatario"] = pedido.cliente_pedido.estado
    nfe["cep_destinatario"] = pedido.cliente_pedido.CEP
    nfe["complemento_destinatario"] = pedido.cliente_pedido.complemento
    if(data.get('cep_entrega') != pedido.cliente_pedido.CEP):
        nfe["logradouro_entrega"]= data.get('logradouro_entrega')
        nfe["numero_entrega"] = data.get('numero_entrega')
        nfe["complemento_entrega"] = data.get('complemento_entrega')
        nfe["bairro_entrega"] = data.get('bairro_entrega')
        nfe["municipio_entrega"] = data.get('cidade_entrega')
        nfe["uf_entrega"] = data.get('uf_entrega')
        nfe["cep_entrega"] = data.get('cep_entrega')

    if (pedido.cliente_pedido.nome_contato):
        if (pedido.cliente_pedido.inscricao_estadual == 'ISENTO'):
            nfe["indicador_inscricao_estadual_destinatario"] = "2"

        elif (pedido.cliente_pedido.inscricao_estadual):
            nfe["indicador_inscricao_estadual_destinatario"] = "1"
            nfe["inscricao_estadual_destinatario"]= pedido.cliente_pedido.inscricao_estadual

    else:
        nfe["indicador_inscricao_estadual_destinatario"] = "2"
    nfe["icms_base_calculo"] = "0"
    nfe["icms_valor_total"] = "0"
    nfe["icms_valor_total_desonerado"] = "0"
    nfe["icms_base_calculo_st"] = "0"
    nfe["icms_valor_total_st"] = "0"
    nfe["valor_produtos"] = valorTotal
    nfe["valor_frete"] = "0"
    nfe["valor_seguro"] = "0"
    nfe["valor_desconto"] = "0"
    nfe["valor_total_ii"] = "0"
    nfe["valor_ipi"] = "0"
    nfe["valor_pis"] = "0"
    nfe["valor_cofins"] = "0"
    nfe["valor_outras_despesas"] = "0"
    nfe["valor_total"] = valorTotal
    nfe["modalidade_frete"] = "0"

    if data.get('transportadora'):
        nfe["volumes"] = []
        volumes = {}
        volumes["quantidade"] = data.get('volume')
        volumes["peso_bruto"] = data.get('peso')
        volumes["peso_liquido"] = data.get('peso')
        nfe["volumes"].append(volumes)
        nfe["nome_transportador"] = data.get('transportadora')
        if data.get('transportadora_cnpj'):
            nfe["cnpj_transportador"] = data.get('transportadora_cnpj')
    if informacoes_adicionais:
        nfe["informacoes_adicionais_contribuinte"] = informacoes_adicionais
    else:
        nfe["informacoes_adicionais_contribuinte"] = None


    numero_item = 1
    nfe["items"] = []  # Inicializa a lista de itens
    for item in items:

        itens = {}
        itens["numero_item"] = numero_item
        itens["codigo_produto"] = item['codigo']
        itens["descricao"] = item['nome']
        if pedido.uf_entrega == 'RJ' and pedido.cliente_pedido.estado == 'RJ':
            itens["cfop"] = "5101"
        else:
            itens["cfop"] = "6101"
        itens["unidade_comercial"] = "UN"
        itens["quantidade_comercial"] = item['quantidade']
        itens["valor_unitario_comercial"] = item['preco']
        itens["valor_bruto"] = item['total_item']
        itens["valor_desconto"] = 0.00
        itens["unidade_tributavel"] = "UN"
        itens["codigo_ncm"] = item['ncm']
        itens["quantidade_tributavel"] = item['quantidade']
        itens["valor_unitario_tributavel"] = item['preco']
        itens["inclui_no_total"] = "1"
        itens["icms_origem"] = "0"
        itens["icms_situacao_tributaria"] = "400"
        itens["pis_situacao_tributaria"] = "99"
        itens["cofins_situacao_tributaria"] = "99"
        nfe["items"].append(itens)
        numero_item += 1







    url = "https://api.focusnfe.com.br/v2/nfe"
    ref = {

        "ref": str(pedido_numero) + str(pedido.operacao_nota),
        "token": token
    }
    print(url,nfe)
    ref = urllib.parse.urlencode(ref)



    r = requests.post(url, params=ref, data=json.dumps(nfe), auth=(token, ""))
    print(r.status_code, r.text)
    response = r.json()

    if 'status' in response:
        status = response['status']

        if r.status_code == 202:
            while status == 'processando_autorizacao':
                json_response = r.json()
                ref = json_response['ref']
                pedido.referencia_nfe = ref

                url_consulta = 'https://api.focusnfe.com.br/v2/nfe/'
                completa = 'completa=1'
                response = requests.get(url_consulta+ref, params=completa, auth=(token, "") )
                print(response.status_code, response.text)
                resposta = response.json()
                status = resposta['status']

        if response.status_code == 200:
            json_response = response.json()
            xml = json_response['caminho_xml_nota_fiscal']
            nfe = json_response['caminho_danfe']
            nfe_numero = json_response['numero']
            url_focus = 'https://api.focusnfe.com.br'
            pedido.caminho_xml = url_focus+xml
            pedido.caminho_nfe = url_focus+nfe
            pedido.numero_nfe = nfe_numero



            pedido.save()
            response_data = {
                'nfe': pedido.caminho_nfe,
                'xml': pedido.caminho_xml
            }
            return JsonResponse(response_data)
        else:

            json_response = response.json()
            error_code = json_response['codigo']
            error_message = json_response['mensagem']

            response_data = {
                'code': error_code,
                'error': error_message
            }

            return JsonResponse(response_data, status=response.status_code)
    else:

        error_code = response['codigo']
        error_message = response['mensagem']

        response_data = {
            'code': error_code,
            'error': error_message
        }
        return JsonResponse(response_data)



    # Mostra na tela o código HTTP da requisição e a mensagem de retorno da API

def nota_aberta2 (request):
    data = json.loads(request.body)
    items = data['itens']
    print(items)
    total = 0

    # Itere pelos itens e some os valores da chave 'total_item'
    for item in items:

        total += float(item['total_item'])

    # Agora, 'total' conterá a soma dos 'total_item' de todos os itens
    print(total)

    return JsonResponse(True)




    # Mostra na tela o código HTTP da requisição e a mensagem de retorno da API


def calcular_mes(data):
    meses = [
        'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
        'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
    ]
    return meses[data.month - 1]
def report_comissoes_xls (request):
    wb = Workbook()
    colaboradores_ativos = Colaborador.objects.filter(status__descricao='ATIVO')
    # Recupere todos os vendedores
    vendedores = colaboradores_ativos.filter(Q(funcao__icontains='Vendedor')|Q(funcao__icontains='Gerente'))
    currency_style = NamedStyle(name='currency')
    currency_style.number_format = '"R$ "#,##0.00_);[Red]("R$ "#,##0.00)'

    # Para cada vendedor, crie uma nova aba
    for vendedor in vendedores:
        # Crie uma aba com o nome do vendedor
        ws = wb.create_sheet(title=vendedor.nome)


        # Defina os títulos das colunas na aba
        ws['A1'] = 'Mês da Venda'
        ws['B1'] = 'Número'
        #ws['C1'] = 'Nome do Cliente'
        ws['C1'] = 'Valor Total do Pedido'
        #ws['E1'] = 'Valor de Repasse'
        #ws['F1'] = 'Valor Base'
        ws['D1'] = 'Comissão (%)'
        ws['E1'] = 'Valor de Comissão'

        # Consulta ao banco de dados para obter os pedidos do vendedor
        print ('Nome do Vendedor', vendedor.nome)

        if vendedor.funcao == 'Vendedor Interno' or 'Marcia' in vendedor.nome:
            pedidos_vendedor = Pedidos.objects.filter(Vendedor=vendedor.nome)





            # Preencha as linhas com os dados dos pedidos
            for row, pedido in enumerate(pedidos_vendedor, start=2):
                mes = calcular_mes(pedido.datapedido)

                ws.cell(row=row, column=1, value=mes)
                ws.cell(row=row, column=2, value=str(pedido.numero))
                #ws.cell(row=row, column=3, value=pedido.cliente_pedido.nome)
                ws.cell(row=row, column=3, value=pedido.valorTotal).style = currency_style
                # Calcule os valores de repasse e comissão conforme necessário
                #ws.cell(row=row, column=5, value=pedido.valor_repasse).style = currency_style
                #valor_base = pedido.valorTotal - pedido.valor_repasse if pedido.valor_repasse else pedido.valorTotal
                #ws.cell(row=row, column=6, value=valor_base).style = currency_style
                ws.cell(row=row, column=4, value=pedido.vendedor_comissao)
                ws.cell(row=row, column=5, value=pedido.vendedor_comissaov).style = currency_style

        elif vendedor.funcao == 'Vendedor Externo' or 'RAUL' in vendedor.nome:

            pedidos_vendedor = Pedidos.objects.filter(Vendedor2=vendedor.nome)

            # Preencha as linhas com os dados dos pedidos
            for row, pedido in enumerate(pedidos_vendedor, start=2):
                mes = calcular_mes(pedido.datapedido)

                ws.cell(row=row, column=1, value=mes)
                ws.cell(row=row, column=2, value=str(pedido.numero))
                #ws.cell(row=row, column=3, value=pedido.cliente_pedido.nome)
                ws.cell(row=row, column=3, value=pedido.valorTotal).style = currency_style
                # Calcule os valores de repasse e comissão conforme necessário
                #ws.cell(row=row, column=5, value=pedido.valor_repasse).style = currency_style
                #valor_base = pedido.valorTotal - pedido.valor_repasse if pedido.valor_repasse else pedido.valorTotal
                #ws.cell(row=row, column=6, value=valor_base).style = currency_style
                ws.cell(row=row, column=4, value=pedido.vendedor2_comissao)
                ws.cell(row=row, column=5, value=pedido.vendedor2_comissaov).style = currency_style

    # Remova a aba padrão criada automaticamente
    wb.remove(wb.active)

    # Crie uma resposta HTTP para o arquivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relatorio_comissoes.xlsx'
    wb.save(response)

    return response

def report_comissoes_mensal_xls (request):
    wb = Workbook()
    data_inicial = request.POST.get('data_inicial')
    data_final = request.POST.get('data_final')
    print (data_inicial, data_final)

    if data_inicial and data_final:
        # Converte as strings de data para objetos datetime
        data_inicial = datetime.strptime(data_inicial, '%Y-%m-%d').date()
        data_final = datetime.strptime(data_final, '%Y-%m-%d').date()

    # Recupere todos os vendedores
    colaboradores_ativos = Colaborador.objects.filter(status__descricao='ATIVO')
    vendedores = colaboradores_ativos.filter(Q(funcao__icontains='Vendedor')|Q(funcao__icontains='Gerente'))
    currency_style = NamedStyle(name='currency')
    currency_style.number_format = '"R$ "#,##0.00_);[Red]("R$ "#,##0.00)'

    # Para cada vendedor, crie uma nova aba
    for vendedor in vendedores:
        # Crie uma aba com o nome do vendedor
        ws = wb.create_sheet(title=vendedor.nome)


        # Defina os títulos das colunas na aba
        ws['A1'] = 'Mês da Venda'
        ws['B1'] = 'Número'
        #ws['C1'] = 'Nome do Cliente'
        ws['C1'] = 'Valor Total do Pedido'
        #ws['E1'] = 'Valor de Repasse'
        #ws['F1'] = 'Valor Base'
        ws['D1'] = 'Comissão (%)'
        ws['E1'] = 'Valor de Comissão'

        # Consulta ao banco de dados para obter os pedidos do vendedor
        print ('Nome do Vendedor', vendedor.nome)

        if vendedor.funcao == 'Vendedor Interno' or 'Marcia' in vendedor.nome:
            pedidos_vendedor = Pedidos.objects.filter(Vendedor=vendedor.nome, datapedido__range=(data_inicial, data_final))





            # Preencha as linhas com os dados dos pedidos
            for row, pedido in enumerate(pedidos_vendedor, start=2):
                mes = calcular_mes(pedido.datapedido)

                ws.cell(row=row, column=1, value=mes)
                ws.cell(row=row, column=2, value=str(pedido.numero))
                #ws.cell(row=row, column=3, value=pedido.cliente_pedido.nome)
                ws.cell(row=row, column=3, value=pedido.valorTotal).style = currency_style
                # Calcule os valores de repasse e comissão conforme necessário
                #ws.cell(row=row, column=5, value=pedido.valor_repasse).style = currency_style
                #valor_base = pedido.valorTotal - pedido.valor_repasse if pedido.valor_repasse else pedido.valorTotal
                #ws.cell(row=row, column=6, value=valor_base).style = currency_style
                ws.cell(row=row, column=4, value=pedido.vendedor_comissao)
                ws.cell(row=row, column=5, value=pedido.vendedor_comissaov).style = currency_style

        elif vendedor.funcao == 'Vendedor Externo' or 'RAUL' in vendedor.nome:

            pedidos_vendedor = Pedidos.objects.filter(Vendedor2=vendedor.nome,  datapedido__range=(data_inicial, data_final))

            # Preencha as linhas com os dados dos pedidos
            for row, pedido in enumerate(pedidos_vendedor, start=2):
                mes = calcular_mes(pedido.datapedido)

                ws.cell(row=row, column=1, value=mes)
                ws.cell(row=row, column=2, value=str(pedido.numero))
                #ws.cell(row=row, column=3, value=pedido.cliente_pedido.nome)
                ws.cell(row=row, column=3, value=pedido.valorTotal).style = currency_style
                # Calcule os valores de repasse e comissão conforme necessário
                #ws.cell(row=row, column=5, value=pedido.valor_repasse).style = currency_style
                #valor_base = pedido.valorTotal - pedido.valor_repasse if pedido.valor_repasse else pedido.valorTotal
                #ws.cell(row=row, column=6, value=valor_base).style = currency_style
                ws.cell(row=row, column=4, value=pedido.vendedor2_comissao)
                ws.cell(row=row, column=5, value=pedido.vendedor2_comissaov).style = currency_style

    # Remova a aba padrão criada automaticamente
    wb.remove(wb.active)

    # Crie uma resposta HTTP para o arquivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relatorio_comissoes.xlsx'
    wb.save(response)

    return response

@login_required
def report_pedidos_xls (request):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()



    wb = Workbook()

    # Recupere todos os vendedores
    pedidos = Pedidos.objects.all()
    currency_style = NamedStyle(name='currency')
    currency_style.number_format = '"R$ "#,##0.00_);[Red]("R$ "#,##0.00)'

    # Para cada vendedor, crie uma nova aba

        # Crie uma aba com o nome do vendedor
    ws = wb.create_sheet(title='Pedidos')


    # Defina os títulos das colunas na aba
    ws['A1'] = 'Numero do Pedido'
    ws['B1'] = 'Nome do Cliente'
    ws['C1'] = 'Data do Pedido'
    ws['D1'] = 'Prazo de Entrega'
    ws['E1'] = 'Vendedor 1'
    ws['F1'] = '% Comis. 1'
    ws['G1'] = 'Valor Comis.1'
    ws['H1'] = 'Vendedor 2'
    ws['I1'] = '% Comiss 2'
    ws['J1'] = 'Valor Comis.1'
    ws['K1'] = 'Valor Total'
    ws['L1'] = 'Bonificado'
    ws['M1'] = 'Valor Repasse'
    ws['N1'] = 'Empresa'


    # Preencha as linhas com os dados dos pedidos
    for row, pedido in enumerate(pedidos, start=2):

        ws.cell(row=row, column=1, value=str(pedido.numero))
        ws.cell(row=row, column=2, value=str(pedido.cliente_pedido.nome))
        ws.cell(row=row, column=3, value=pedido.datapedido)
        ws.cell(row=row, column=4, value=pedido.prazo_entrega)
        ws.cell(row=row, column=5, value=pedido.Vendedor)
        ws.cell(row=row, column=6, value=pedido.vendedor_comissao)
        ws.cell(row=row, column=7, value=pedido.vendedor_comissaov).style = currency_style
        ws.cell(row=row, column=8, value=pedido.Vendedor2)
        ws.cell(row=row, column=9, value=pedido.vendedor2_comissao)
        ws.cell(row=row, column=10, value=pedido.vendedor2_comissaov).style = currency_style
        ws.cell(row=row, column=11, value=pedido.valorTotal).style = currency_style
        ws.cell(row=row, column=12, value=pedido.bonificado.nome) if pedido.bonificado else 'N'
        ws.cell(row=row, column=13, value=pedido.valor_repasse if pedido.valor_repasse else 0).style = currency_style
        ws.cell(row=row, column=14, value=pedido.empresa_pedido.nome)



    # Remova a aba padrão criada automaticamente
    wb.remove(wb.active)

    # Crie uma resposta HTTP para o arquivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relatorio_pedidos.xlsx'
    wb.save(response)
    return response

@login_required
def report_contas_receber_xls (request):

    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()



    wb = Workbook()

    # Recupere todos os vendedores
    contas = Contas_Receber.objects.all()
    currency_style = NamedStyle(name='currency')
    currency_style.number_format = '"R$ "#,##0.00_);[Red]("R$ "#,##0.00)'

    # Para cada vendedor, crie uma nova aba

        # Crie uma aba com o nome do vendedor
    ws = wb.create_sheet(title='Contas_Receber')


    # Defina os títulos das colunas na aba
    ws['A1'] = 'Numero do Pedido'
    ws['B1'] = 'Nome do Cliente'
    ws['C1'] = 'Data de Vencimento'
    ws['D1'] = 'Data de Pagamento'
    ws['E1'] = 'Forma de Pagamento'
    ws['F1'] = 'Valor da Parcela'
    ws['G1'] = 'Parcela'
    ws['H1'] = 'Total de Parcelas'
    ws['I1'] = 'Total do Pedido'
    ws['J1'] = 'Status'

    # Preencha as linhas com os dados dos pedidos
    for row, conta in enumerate(contas, start=2):



        ws.cell(row=row, column=1, value=str(conta.pedido))
        ws.cell(row=row, column=2, value=str(conta.cliente.nome))
        ws.cell(row=row, column=3, value=conta.data_vencimento)
        ws.cell(row=row, column=4, value=conta.data_pagamento)
        ws.cell(row=row, column=5, value=conta.tipo_pgto)
        # Calcule os valores de repasse e comissão conforme necessário
        ws.cell(row=row, column=6, value=conta.valor).style = currency_style
        ws.cell(row=row, column=7, value=conta.numero_parcela)
        ws.cell(row=row, column=8, value=conta.total_parcelas)
        if conta.pedido_conta:
            ws.cell(row=row, column=9, value=conta.pedido_conta.valorTotal).style = currency_style
        else:
            ws.cell(row=row, column=9, value=0).style = currency_style

        ws.cell(row=row, column=10, value=conta.status_conta)



    # Remova a aba padrão criada automaticamente
    wb.remove(wb.active)

    # Crie uma resposta HTTP para o arquivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relatorio_contas_receber.xlsx'
    wb.save(response)
    return response

@login_required
def report_contas_pagar_xls (request):

    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()




    wb = Workbook()

    # Recupere todos os vendedores
    contas = Contas_Pagar.objects.all()
    currency_style = NamedStyle(name='currency')
    currency_style.number_format = '"R$ "#,##0.00_);[Red]("R$ "#,##0.00)'

    # Para cada vendedor, crie uma nova aba

        # Crie uma aba com o nome do vendedor
    ws = wb.create_sheet(title='Contas_Pagar')


    # Defina os títulos das colunas na aba
    ws['A1'] = 'Numero do Pedido'
    ws['B1'] = 'Descricao'
    ws['C1'] = 'Data de Vencimento'
    ws['D1'] = 'Data de Pagamento'
    ws['E1'] = 'Forma de Pagamento'
    ws['F1'] = 'Valor da Parcela'
    ws['G1'] = 'Parcela'
    ws['H1'] = 'Total de Parcelas'
    ws['I1'] = 'Nota Fiscal'
    ws['J1'] = 'Status'
    ws['K1'] = 'Banco'
    ws['L1'] = 'Empresa'
    ws['M1'] = 'Fornecedor'


    # Preencha as linhas com os dados dos pedidos
    for row, conta in enumerate(contas, start=2):

        ws.cell(row=row, column=1, value=str(conta.pedido))
        ws.cell(row=row, column=2, value=str(conta.descricao))
        ws.cell(row=row, column=3, value=conta.data_vencimento)
        ws.cell(row=row, column=4, value=conta.data_pagamento)
        ws.cell(row=row, column=5, value=conta.tipo_pgto.descricao) if conta.tipo_pgto else 'N'
        # Calcule os valores de repasse e comissão conforme necessário
        ws.cell(row=row, column=6, value=conta.valor).style = currency_style

        ws.cell(row=row, column=7, value=conta.numero_parcela)
        ws.cell(row=row, column=8, value=conta.total_parcelas)
        ws.cell(row=row, column=9, value=conta.notafiscal)
        ws.cell(row=row, column=10, value=conta.status)
        ws.cell(row=row, column=11, value=conta.banco.nome) if conta.banco else 'N'
        ws.cell(row=row, column=12, value=conta.empresa.nome) if conta.empresa else 'N'
        ws.cell(row=row, column=13, value=conta.fornecedor.nome) if conta.fornecedor else 'N'



    # Remova a aba padrão criada automaticamente
    wb.remove(wb.active)

    # Crie uma resposta HTTP para o arquivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relatorio_contas_pagar.xlsx'
    wb.save(response)
    return response


@login_required
def painel_boletos(request, empresa):

    empresa=Empresas.objects.filter(nome__icontains=empresa).first()
    empresa_id = empresa.id
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    data_atual = date.today()
    mes_corrente = data_atual.month
    ano_corrente = data_atual.year
    if colaborador.funcao == 'Gerente':
        boletos = Contas_Receber.objects.filter\
            (tipo_pgto='BOLETO',
             id_cobranca__isnull=False,
             pedido_conta__empresa_pedido = empresa_id,
             data_vencimento__month=mes_corrente,
             data_vencimento__year=ano_corrente
             )
        tboletos = boletos.count()
        vboletos = 0.00
        for boleto in boletos:
            vboletos += boleto.valor

        vboletos = format_currency(float(vboletos),'BRL', locale='pt_BR')
        print(vboletos)
        boletos_recebidos = boletos.filter(status_cobranca='RECEIVED')
        tboletos_recebidos = boletos_recebidos.count()
        vboletos_recebidos = 0.00
        for boleto in boletos_recebidos:
            vboletos_recebidos += boleto.valor
        vboletos_recebidos = format_currency(float(vboletos_recebidos),'BRL', locale='pt_BR')
        boletos_atrasados = boletos.filter(status_cobranca='OVERDUE')
        tboletos_atrasados = boletos_atrasados.count()
        vboletos_atrasados = 0.00
        for boleto in boletos_atrasados:
            vboletos_atrasados += boleto.valor

        vboletos_atrasados = format_currency(float(vboletos_atrasados), 'BRL', locale='pt_BR')

        boletos_vencer = boletos.filter(status_cobranca='PENDING')
        tboletos_vencer = boletos_vencer.count()
        vboletos_vencer = 0.00
        for boleto in boletos_vencer:
            vboletos_vencer += boleto.valor
        vboletos_vencer = format_currency(float(vboletos_vencer), 'BRL', locale='pt_BR')

        context = {

            'boletos': boletos,
            'tboletos': tboletos,
            'vboletos': vboletos,
            'boletos_recebidos': boletos_recebidos,
            'tboletos_recebidos': tboletos_recebidos,
            'vboletos_recebidos': vboletos_recebidos,
            'boletos_atrasados': boletos_atrasados,
            'tboletos_atrasados': tboletos_atrasados,
            'vboletos_atrasados': vboletos_atrasados,
            'boletos_vencer': boletos_vencer,
            'tboletos_vencer': tboletos_vencer,
            'vboletos_vencer': vboletos_vencer,
            'nome': nome_usuario,
            'empresa': empresa.nome


        }

        print (tboletos, tboletos_recebidos, tboletos_atrasados, tboletos_vencer)

        return render(request, 'painel_boletos.html', context)

    else:
        return HttpResponse('USUARIO SEM AUTORIZAÇÃO')
@login_required
def boletos_total(request, empresa, status):

    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    empresa = Empresas.objects.filter(nome__icontains=empresa).first()
    empresa_id = empresa.id
    data_atual = date.today()
    mes_corrente = data_atual.month
    ano_corrente = data_atual.year
    if colaborador.funcao == 'Gerente':
        boletos = Contas_Receber.objects.filter\
            (tipo_pgto='BOLETO',
             id_cobranca__isnull=False,
             pedido_conta__empresa_pedido=empresa_id,
             data_vencimento__month=mes_corrente,
             data_vencimento__year=ano_corrente
             )
        if status == 'total':
            boletos = boletos
            tboletos = boletos.count()
        elif status =='vencer':
            boletos_vencer = boletos.filter(status_cobranca='PENDING')
            tboletos_vencer = boletos_vencer.count()
            boletos = boletos_vencer
        elif status == 'recebidos':
            boletos_recebidos = boletos.filter(status_cobranca='RECEIVED')
            tboletos_recebidos = boletos_recebidos.count()
            boletos = boletos_recebidos
        elif status == 'vencidos':
            boletos_atrasados = boletos.filter(status_cobranca='OVERDUE')
            tboletos_atrasados = boletos_atrasados.count()
            boletos = boletos_atrasados
        else:
            return HttpResponse('STATUS INVALIDO')
        total_boletos_lista = 0
        for boleto in boletos:
            total_boletos_lista += boleto.valor

        context = {

            'boletos': boletos.order_by('data_vencimento'),
            'nome': nome_usuario,
            'empresa': empresa.nome,
            'total_boletos_lista': total_boletos_lista

        }

        return render(request, 'boletos_total.html', context)


    else:
        return HttpResponse('USUARIO SEM AUTORIZAÇÃO')

@login_required
def filtrar_boletos(request):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()

    # Obter os dados da requisição POST
    data = json.loads(request.body)
    empresa_nome= data['empresa']
    print (empresa_nome)
    empresa = Empresas.objects.filter(nome__icontains = empresa_nome).first()
    empresa_id = empresa.id


    # Filtrar os boletos com base nos critérios
    boletos = Contas_Receber.objects.filter(tipo_pgto='BOLETO',
             id_cobranca__isnull=False,
             pedido_conta__empresa_pedido = empresa_id,
             )
    filtro_q = Q()
    if data['pedido']:
        filtro_q &= Q(pedido__icontains=data['pedido'])
    if data['cliente_nome']:
        filtro_q &= Q(cliente__nome__icontains=data['cliente_nome'])
    if data['cliente_documento']:
        filtro_q &= Q(cliente__documento__icontains=data['cliente_documento'])

    if data['idatav'] and data['fdatav']:
        filtro_q &= Q(data_vencimento__range=[data['idatav'], data['fdatav']])
    if data['idatap'] and data['fdatap']:
        filtro_q &= Q(data_pagamento__range=[data['idatap'], data['fdatap']])

    boletos = boletos.filter(filtro_q)
    print (boletos)

    # Criar uma lista de dicionários com os dados dos boletos
    boletos_data = []
    total_boletos_lista = 0
    for boleto in boletos:
        total_boletos_lista += boleto.valor
        boleto_data = {
            'numero': boleto.pedido,
            'cliente_pedido': boleto.cliente.nome,
            'datav': boleto.data_vencimento.strftime('%d/%m/%Y'),
            'datap': boleto.data_pagamento.strftime('%d/%m/%Y') if boleto.data_pagamento else " ",
            'valor': format_currency(boleto.valor, 'BRL', locale='pt_BR'),
            'numero_parcela': boleto.numero_parcela,
            'total_parcelas': boleto.total_parcelas,
            'status': boleto.status_conta,
            'total_boletos_lista': format_currency(total_boletos_lista, 'BRL', locale='pt_BR')

        }
        boletos_data.append(boleto_data)

    # Retornar os dados como resposta JSON
    return JsonResponse({'boletos': boletos_data})
@login_required
def painel_notas(request, empresa):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()

    empresa=Empresas.objects.filter(nome__icontains=empresa).first()
    empresa_id = empresa.id
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    data_atual = date.today()
    mes_corrente = data_atual.month
    ano_corrente = data_atual.year
    if data_atual.month == 1:
        mes_anterior = 12  # Se o mês atual é janeiro, o mês anterior é dezembro (mês 12)
    else:
        mes_anterior = data_atual.month - 1
    if colaborador.funcao == 'Gerente':
        notas = Pedidos.objects.filter( numero_nfe__isnull=False, empresa_pedido = empresa_id, dataEmissao_nfe__month = mes_corrente, dataEmissao_nfe__year = ano_corrente)
        tnotas = notas.count()
        vnotas = 0.00
        for nota in notas:
            vnotas += nota.valor_nota

        vnotas = format_currency(float(vnotas),'BRL', locale='pt_BR')
        print(vnotas)

        notas_ma = Pedidos.objects.filter( numero_nfe__isnull=False, empresa_pedido = empresa_id, dataEmissao_nfe__month = mes_anterior, dataEmissao_nfe__year = ano_corrente)
        tnotas_ma = notas_ma.count()
        vnotas_ma = 0.00
        for nota in notas_ma:
            vnotas_ma += nota.valor_nota
        vnotas_ma = format_currency(float(vnotas_ma), 'BRL', locale='pt_BR')

        context = {

            'notas': notas,
            'tnotas': tnotas,
            'vnotas': vnotas,
            'notas_ma': notas_ma,
            'vnotas_ma': vnotas_ma,
            'tnotas_ma': tnotas_ma,
            'nome': nome_usuario,
            'empresa': empresa.nome


        }

        print (notas, tnotas, vnotas)

        return render(request, 'painel_notas.html', context)

    else:
        return HttpResponse('USUARIO SEM AUTORIZAÇÃO')
@login_required
def notas_total(request, empresa, periodo):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()


    empresa = Empresas.objects.filter(nome__icontains=empresa).first()
    empresa_id = empresa.id
    nome_usuario = request.user.username
    colaborador = Colaborador.objects.get(usuario__username=nome_usuario)
    data_atual = date.today()
    mes_corrente = data_atual.month
    ano_corrente = data_atual.year
    if data_atual.month == 1:
        mes_anterior = 12  # Se o mês atual é janeiro, o mês anterior é dezembro (mês 12)
    else:
        mes_anterior = data_atual.month - 1
    if colaborador.funcao == 'Gerente':
        if periodo == 'atual':
            notas = Pedidos.objects.filter(numero_nfe__isnull=False, empresa_pedido=empresa_id,
                                           dataEmissao_nfe__month=mes_corrente, dataEmissao_nfe__year=ano_corrente).order_by('dataEmissao_nfe')
            tnotas = notas.count()
            vnotas = 0.00
            for nota in notas:
                vnotas += nota.valor_nota

            vnotas = format_currency(float(vnotas), 'BRL', locale='pt_BR')
            print(vnotas)
        elif periodo == 'ma':
            notas_ma = Pedidos.objects.filter(numero_nfe__isnull=False, empresa_pedido=empresa_id,
                                              dataEmissao_nfe__month=mes_anterior, dataEmissao_nfe__year=ano_corrente).order_by('dataEmissao_nfe')
            tnotas_ma = notas_ma.count()
            vnotas_ma = 0.00
            for nota in notas_ma:
                vnotas_ma += nota.valor_nota
            vnotas_ma = format_currency(float(vnotas_ma), 'BRL', locale='pt_BR')
            notas = notas_ma
        context = {

            'notas': notas,
            'nome': nome_usuario,
            'empresa': empresa.nome

        }


        return render(request, 'notas_total.html', context)


    else:
        return HttpResponse('USUARIO SEM AUTORIZAÇÃO')
@login_required
def cobranca_unificada(request, documento):

    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()



    cliente = Cliente.objects.filter(documento=documento).first()
    contas = Contas_Receber.objects.filter (cliente=cliente)
    contas = contas.filter (Q(status_conta='Pendente Faturamento')|Q(status_conta='A Vencer', id_cobranca__isnull = True ))

    pedidos = Pedidos.objects.filter(cliente_pedido=cliente)

    context ={
        'contas': contas,
        'cliente': cliente,
        'pedidos': pedidos
    }
    return render (request, 'unificar_contas.html', context)


@login_required
def unificar_cobranca(request):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()


    if request.method == 'POST':
        try:
            data = json.loads(request.body)  # Carrega o JSON do corpo da requisição
            pedidos_marcados = data['pedidos_marcados']
            if pedidos_marcados:
                #Checar se tem pedidos com 1 parcela paga ou não

                for numero_pedido in pedidos_marcados:
                    contas_pedido = Contas_Receber.objects.filter(pedido=numero_pedido)
                    for conta in contas_pedido:
                        if conta.status_conta != 'Pendente Faturamento':
                            return JsonResponse({'message': 'Não Pode Unificar. Um ou mais pedidos com status invalido!'}, status=200)
                        else:
                            conta.status_conta = 'Unificando'
                            conta.save()
                            pass

                valor_unificado = 0

                for numero_pedido in pedidos_marcados:
                    pedido = Pedidos.objects.filter (numero = numero_pedido).first()
                    valor_unificado += pedido.valorTotal

                nova_conta = Contas_Receber()
                nova_conta.cliente = pedido.cliente_pedido
                nova_conta.valor = valor_unificado
                nova_conta.numero_parcela = 1
                nova_conta.total_parcelas  = 1
                nova_conta.pedido = pedido.numero
                nova_conta.descricao = f'BOLETO UNIFICADO ({", ".join(str(pedido) for pedido in pedidos_marcados)})'
                nova_conta.tipo_pgto= 'BOLETO'
                nova_conta.status_conta = 'Pendente Faturamento Unificado'
                nova_conta.momento_pagamento = 'FATURADO'
                nova_conta.flag_uni = 1
                nova_conta.pedido_conta = pedido

                nova_conta.save()

                for numero_pedido in pedidos_marcados:

                    contas_pedido = Contas_Receber.objects.filter(pedido=numero_pedido)
                    for conta in contas_pedido:
                        if conta.status_conta == 'Unificando':
                            conta.delete()
                        else:
                            pass



                    pedido = Pedidos.objects.filter(numero=numero_pedido).first()
                    nova_cob = Cob_Unificada()
                    nova_cob.cliente = pedido.cliente_pedido
                    nova_cob.pedido = pedido

                    nova_cob.conta = nova_conta
                    nova_cob.save()

                print(nova_conta)
                return JsonResponse({'message': 'Cobrança Unificada gerada!', 'pedido': nova_conta.pedido}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'message': 'Erro ao unificar os pedidos'}, status=400)

def upload_foto_vendedor (request):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()



    if request.method == 'POST':

        id_colaborador = request.POST.get('id_colaborador')
        if 'arquivo' not in request.FILES:
            return JsonResponse({'success': False, 'message': 'Nenhum arquivo enviado.'})
        arquivo = request.FILES.get('arquivo')
        empresaaws = Empresas.objects.filter(id=1).first()
        s3_client = boto3.client('s3', aws_access_key_id=empresaaws.aws_id,
                                 aws_secret_access_key=empresaaws.aws_chave)
        colaborador = Colaborador.objects.filter(id=id_colaborador).first()
        if colaborador:
            nome_arquivo = f"foto_{colaborador.nome}.png"
            s3_client.upload_fileobj(arquivo, 'storagesw', nome_arquivo)
            colaborador.caminho_foto = f"https://storagesw.s3.amazonaws.com/{nome_arquivo}"
            colaborador.save()
            return JsonResponse({'success': True, 'message': f'Foto {colaborador.nome} cadastrada com sucesso!'})
        else:
            return JsonResponse({'success': False, 'message': 'Erro ao cadastrar Foto. Colaborador Não Encontrado!'})






def template_fornecedor(request):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()


    fornecedores = OpcoesFornecedor.objects.all()
    context = {

        'fornecedores': fornecedores,
        'empresas': Empresas.objects.all()
    }

    if request.method == 'POST':
        try:
            fornecedor_id = request.POST.get('fornecedor')
            fornecedor = OpcoesFornecedor.objects.filter(id=fornecedor_id).first()
            descricao = request.POST.get('descricao')
            empresa_id = request.POST.get('empresa')
            empresa = Empresas.objects.filter(id=empresa_id).first()




            if 'arquivo' not in request.FILES:
                return JsonResponse({'success': False, 'message': 'Nenhum arquivo enviado.'})
            arquivo = request.FILES['arquivo'].file
            empresaaws=Empresas.objects.filter(id=1).first()
            s3_client = boto3.client('s3', aws_access_key_id=empresaaws.aws_id,
                                     aws_secret_access_key=empresaaws.aws_chave)

            nome_arquivo = f"{fornecedor.nome}{descricao}.png"

            s3_client.upload_fileobj(arquivo, 'storagesw', nome_arquivo)

            novo_template = TemplateFornecedor()
            novo_template.fornecedor = fornecedor
            novo_template.descricao = descricao
            novo_template.empresa = empresa
            novo_template.caminho_imagem = f"https://storagesw.s3.amazonaws.com/{nome_arquivo}"
            novo_template.save()

            return JsonResponse({'success': True, 'message': 'Template cadastrado com sucesso!'})


        except KeyError as e:

            return JsonResponse({'success': False, 'message': f'Erro: {str(e)}'})

        except Exception as e:

            return JsonResponse({'success': False, 'message': 'Erro ao cadastrar template.'})

    return render (request, 'cadastra_template_fornecedor.html', context)




def gerar_codigo_orcamento(empresa):

    empresa = Empresas.objects.filter(nome=empresa).first()
    orcamento = Orcamentos.objects.filter(empresa=empresa).order_by('-id').first()
    if 'GRF' in empresa.nome:
        if orcamento:
            codigo = orcamento.codigo + 1
        else:
            codigo = 2433001
    else:
        if orcamento:
            codigo = orcamento.codigo + 1
        else:
            codigo = 2494001

    return codigo

@login_required
def novo_orcamento(request):

    nome_usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=nome_usuario).first()


    if 'Vendedor' in colaborador.funcao:
        templates = TemplateFornecedor.objects.filter(empresa=colaborador.empresa)
    else:
        templates = TemplateFornecedor.objects.all()


    if request.method == "POST":
        novo_orcamento = Orcamentos()
        empresa_nome = request.POST.get('empresa')
        novo_orcamento.empresa = Empresas.objects.filter(nome=empresa_nome).first()
        novo_orcamento.vendedor = colaborador
        novo_lead = Leads()
        novo_lead.nome = request.POST.get('nome_cliente')
        novo_lead.email = request.POST.get('email_cliente')
        novo_lead.telefone = remover_caracteres_especiais(request.POST.get('cliente_telefone'))
        novo_lead.save()
        novo_orcamento.cliente_lead = novo_lead

        novo_orcamento.validade = int(request.POST.get('validade'))
        novo_orcamento.dataexpiracao =(date.today() + timedelta (novo_orcamento.validade))
        novo_orcamento.forma_pagamento = request.POST.get('forma_pagamento')
        novo_orcamento.codigo = gerar_codigo_orcamento(empresa_nome)
        novo_orcamento.datacriacao = date.today()
        novo_orcamento.texto_garantia= request.POST.get('texto_garantia')
        novo_orcamento.status = StatusOr.objects.filter(descricao='Criado').first()
        templates_id = request.POST.getlist('templates')
        print (templates_id)
        quantidade = request.POST.getlist('quantidade')
        descricao_item = request.POST.getlist('descricao_item')
        comprimento = request.POST.getlist('comprimento')
        largura = request.POST.getlist('largura')
        tipo_item = request.POST.getlist('tipo_item')
        preco_unitario = request.POST.getlist('preco_unitario')
        valor_total = request.POST.getlist('valor_total')

        endereco = request.POST.get('endereco')
        cidade = request.POST.get('cidade')
        bairro = request.POST.get('bairro')
        estado = request.POST.get('estado')

        valor_orcamento =0

        # Agora você pode iterar sobre os itens capturados e salvar no banco de dados ou realizar outra lógica
        for i in range(len(quantidade)):
            # Acessa cada campo individualmente para cada linha da tabela
            item_novo = ItensOrcamento()
            item_novo.quantidade = quantidade[i]
            item_novo.descricao = descricao_item[i]
            item_novo.comprimento = comprimento[i]
            item_novo.largura = largura[i]
            tipo_descricao = tipo_item[i]
            item_novo.tipo_medida = TipoMedida.objects.filter(descricao=tipo_descricao).first()
            item_novo.valor_item = float (preco_unitario[i])
            item_novo.valor_total_item = valor_total[i]
            valor_orcamento = valor_orcamento+ float(valor_total[i])
            item_novo.codigo_orcamento=novo_orcamento.codigo
            item_novo.save()
            # Exemplo de impressão dos dados para depuração
            print(f'Quantidade: {item_novo.quantidade}, Descrição: { item_novo.descricao}, Comprimento: {item_novo.comprimento}, Largura: {item_novo.largura}, '
                  f'Tipo: {item_novo.tipo_medida}, Valor Total: { item_novo.valor_item}')
            novo_orcamento.valor_total=valor_orcamento

            novo_orcamento.endereco_entrega = endereco
            novo_orcamento.estado_entrega = estado
            novo_orcamento.cidade_entrega = cidade
            novo_orcamento.bairro_entrega = bairro
            novo_orcamento.save()

            item_novo.orcamento=Orcamentos.objects.filter(codigo=item_novo.codigo_orcamento).first()
            item_novo.save()

        for template in templates_id:
            novo_orcamento_template = Orcamentos_Templates()
            if template:
                template_orcamento = TemplateFornecedor.objects.filter(id=template).first()
                if template_orcamento:
                    novo_orcamento_template.template=template_orcamento
                    novo_orcamento_template.orcamento = novo_orcamento
                    novo_orcamento_template.save()
                else:
                    pass

        if novo_orcamento.empresa.id == 1:

            return redirect(f'/template_orcamento/{novo_orcamento.codigo}/')
        else:
            return redirect(f'/template_orcamento_tem/{novo_orcamento.codigo}/')

    else:
        context = {

            'empresas': Empresas.objects.all(),
            'clientes': Cliente.objects.all(),
            'tipos_medidas': TipoMedida.objects.all(),
            'colaborador': colaborador,
            'templates': templates

        }
        return render (request, 'new_orcamento.html', context)
@login_required
def painel_orcamento (request):


    nome_usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=nome_usuario).first()

    status = StatusOr.objects.all()

    mes_atual = datetime.now().month
    ano_atual = datetime.now().year
    if colaborador.funcao == 'Gerente':

        orcamentos_mes = Orcamentos.objects.all()
        total_orcamentos_mes = orcamentos_mes.count()
        total_acima_mil = orcamentos_mes.filter(valor_total__gte=1000).count()
        data_atual = timezone.now()

        # Calcula a data que representa 7 dias atrás
        data_limite = data_atual - timedelta(days=7)
        sem_atualizacao = Orcamentos.objects.filter(data_ultima_atualizacao__lte=data_limite).exclude(status__descricao__in=['Confirmado', 'Expirado'])
        total_sem_atualizacao = sem_atualizacao.count()
        for orcamento in sem_atualizacao:
            orcamento.status = StatusOr.objects.filter(descricao='7 Dias').first()
            orcamento.save()



        context = {
            'total_orcamentos_mes': total_orcamentos_mes,
            'total_acima_mil': total_acima_mil,
            'total_sem_atualizacao': total_sem_atualizacao,
            'orcamentos': orcamentos_mes,
            'colaborador': colaborador,
            'nome_colaborador': colaborador.nome.upper(),
            'itens': ItensOrcamento.objects.all(),
            'status': status
        }
    else:
        orcamentos = Orcamentos.objects.filter(vendedor__nome=colaborador.nome).order_by('-id')
        orcamentos_mes = orcamentos.filter(datacriacao__month=mes_atual, datacriacao__year=ano_atual)
        total_orcamentos_mes = orcamentos_mes.count()
        total_acima_mil = orcamentos_mes.filter(valor_total__gte=1000).count()
        data_atual = timezone.now()

        # Calcula a data que representa 7 dias atrás
        data_limite = data_atual - timedelta(days=7)
        sem_atualizacao = orcamentos.filter(data_ultima_atualizacao__lte=data_limite).exclude(
            status__descricao__in=['Confirmado', 'Expirado'])
        total_sem_atualizacao = sem_atualizacao.count()
        for orcamento in sem_atualizacao:
            orcamento.status = StatusOr.objects.filter(descricao='7 Dias').first()
            orcamento.save()


        context = {

            'orcamentos': orcamentos,
            'total_orcamentos_mes': total_orcamentos_mes,
            'total_acima_mil': total_acima_mil,
            'total_sem_atualizacao': total_sem_atualizacao,
            'nome_colaborador': colaborador.nome.upper(),
            'colaborador': colaborador,
            'status': status,
            'itens': ItensOrcamento.objects.all()
        }

    return render (request, 'orcamentos/painel_orcamentos.html', context)

def edita_orcamento (request, orcamento_id):
    campos = [field.name for field in Orcamentos._meta.get_fields()]
    print(campos)
    orcamento = Orcamentos.objects.filter(id=orcamento_id).first()
    context = {
        'campos':campos,
        'orcamento':orcamento

    }

    return render (request, 'orcamentos/edita_orcamento.html', context)
@csrf_exempt
@login_required
def atualiza_status_orcamento(request):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()

    if request.method == 'POST':
        data = json.loads(request.body)
        print(data)
        codigo_orcamento = data.get('orcamento_codigo')
        novo_status_id = data.get('status')
        nome_usuario = request.user.username
        colaborador = Colaborador.objects.filter(usuario__username=nome_usuario).first()
        status = StatusOr.objects.filter(id=novo_status_id).first()
        if status:
            orcamento = Orcamentos.objects.get(codigo=codigo_orcamento)
            if orcamento:
                orcamento.status = status
                orcamento.save()
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
@login_required
def atualiza_valor_final(request):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()

    if request.method == 'POST':
        data = json.loads(request.body)
        print(data)
        codigo_orcamento = data.get('orcamento_codigo')
        novo_valor = data.get('novo_valor')
        novo_valor_final = re.sub(r'^R\$|\.', '', novo_valor).replace(',', '.')
        nome_usuario = request.user.username
        colaborador = Colaborador.objects.filter(usuario__username=nome_usuario).first()

        if float(novo_valor_final) > 0:
            orcamento = Orcamentos.objects.get(codigo=codigo_orcamento)
            if orcamento:
                orcamento.valor_com_desconto = float(novo_valor_final)
                orcamento.data_ultima_atualizacao = datetime.now()
                orcamento.save()

                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
        else:
            return JsonResponse({'status': 'error', 'message': 'Valor Não Pode Ser Menor ou Igual a Zero!!!'}, status=405)

def get_observacoes(request, orcamento_id):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()


    orcamento = Orcamentos.objects.filter(id=orcamento_id).first()
    observacoes = Orcamento_Comentarios.objects.filter(orcamento=orcamento).values('texto', 'data').order_by('-id')
    context ={
        'orcamento': orcamento,
        'comentarios':observacoes,
        'vendedor': orcamento.vendedor.nome
    }
    return render (request, 'orcamentos/comentarios.html', context)

def gerar_orcamento (request, numero_orcamento):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()

    orcamento = Orcamentos.objects.filter(codigo=numero_orcamento).first()
    templates = Orcamentos_Templates.objects.filter(orcamento=orcamento)
    itens_orcamento = ItensOrcamento.objects.filter(orcamento=orcamento)
    flag_kapazi = 0
    for item in itens_orcamento:
        if 'kapazi' in item.descricao or 'Kapazi' in item.descricao or 'KAPAZI' in item.descricao :
            print(item.descricao)
            flag_kapazi = 1



    context = {

        'numero_orcamento': numero_orcamento,
        'orcamento': orcamento,
        'valor_total': format_currency(orcamento.valor_total,'BRL', locale='pt_BR'),
        'valor_com_desconto': format_currency(orcamento.valor_com_desconto,'BRL', locale='pt_BR') if orcamento.valor_com_desconto > 0 else 0.00 ,
        'itens': ItensOrcamento.objects.filter(codigo_orcamento=numero_orcamento),
        'nome_vendedor': orcamento.vendedor.nome.upper(),
        'templates': templates,
        'flag_kapazi': flag_kapazi

    }


    return render (request, 'template_orcamento.html', context)

def gerar_orcamento_tem (request, numero_orcamento):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()


    #locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')  # Se estiver no Windows, pode ser 'Portuguese_Brazil.1252'

    # Obtendo a data atual
    hoje =datetime.now()

    # Formatando o dia, mês e ano
    dia = hoje.day
    mes = hoje.strftime('%B')  # Obtém o nome completo do mês
    ano = hoje.year

    orcamento = Orcamentos.objects.filter(codigo=numero_orcamento).first()

    templates = Orcamentos_Templates.objects.filter(orcamento=orcamento)
    itens_orcamento = ItensOrcamento.objects.filter(orcamento=orcamento)
    flag_kapazi = 0
    for item in itens_orcamento:
        if 'kapazi' in item.descricao or 'Kapazi' in item.descricao:
            print(item.descricao)
            flag_kapazi=1

    print(flag_kapazi)

    context = {
        'dia': dia,
        'mes_string': mes,
        'ano_atual': ano,
        'numero_orcamento': numero_orcamento,
        'orcamento': orcamento,
        'valor_total': format_currency(orcamento.valor_total,'BRL', locale='pt_BR'),
        'valor_com_desconto': format_currency(orcamento.valor_com_desconto, 'BRL',
                                              locale='pt_BR') if orcamento.valor_com_desconto > 0 else 0.00,
        'templates': templates,
        'itens': ItensOrcamento.objects.filter(codigo_orcamento=numero_orcamento),
        'nome_vendedor': orcamento.vendedor.nome.upper(),
        'flag_kapazi': flag_kapazi

    }


    return render (request, 'template_orcamento_tem.html', context)

@login_required
def orcamento_comentario(request):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()


    if request.method == 'POST':
        data = json.loads(request.body)
        print(data)
        codigo_orcamento = data.get('numero_orcamento')
        comentario = data.get('comentario')
        nome_usuario = request.user.username
        colaborador = Colaborador.objects.filter(usuario__username=nome_usuario).first()
        orcamento = Orcamentos.objects.filter(codigo=codigo_orcamento).first()

        novo_comentario = Orcamento_Comentarios(
            orcamento = orcamento,
            texto = comentario,
            colaborador = colaborador,
            data = date.today()

        )
        novo_comentario.save()
        orcamento.data_ultima_atualizacao = date.today()
        orcamento.status = StatusOr.objects.filter(descricao='Aguardando').first()
        orcamento.save()

        return JsonResponse({'status': 'success'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

@csrf_exempt
def orcamento_carimbo (request):
    if request.method == 'POST':
        data = json.loads(request.body)  # Obtém o valor do parâmetro 'data' da URL
        id_orcamento = data.get('idOrcamento')
        orcamento = Orcamentos.objects.get(id=id_orcamento)
        if orcamento.flag_carimbo == 0:
            orcamento.flag_carimbo = 1
            orcamento.save()
            return JsonResponse({'message': 'Carimbo Habilitado! '})
        elif orcamento.flag_carimbo == 1:
            orcamento.flag_carimbo = 0
            orcamento.save()
            return JsonResponse({'message': 'Carimbo Desabilitado! '})
        else:
            return JsonResponse({'error': 'Orçamento Não Encontrado!'}, status=405)
    else:
        return JsonResponse({'error': 'Erro na alteração de Uso do Carimbo!'}, status=405)


def delete_orcamento(request):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()


    if request.method == 'POST':
        orcamento_codigo = request.POST.get('orcamento_codigo')
        orcamento = Orcamentos.objects.filter(codigo=orcamento_codigo).first()
        itens_orcamentos = ItensOrcamento.objects.filter(orcamento=orcamento)
        comentarios_orcamentos = Orcamento_Comentarios.objects.filter(orcamento=orcamento)
        itens_template = Orcamentos_Templates.objects.filter(orcamento=orcamento)
        if itens_template:
            itens_template.delete()

        if comentarios_orcamentos:
            comentarios_orcamentos.delete()
        if itens_orcamentos:
            itens_orcamentos.delete()
        orcamento.delete()
        mensagem = f'Orçamento {orcamento.codigo} excluido'
        novo_log(123, colaborador, mensagem)
        return JsonResponse({'mensagem': 'Orçamento deletado com sucesso!'})

    else:

        return JsonResponse({'mensagem': 'Método inválido.'}, status=400)


def relatorio_log(request):

    logs = RegistrosLog.objects.all().order_by('data')
    context = {
        'logs': logs
    }
    return render (request, 'relatorios/relatorio_log.html', context)


@login_required
def report_orcamentos_xls (request):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()


    wb = Workbook()

    # Recupere todos os vendedores
    orcamentos = Orcamentos.objects.all()
    currency_style = NamedStyle(name='currency')
    currency_style.number_format = '"R$ "#,##0.00_);[Red]("R$ "#,##0.00)'

    # Para cada vendedor, crie uma nova aba

        # Crie uma aba com o nome do vendedor
    ws = wb.create_sheet(title='Pedidos')


    # Defina os títulos das colunas na aba
    ws['A1'] = 'Codigo do Orçamento'
    ws['B1'] = 'Validade'
    ws['C1'] = 'Data do Orçamento'
    ws['D1'] = 'Nome do Vendedor'
    ws['E1'] = 'Status'
    ws['F1'] = 'Valor'
    ws['G1'] = 'Nome do CLiente'
    ws['H1'] = 'Telefone do Cliente'
    ws['I1'] = 'Empresa'



    # Preencha as linhas com os dados dos pedidos
    for row, orcamento in enumerate(orcamentos, start=2):

        ws.cell(row=row, column=1, value=str(orcamento.codigo))
        ws.cell(row=row, column=2, value=str(orcamento.validade))
        ws.cell(row=row, column=3, value=orcamento.datacriacao)
        ws.cell(row=row, column=4, value=orcamento.vendedor.nome)
        ws.cell(row=row, column=5, value=orcamento.status.descricao)
        ws.cell(row=row, column=6, value=orcamento.valor_total).style = currency_style
        ws.cell(row=row, column=7, value=orcamento.cliente_lead.nome)
        ws.cell(row=row, column=8, value=orcamento.cliente_lead.telefone)
        ws.cell(row=row, column=9, value=orcamento.empresa.nome)


    # Remova a aba padrão criada automaticamente
    wb.remove(wb.active)

    # Crie uma resposta HTTP para o arquivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relatorio_orcamentos.xlsx'
    wb.save(response)
    return response

@login_required
def report_clientes_xls (request):
    usuario = request.user.username
    colaborador = Colaborador.objects.filter(usuario__username=usuario).first()

    if colaborador.funcao == 'Gerente':

        wb = Workbook()

        clientes = Cliente.objects.all()
        currency_style = NamedStyle(name='currency')
        currency_style.number_format = '"R$ "#,##0.00_);[Red]("R$ "#,##0.00)'




        ws = wb.create_sheet(title='Clientes')



        ws['A1'] = 'Nome do Cliente'
        ws['B1'] = 'Data Cadastro'
        ws['C1'] = 'Origem Cadastro'
        ws['D1'] = 'Email'
        ws['E1'] = 'Telefone'
        ws['F1'] = 'Cidade'
        ws['G1'] = 'Nome Contato'


        for row, cliente in enumerate(clientes, start=2):

            ws.cell(row=row, column=1, value=str(cliente.nome))
            ws.cell(row=row, column=2, value=cliente.data_cadastro)
            ws.cell(row=row, column=3, value=cliente.tipo_entrada.descricao if cliente.tipo_entrada else None)
            ws.cell(row=row, column=4, value=cliente.email)
            ws.cell(row=row, column=5, value=cliente.telefone1)
            ws.cell(row=row, column=6, value=f'{cliente.endereco},{cliente.bairro} - {cliente.cidade} - {cliente.estado}')
            ws.cell(row=row, column=7, value=cliente.nome_contato)



        wb.remove(wb.active)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=relatorio_clientes.xlsx'
        wb.save(response)
        return response
    else:

        return HttpResponse('Sem Permissão de Acesso')

@login_required
def resumo_origens_clientes(request):
    hoje = timezone.now()
    mes_atual = hoje.month
    ano_atual = hoje.year

    # Definir o mês anterior
    if mes_atual == 1:
        mes_anterior = 12
        ano_anterior = ano_atual - 1
    else:
        mes_anterior = mes_atual - 1
        ano_anterior = ano_atual


    pedidos_grf =Pedidos.objects.filter(empresa_pedido__id=1)
    pedidos_tem = Pedidos.objects.filter(empresa_pedido__id=2)
    # Todos os clientes e tipos de entrada (sem filtro) GRF
    clientes_por_tipo_grf = pedidos_grf.values(
        'cliente_pedido__tipo_entrada__descricao'  # Agrupa apenas por tipo de entrada
    ).annotate(
        total_clientes=Count('cliente_pedido__id', distinct=True),
        valorTotal=Sum('valorTotal')  # Conta o número de clientes para cada tipo de entrada
    ).order_by('cliente_pedido__tipo_entrada__descricao')

    # Clientes do mês atual
    clientes_mes_atual_grf =  pedidos_grf.filter(
        cliente_pedido__data_cadastro__year=ano_atual,
        cliente_pedido__data_cadastro__month=mes_atual
    ).annotate(
        mes_cadastro=TruncMonth('cliente_pedido__data_cadastro')
    ).values(
        'cliente_pedido__tipo_entrada__descricao', 'mes_cadastro'
    ).annotate(
        total_clientes=Count('cliente_pedido__id', distinct=True),
        valorTotal=Sum('valorTotal')
    ).order_by('cliente_pedido__tipo_entrada')

    # Clientes do mês anterior
    clientes_mes_anterior_grf = pedidos_grf.filter(
        cliente_pedido__data_cadastro__year=ano_anterior,
        cliente_pedido__data_cadastro__month=mes_anterior,

    ).annotate(
        mes_cadastro=TruncMonth('cliente_pedido__data_cadastro')
    ).values(
        'cliente_pedido__tipo_entrada__descricao', 'mes_cadastro'
    ).annotate(
        total_clientes=Count('cliente_pedido__id', distinct=True),
        valorTotal=Sum('valorTotal')
    ).order_by('cliente_pedido__tipo_entrada')

   ########### Clientes Tem
    clientes_por_tipo_tem = pedidos_tem.values(
        'cliente_pedido__tipo_entrada__descricao'  # Agrupa apenas por tipo de entrada
    ).annotate(
        total_clientes=Count('cliente_pedido__id', distinct=True),
        valorTotal=Sum('valorTotal')
        # Conta o número de clientes para cada tipo de entrada
    ).order_by('cliente_pedido__tipo_entrada__descricao')

    # Clientes do mês atual
    clientes_mes_atual_tem = pedidos_tem.filter(
        cliente_pedido__data_cadastro__year=ano_atual,
        cliente_pedido__data_cadastro__month=mes_atual
    ).annotate(
        mes_cadastro=TruncMonth('cliente_pedido__data_cadastro')
    ).values(
        'cliente_pedido__tipo_entrada__descricao', 'mes_cadastro'
    ).annotate(
        total_clientes=Count('cliente_pedido__id', distinct=True),
        valorTotal=Sum('valorTotal')
    ).order_by('cliente_pedido__tipo_entrada')

    # Clientes do mês anterior
    clientes_mes_anterior_tem = pedidos_tem.filter(
        cliente_pedido__data_cadastro__year=ano_anterior,
        cliente_pedido__data_cadastro__month=mes_anterior,

    ).annotate(
        mes_cadastro=TruncMonth('cliente_pedido__data_cadastro')
    ).values(
        'cliente_pedido__tipo_entrada__descricao', 'mes_cadastro',
    ).annotate(
        total_clientes=Count('cliente_pedido__id', distinct=True),
        valorTotal=Sum('valorTotal')
    ).order_by('cliente_pedido__tipo_entrada')

    # Recuperando todos os tipos de indicação
    tipos = Tipo_Indicacao.objects.all()



    # Preparando o contexto para a renderização da página
    context = {
        'clientes_por_tipo': clientes_por_tipo_grf,
        'clientes_mes_atual': clientes_mes_atual_grf,
        'clientes_mes_anterior': clientes_mes_anterior_grf,
        'clientes_por_tipo_tem': clientes_por_tipo_tem,
        'clientes_mes_atual_tem': clientes_mes_atual_tem,
        'clientes_mes_anterior_tem': clientes_mes_anterior_tem,
        'tipos': tipos
    }

    return render (request,'clientes/painel_crm.html', context)


@csrf_exempt
def webhook_notas_grf(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        print (data,'NOTA GRF')

@csrf_exempt
def webhook_notas_tem(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        print (data,'NOTA TEM')
